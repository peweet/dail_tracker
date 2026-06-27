"""LA Purchase-Orders / Payments over €20,000 — 31-council per-transaction fact.

The per-transaction COMMITTED/SPENT layer of the procurement lifecycle (see
doc/PROCUREMENT_INVESTIGATION.md "LA PO corpus" + doc/PROCUREMENT_BUILD_PLAN.md §4b/§8b).
Circular Fin 07/2012 obliges every local authority to publish its purchase-orders (or
payments) over €20k; each does so on its OWN website in a near-converged
`Supplier · Amount(€) · Description` shape. This builds them into ONE silver fact.

DESIGN — a council is a ROW in SCHEMA_MAP, never a new parser:
  - SCHEMA_MAP    one config per council (listing/direct URLs + value_kind + quirks)
  - ONE generic reader  fitz largest-x-gap for PDFs (order-independent, strips PO#/ID
                        prefixes); openpyxl header+content-fallback for XLSX; polars CSV
  - per-council QUIRKS  ≈1 flag each (neg-abs / supplier-is-ID / aggregate-guard / …)

SCHEMA — emits the SAME row shape as data/silver/parquet/public_payments_fact.parquet
(the 17 central/semi-state publishers) so the two facts UNION at gold-view time, but on
the MASTER taxonomy: `value_kind` (renamed from that fact's drifted `amount_semantics`)
+ `realisation_tier` + derived `value_safe_to_sum`. One fact, one vocab — option (a).

MEDALLION — raw files self-fetched to bronze/pdfs/la_procurement/<council>/; the
reconciled fact to silver/parquet/la_payments_fact.parquet (zstd/3/stats, gitignore-
negated → Cloud-readable). NOT wired into pipeline.py yet (promote on go-ahead, like afs).

PRIVACY ([[feedback_personal_insolvency_privacy]] / no-inference): sole-trader /
individual / bare-ID payees are personal data → supplier_class drives a quarantine
(public_display=False); company-suffix and public-body payees are kept.

Run:
  ./.venv/Scripts/python.exe extractors/procurement_la_payments_extract.py --list
  ./.venv/Scripts/python.exe extractors/procurement_la_payments_extract.py --only south_dublin,cork_county
  ./.venv/Scripts/python.exe extractors/procurement_la_payments_extract.py --max-files 6
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import io
import json
import re
import subprocess
import sys
import time
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import quote, unquote, urljoin, urlparse

import fitz  # PyMuPDF
import polars as pl
import requests

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8")

import config  # noqa: E402
from services.fetch_report import Breaker, FetchReport, classify_body, classify_exception, write_sentinel  # noqa: E402
from services.parquet_io import save_parquet  # noqa: E402
from shared.name_norm import name_norm_expr  # noqa: E402
from shared.text_encoding import decode_table_bytes  # noqa: E402

H = {"User-Agent": "Mozilla/5.0 (dail-tracker research)"}
BRONZE = config.BRONZE_PDF_DIR / "la_procurement"
OUT_FACT = config.SILVER_PARQUET_DIR / "la_payments_fact.parquet"
OUT_COV = ROOT / "data/_meta/la_payments_coverage.json"
CRO = ROOT / "data/silver/cro/companies.parquet"
PARSER_VERSION = "0.1.0"
MIN_EUR = 1000.0  # PO-over-20k lists carry the occasional sub-20k line; 1k drops noise/refs
# Row floor for the overwrite-each-run fact (84,706 rows 2026-06-20). The fact is
# replaced wholesale every run, so a partial `--only` slice or a mass harvest
# failure would silently downgrade silver and the next consolidate would wipe gold
# history (see the --max-files note in main). ~29% headroom below the full count;
# a partial/debug run is the intended use of DAIL_SKIP_ROW_FLOOR=1.
MIN_FACT_ROWS = 60_000

DATA_EXT = (".pdf", ".xlsx", ".xls", ".csv")

# ----------------------------------------------------------------------------- regexes
MONEY_RE = re.compile(r"(?:€|EUR)?\s?\d{1,3}(?:,\d{3})+(?:\.\d{2})?|\d+\.\d{2}")
NUM_RE = re.compile(r"-?\d[\d,]*(?:\.\d+)?")
HREF_RE = re.compile(r"""href\s*=\s*["']([^"']+)["']""", re.I)
DIGIT_PREFIX = re.compile(
    r"^\d{3,}(?:\s+\d+)?\s+"
)  # leading PO# + optional vendor-ID run (Mayo/Donegal/Waterford/Kilkenny)
QUARTER_RE = re.compile(r"q\s?([1-4])|quarter[\s_-]?([1-4])|qtr[\s_-]?([1-4])", re.I)

SUP_RE = re.compile(r"supplier|payee|vendor|provider|creditor|benefic|\bname\b|company", re.I)
AMT_RE = re.compile(r"amount|total|value|gross|\beuro\b|€|\beur\b|\bpaid\b|cost|net", re.I)
DESC_RE = re.compile(r"descript|\bdesc\b|detail|categor|service|goods|nature|product|\bgl\b", re.I)
PAID_RE = re.compile(r"\bpaid\b|payment status|\bstatus\b", re.I)
PO_RE = re.compile(r"\border\s*no|\bpo\b|\bpor\b|order number|\bref\b|transaction", re.I)
TOTAL_RE = re.compile(r"^\s*(grand\s+|sub-?)?total\b|^\s*all suppliers\b|^\s*various\b", re.I)

# supplier-class classifiers (mirror procurement_public_body_extract.py so the two facts
# resolve identically; keep in lockstep if either changes).
COMPANY_SUFFIX = re.compile(
    r"\b(ltd|limited|dac|plc|clg|uc|teo|teoranta|t/a|inc|llc|gmbh|company|co\.|group|"
    r"services|solutions|consult|engineer|partners|associates|holdings|university|college|"
    r"council|board|institute|ireland|technolog|systems|media|construction|contractors|&)\b",
    re.I,
)
FOREIGN_FORM = re.compile(
    r"\b(gmbh|s\.?a\.?|n\.?v\.?|s\.?a\.?s|s\.?p\.?a|inc|llc|\bpty\b|\bab\b|\bbv\b|\boy\b|srl|sl|sarl|aps)\b", re.I
)
PUBLIC_BODY = re.compile(
    r"\b(county council|city council|university|institute of technology|department of|office of|"
    r"\bHSE\b|health service|an garda|údarás|udaras|education and training board|\bETB\b|"
    r"local authority|\bOPW\b|hospital|\birish water\b|uisce)\b",
    re.I,
)
ID_CODE_RE = re.compile(r"^[\d\W]+$|^[A-Z]?\d{4,}$")  # Fingal: supplier published as a bare code


# ============================================================================ CONFIG
def la(
    slug,
    council,
    region,
    entity,
    *,
    fmt,
    listing,
    value_kind,
    direct=None,
    include=None,
    neg_abs=False,
    supplier_is_id=False,
    aggregate_guard=False,
    anchor_listing=False,
    status="READY",
    caveat="",
    pdf_reader=None,
) -> dict:
    """One local authority. value_kind ∈ {po_committed (→COMMITTED), payment_actual (→SPENT)}."""
    return {
        "slug": slug,
        "council": council,
        "region": region,
        "entity": entity,
        "fmt": fmt,
        "listing_url": listing,
        "value_kind": value_kind,
        "direct_files": direct or [],
        "include": re.compile(include, re.I) if include else None,
        "neg_abs": neg_abs,
        "supplier_is_id": supplier_is_id,
        "aggregate_guard": aggregate_guard,
        "anchor_listing": anchor_listing,
        "status": status,
        "caveat": caveat,
    }


# Limerick deletes each prior quarter's PDF on re-upload (live site keeps only the current
# quarter; all older URLs 404), so the back-catalogue exists ONLY in the Internet Archive.
# These are permanent Wayback snapshots (`web/{ts}id_/{original}` returns the raw PDF). Sourced
# via the CDX API (url=limerick.ie&filter=mimetype:application/pdf&filter=original:.*urchase.*rders.*);
# baked in as a deterministic seed so extraction has no live-CDX dependency (CDX 504s often).
# pre-2016 quarters are auto-dropped by harvest's <2016 gate (consistent with the rest of the corpus).
_WB = "https://web.archive.org/web/{}id_/{}"
LIMERICK_WAYBACK: list[str] = [
    _WB.format(ts, u)
    for ts, u in [
        (
            "20171025144903",
            "https://www.limerick.ie/sites/default/files/media/documents/2017-05/purchase_orders_over_20000_quarter_1_2016_new.pdf",
        ),
        (
            "20171025144900",
            "https://www.limerick.ie/sites/default/files/media/documents/2017-05/purchase_orders_over_20000_quarter_2_2016.pdf",
        ),
        (
            "20171025144855",
            "https://www.limerick.ie/sites/default/files/media/documents/2017-05/purchase_orders_over_20k_q4_2016.pdf",
        ),
        (
            "20171025144857",
            "https://www.limerick.ie/sites/default/files/media/documents/2017-05/purchase_orders_over_eu20000_-_quarter_3_2016.pdf",
        ),
        (
            "20171025144849",
            "https://www.limerick.ie/sites/default/files/media/documents/2017-08/Purchase%20Orders%20over%20%E2%82%AC20%2C000%20Quarter%202%2C%202017_0.pdf",
        ),
        (
            "20190526073200",
            "https://www.limerick.ie/sites/default/files/media/documents/2018-12/Purchase%20Orders%20over%20%E2%82%AC20000%20Q3%202018%20v2.pdf",
        ),
        (
            "20190526073154",
            "https://www.limerick.ie/sites/default/files/media/documents/2019-02/Purchase%20Orders%20over%20%E2%82%AC20000%20Quarter%204%202018.pdf",
        ),
        (
            "20200917233827",
            "https://www.limerick.ie/sites/default/files/media/documents/2019-09/Purchase-Orders-over-20000-Q2-2019.pdf",
        ),
        (
            "20200917233836",
            "https://www.limerick.ie/sites/default/files/media/documents/2019-09/Purchase-Orders-over-20000-Quarter1-2019.pdf",
        ),
        (
            "20200917233821",
            "https://www.limerick.ie/sites/default/files/media/documents/2020-01/Purchase-Orders-over-20000-Quarter-3-2019.pdf",
        ),
        (
            "20200917233817",
            "https://www.limerick.ie/sites/default/files/media/documents/2020-04/purchase-orders-over-20000-q4-2019.pdf",
        ),
        (
            "20210607110505",
            "https://www.limerick.ie/sites/default/files/media/documents/2020-10/purchase-orders-over-eu20000-quarter-2-2020.pdf",
        ),
        (
            "20210607110501",
            "https://www.limerick.ie/sites/default/files/media/documents/2021-01/purchase-orders-over-eu20000-quarter-3-2020.pdf",
        ),
        (
            "20210607110457",
            "https://www.limerick.ie/sites/default/files/media/documents/2021-04/purchase-orders-over-eu20000-q4-2020.pdf",
        ),
        (
            "20220613103544",
            "https://www.limerick.ie/sites/default/files/media/documents/2022-02/purchase-orders-over-eu20000-q1-2021.pdf",
        ),
        (
            "20220613103540",
            "https://www.limerick.ie/sites/default/files/media/documents/2022-02/purchase-orders-over-eu20000-q2-2021.pdf",
        ),
        (
            "20230710100906",
            "https://www.limerick.ie/sites/default/files/media/documents/2022-08/Purchase-Orders-over-20000-Quarter-3-2021.pdf",
        ),
        (
            "20230710100848",
            "https://www.limerick.ie/sites/default/files/media/documents/2022-08/Purchase-Orders-over-20000-Quarter-4-2021.pdf",
        ),
        (
            "20230710100831",
            "https://www.limerick.ie/sites/default/files/media/documents/2022-08/Purchase-Orders-over-20000-Quarter-1-2022.pdf",
        ),
        (
            "20230710100805",
            "https://www.limerick.ie/sites/default/files/media/documents/2023-01/Purchase-Orders-over-%E2%82%AC20%2C000-Quarter-2-2022.pdf",
        ),
        (
            "20230710100759",
            "https://www.limerick.ie/sites/default/files/media/documents/2023-01/Purchase-Orders-over-%E2%82%AC20%2C000-Quarter-3-2022.pdf",
        ),
        (
            "20230710100752",
            "https://www.limerick.ie/sites/default/files/media/documents/2023-02/Purchase-Orders-over-%E2%82%AC20%2C000-Quarter-4-2022.pdf",
        ),
        (
            "20230710100742",
            "https://www.limerick.ie/sites/default/files/media/documents/2023-06/Purchase-Orders-over-%E2%82%AC20%2C000-Quarter-1-2023.pdf",
        ),
    ]
]


# Routes merged from procurement_la_registry.py (best file-list / direct URLs). value_kind
# is set from the council's publication TYPE: "Purchase Orders over 20k" → po_committed
# (COMMITTED); "Payments greater than 20k" → payment_actual (SPENT).
SCHEMA_MAP: list[dict] = [
    # ---- tabular (XLSX / CSV) — no PDF parsing ----
    la(
        "south_dublin",
        "South Dublin",
        "Dublin",
        "dublin",
        fmt="xlsx",
        listing="https://www.sdcc.ie/en/services/business/payments/",
        value_kind="po_committed",
        caveat="PO list carries a PAID Y/N flag (carried per row)",
    ),
    la(
        "cork_city",
        "Cork City",
        "Munster",
        "city",
        fmt="xlsx",
        listing="https://www.corkcity.ie/en/council-services/public-info/spending-and-revenue/",
        value_kind="po_committed",
        caveat="xlsx; old media-folder PDFs 404",
    ),
    la(
        "wicklow",
        "Wicklow",
        "Leinster",
        "county",
        fmt="xlsx",
        listing="https://www.wicklow.ie/Living/Your-Council/Finance/Procurement/Purchase-Orders-Over-20-000",
        value_kind="po_committed",
    ),
    la(
        "monaghan",
        "Monaghan",
        "Ulster",
        "county",
        fmt="xlsx",
        listing="https://monaghan.ie/finance/publication-of-purchase-orders/",
        value_kind="po_committed",
    ),
    la(
        "kilkenny",
        "Kilkenny",
        "Leinster",
        "county",
        fmt="xlsx",
        listing="https://kilkennycoco.ie/eng/services/finance/purchase-orders-over-%E2%82%AC20-000/",
        value_kind="po_committed",
        neg_abs=True,
        caveat="amounts stored NEGATIVE (debit sign) → abs(); supplier col 'Ap/Ar ID(T)' resolved by content-fallback",
    ),
    la(
        "wexford",
        "Wexford",
        "Leinster",
        "county",
        fmt="xlsx",
        listing="https://www.wexfordcoco.ie/council-and-democracy/procurement-finance-and-credit-control/council-spend",
        value_kind="po_committed",
        caveat="recent quarters .xlsx; 2016-2024 history is legacy .xls (read via xlrd 2.x)",
    ),
    # ---- digital PDF (fitz largest-x-gap; NO OCR) ----
    la(
        "cork_county",
        "Cork County",
        "Munster",
        "county",
        fmt="pdf",
        listing="https://www.corkcoco.ie/en/council/accessibility-maps-and-publications/purchase-orders-in-excess-of-eu20000",
        value_kind="po_committed",
        direct=[
            "https://www.corkcoco.ie/sites/default/files/2025-08/2025-q2-purchase-orders-in-excess-of-eu20000.pdf",
            "https://www.corkcoco.ie/sites/default/files/2025-05/2025-q1-purchase-orders-in-excess-of-eu20000-pdf.pdf",
        ],
        caveat="row carries a Paid flag + free-text description; x-gap keeps supplier+amount",
    ),
    la(
        "kildare",
        "Kildare",
        "Leinster",
        "county",
        fmt="pdf",
        listing="https://kildarecoco.ie/YourCouncil/Publications/Finance/PurchaseOrdersover20000/",
        value_kind="po_committed",
    ),
    la(
        "westmeath",
        "Westmeath",
        "Leinster",
        "county",
        fmt="pdf",
        listing="https://www.westmeathcoco.ie/en/ourservices/finance/procurement/purchaseorders/",
        value_kind="po_committed",
    ),
    la(
        "waterford",
        "Waterford",
        "Munster",
        "merged",
        fmt="pdf",
        listing="https://waterfordcouncil.ie/openness-transparency/governance-related-financial-information/procurement/purchase-orders-e20000/",
        value_kind="po_committed",
        caveat="leading OrderNo column → digit-prefix strip",
    ),
    la(
        "limerick",
        "Limerick",
        "Munster",
        "merged",
        fmt="pdf",
        listing="https://www.limerick.ie/council/services/business-and-economy/revenue-collection/accounts-payable",
        value_kind="po_committed",
        direct=[
            "https://www.limerick.ie/sites/default/files/media/documents/2026-05/purchase-orders-over-eu20-000-quarter-1-2026.pdf",
            *LIMERICK_WAYBACK,
        ],
        caveat="row carries a Paid column. Live site keeps ONLY the current quarter (prunes the rest) "
        "→ 2016-2023 back-catalogue seeded from Internet Archive snapshots (LIMERICK_WAYBACK)",
    ),
    la(
        "offaly",
        "Offaly",
        "Leinster",
        "county",
        fmt="pdf",
        listing="https://www.offaly.ie/financial-reports/",
        value_kind="payment_actual",
        include=r"payment|20k|over.?20|greater",
        caveat="GL30 'Payments Greater than €20k' → SPENT grain",
    ),
    la(
        "longford",
        "Longford",
        "Leinster",
        "county",
        fmt="pdf",
        listing="https://www.longfordcoco.ie/services/finance/finance-documents/large-purchase-orders/",
        value_kind="po_committed",
    ),
    la(
        "galway_city",
        "Galway City",
        "Connacht",
        "city",
        fmt="pdf",
        listing="https://www.galwaycity.ie/services/finance-services/budgets-and-financial-publications",
        value_kind="po_committed",
        direct=[
            "https://www.galwaycity.ie/sites/default/files/2026-05/Qtr%201%202026_Purchase%20Orders%20over%20%E2%82%AC20k_0.pdf"
        ],
        caveat="amount is the LAST column (handled by rightmost-money pick); budgets page mixes prompt-pay",
    ),
    la(
        "galway_county",
        "Galway County",
        "Connacht",
        "county",
        fmt="pdf",
        listing="https://www.gaillimh.ie/en/finance/financial-publications/purchase-orders",
        value_kind="po_committed",
        direct=["https://www.galwaycoco.ie/sites/default/files/2026-01/Quarter%201%202025%20%28ENG%29.pdf"],
        caveat="gaillimh.ie alt domain unblocks the galwaycoco WAF listing",
    ),
    la(
        "kerry",
        "Kerry",
        "Munster",
        "county",
        fmt="pdf",
        listing="https://www.kerrycoco.ie/finance/financial-documents/",
        value_kind="po_committed",
    ),
    la(
        "meath",
        "Meath",
        "Leinster",
        "county",
        fmt="pdf",
        listing="https://www.meath.ie/council/your-council/finance-and-procurement/tenders-and-contracts/payments-over-eu20000",
        value_kind="payment_actual",
        caveat="'payments over €20,000' → SPENT; curl fallback (Python TLS quirk)",
    ),
    la(
        "sligo",
        "Sligo",
        "Connacht",
        "county",
        fmt="pdf",
        listing="https://www.sligococo.ie/YourCouncil/Finance/ProcurementPurchasing/PurchasingActivity/",
        value_kind="po_committed",
        caveat="curl fallback (Python TLS quirk)",
    ),
    # ---- one-hop crawl from the landing page ----
    la(
        "clare",
        "Clare",
        "Munster",
        "county",
        fmt="pdf",
        listing="https://www.clarecoco.ie/business-licensing-and-economy/procurement-and-tenders",
        value_kind="po_committed",
        direct=[
            "https://www.clarecoco.ie/sites/default/files/2025-08/purchase-orders-over-20-000-in-the-2nd-quarter-of-2025-58204.pdf"
        ],
    ),
    la(
        "leitrim",
        "Leitrim",
        "Connacht",
        "county",
        fmt="pdf",
        listing="https://www.leitrim.ie/council/services/finance/accounts-payable/purchase-to-pay/",
        value_kind="po_committed",
        aggregate_guard=True,
        caveat="finance page also lists a prompt-pay AGGREGATE return → aggregate_guard skips 1-row files",
    ),
    la(
        "laois",
        "Laois",
        "Leinster",
        "county",
        fmt="pdf",
        listing="https://laois.ie/finance/business-and-enterprise-support/procurement-information-and-advice",
        value_kind="po_committed",
        aggregate_guard=True,
        direct=[
            # files exist on the server but are UNLINKED from any laois.ie HTML page (orphaned);
            # the listing crawl finds nothing, so seed them directly. Upload-month folder + period
            # separator (–/-) are both irregular → not patternable. Verified HTTP 200, application/pdf.
            "https://laois.ie/sites/default/files/2024-06/Procurement%20Report%202023%20Oct%E2%80%93Dec.pdf",
            "https://laois.ie/sites/default/files/2024-06/Procurement Report 2023 Jul–Sept.pdf",
            "https://laois.ie/sites/default/files/2024-06/Procurement Report 2023 Apr–Jun.pdf",
            "https://laois.ie/sites/default/files/2024-10/Procurement Report 2022 Jan-Mar.pdf",
            "https://laois.ie/sites/default/files/2024-10/Procurement Report 2022 Jul-Sept.pdf",
            "https://laois.ie/sites/default/files/2024-10/Procurement Report 2022 Oct-Dec.pdf",
        ],
        caveat="'Procurement Report' grain; guard drops total-row mis-grabs (€397m artefact). "
        "Files orphaned (unlinked) → direct= seeds; cols Supplier|Description|Total",
    ),
    la(
        "fingal",
        "Fingal",
        "Dublin",
        "dublin",
        fmt="pdf",
        listing="https://www.fingal.ie/council/service/procurement",
        value_kind="po_committed",
        include=r"(?i)(pos?-over-20k|purchase.?orders?.?over.?20k)",
        direct=[
            # the procurement landing links ZERO PO files (only policy PDFs); a crawl drifts into
            # heritage docs. Files live at sites/default/files/{upload-month}/...; seed directly +
            # include= filters the listing noise. Verified HTTP 200, born-digital PDF.
            "https://www.fingal.ie/sites/default/files/2025-07/q1-2025-pos-over-20k.pdf",
            "https://www.fingal.ie/sites/default/files/2024-09/q2-pos-over-20k-2023.pdf",
            "https://www.fingal.ie/sites/default/files/2024-08/q3-2022-purchase-orders-over-20k.pdf",
        ],
        caveat="cols SupplierID(T)|Acc element(T)|Amount(C); landing links no PO files. "
        "Current files publish supplier NAMES (BUSHELL INTERIORS LTD, ENERGIA…), not bare codes — "
        "supplier_is_id left False so names classify normally (per-row ID_CODE_RE still catches stray codes)",
    ),
    # ---- NEEDS-RENDER to enumerate, but file URLs are known → fetch direct (no Playwright needed) ----
    la(
        "mayo",
        "Mayo",
        "Connacht",
        "county",
        fmt="pdf",
        listing="https://www.mayo.ie/financial-documents/purchase-orders",
        value_kind="po_committed",
        status="READY",
        anchor_listing=True,
        caveat="getattachment/{GUID}/attachment.aspx (dateless URL, basename collides) → "
        "anchor_listing harvest: enumerate all quarterly PDFs from listing HTML, period from "
        "link text 'Quarter N YYYY'. Full back-catalogue (2016+); leading PO#/ID prefix stripped",
    ),
    la(
        "donegal",
        "Donegal",
        "Ulster",
        "county",
        fmt="pdf",
        listing="https://www.donegalcoco.ie/en/services/other-services/finance/finance-publications",
        value_kind="po_committed",
        status="DIRECT",
        direct=[
            "https://www.donegalcoco.ie/media/h0flvm3b/2025.pdf",
            "https://www.donegalcoco.ie/media/b2aopuh2/2024.pdf",
        ],
        caveat="yearly (not quarterly) PDFs at /media/{code}/{YYYY}.pdf; leading PO#/ID prefix stripped",
    ),
    # ---- NEEDS-RENDER, no known direct URL → Playwright enumeration deferred (v2) ----
    la(
        "carlow",
        "Carlow",
        "Leinster",
        "county",
        fmt="pdf",
        status="NEEDS-RENDER",
        listing="https://carlow.ie/information-technology/statistics-and-reports/financial-statistical-reports",
        value_kind="po_committed",
        caveat="JS/SPA file list — Playwright enumeration deferred",
    ),
    la(
        "cavan",
        "Cavan",
        "Ulster",
        "county",
        fmt="pdf",
        status="NEEDS-RENDER",
        listing="https://www.cavancoco.ie/file-library/business/procurement/over-20k/",
        value_kind="po_committed",
        caveat="JS-rendered file library — Playwright enumeration deferred",
    ),
    la(
        "roscommon",
        "Roscommon",
        "Connacht",
        "county",
        fmt="pdf",
        status="NEEDS-RENDER",
        listing="https://www.roscommoncoco.ie/en/Download-It/Finance-Publications/",
        value_kind="po_committed",
        caveat="JS-rendered download portal — Playwright enumeration deferred",
    ),
    # ---- non-publishers (kept for the full 31 census; never parsed) ----
    la(
        "dublin_city",
        "Dublin City",
        "Dublin",
        "dublin",
        fmt="-",
        status="NON-PUBLISHER",
        listing="https://www.dublincity.ie/council/budgets-and-finance/public-procurement",
        value_kind="po_committed",
        caveat="RE-CONFIRMED non-publisher 2026-06-08 (the biggest LA, but it does NOT publish a "
        "supplier-level PO/payments-over-€20k SPEND list). Its procurement page carries only policy "
        "PDFs + a stale 2018 'List_of_Contracts_Awarded_over_€25000.xls' (AWARD grain, ceilings — "
        "belongs to the awards lane, never the payments fact). data.gov.ie has only the AGGREGATE "
        "prompt-payment schedule (quarterly invoice TOTALS, not supplier-level). Real coverage gap, "
        "structural — not a sourcing miss.",
    ),
    la(
        "dlr",
        "Dún Laoghaire-Rathdown",
        "Dublin",
        "dublin",
        fmt="-",
        status="NON-PUBLISHER",
        listing="https://www.dlrcoco.ie/governance/procurement",
        value_kind="po_committed",
        caveat="policy PDFs only; FOI territory",
    ),
]

TIER = {"po_committed": "COMMITTED", "payment_actual": "SPENT"}

# exclude policy/guidance/aggregate docs that share the listing page
POLICY_RE = re.compile(
    r"guide|guidelin|\bplan\b|policy|circular|strategy|manual|terms|fin.?07|privacy|setup|form|"
    r"prompt.?payment|appendix|procedure|annual.?report|charter|scheme|contract|10.?m\b|over.?10|map|"
    r"budget|estimate|adopted|statutory.?audit|financial.?statement|\bafs\b|revenue.?account|"
    r"quality.?assurance|\bpsc\b|spending.?code|oversight|\bnoac\b|qa.?report",
    re.I,
)
DATA_FILE_RE = re.compile(r"q[1-4]\b|qtr|quarter|20[12]\d|over.?20|20k|purchase|payment|\bpo[s]?\b", re.I)
NAV_HINT = re.compile(r"purchase|procure|over.?20|20k|payment|quarter|qtr|finance|publication|spend|supplier", re.I)


def hr(t: str) -> None:
    print(f"\n{'=' * 78}\n{t}\n{'=' * 78}")


# ============================================================================ fetch
REPORT = FetchReport("la_payments")
LAST_ERR: dict = {}  # set by fetch_bytes on failure, read by the download loop


def _curl(url: str) -> bytes | None:
    try:
        p = subprocess.run(
            ["curl", "-sS", "-k", "-L", "--max-time", "90", "-A", H["User-Agent"], url],
            capture_output=True,
            timeout=120,
        )
        return p.stdout if p.returncode == 0 and p.stdout else None
    except Exception:
        return None


def fetch_bytes(url: str) -> bytes | None:
    # Sligo publishes hrefs with raw spaces — requests/curl reject them as malformed;
    # '%' stays in the safe set so already-encoded hrefs (Galway) don't double-encode.
    url = quote(url, safe="!#$%&'()*+,/:;=?@[]~")
    LAST_ERR.clear()
    try:
        r = requests.get(url, headers=H, timeout=90, allow_redirects=True)
        r.raise_for_status()
        return r.content
    except Exception as e:
        ec, status = classify_exception(e)
        LAST_ERR.update({"error_class": ec, "http_status": status})
        b = _curl(url)
        if b:
            LAST_ERR.clear()
        return b


def fetch_text(url: str) -> str | None:
    b = fetch_bytes(url)
    return b.decode("utf-8", "ignore") if b else None


def fetch_to_bronze(slug: str, url: str, ext: str) -> bytes | None:
    """Self-fetch a source file to bronze/pdfs/la_procurement/<slug>/ and reuse the cached
    copy on re-runs (immutable, re-derivable medallion bronze; same shape as afs)."""
    # Opaque getattachment URLs all share the basename "attachment.aspx" — key the bronze cache
    # by the GUID path segment instead, or every file would collide onto one cached copy.
    if "/getattachment/" in url:
        guid = url.split("/getattachment/", 1)[1].split("/", 1)[0]
        dest = BRONZE / slug / f"getattachment_{re.sub(r'[^A-Za-z0-9]', '', guid)[:40]}.pdf"
    else:
        dest = BRONZE / slug / (re.sub(r"[^A-Za-z0-9._-]", "_", unquote(url.rsplit("/", 1)[-1]))[:80] or "file")
    if not dest.suffix:
        dest = dest.with_suffix(ext if ext in DATA_EXT else ".pdf")
    if dest.exists() and dest.stat().st_size > 1500:
        return dest.read_bytes()
    time.sleep(1.0)  # politeness: only on a real network fetch, never on a cache hit
    b = fetch_bytes(url)
    if b and len(b) > 1500:
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(b)
    return b


# ============================================================================ harvest
def harvest_files(cf: dict, crawl_cap: int = 8) -> list[str]:
    """All real PO/payment data-file links for a council: direct_files + landing page +
    one-hop crawl, minus policy/aggregate docs. Honours an optional `include` filter."""
    found: list[str] = list(cf["direct_files"])
    html = fetch_text(cf["listing_url"])

    # anchor_listing: the council serves files from opaque, dateless URLs (e.g. Mayo's
    # /getattachment/{GUID}/attachment.aspx) but the listing's LINK TEXT carries the period.
    # Enumerate (url, anchor-text) straight from the listing HTML — the normal scan() can't
    # (these hrefs aren't *.pdf and every basename is "attachment.aspx", so it filters + dedups
    # them away). Period is derived from the text and stashed per-url for emit_file to consume.
    if cf.get("anchor_listing") and html:
        cf["_period_by_url"] = {}
        seen_period: set[str] = set()
        for href, inner in re.findall(r'<a[^>]+href="([^"]*getattachment[^"]*)"[^>]*>(.*?)</a>', html, re.I | re.S):
            url = urljoin(cf["listing_url"], href)
            text = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", inner)).strip()
            per, yr, q = period_from_text(text)
            if yr is None or yr < 2016:  # corpus convention: defer pre-2016 layouts
                continue
            if per in seen_period:  # one file per period (newest listed first wins)
                continue
            seen_period.add(per)
            cf["_period_by_url"][url] = (per, yr, q)
            found.append(url)
        return found

    def scan(page_html: str, base: str) -> list[str]:
        out = []
        for href in HREF_RE.findall(page_html):
            low = href.lower().split("?")[0]
            if not any(low.endswith(e) for e in DATA_EXT):
                continue
            dec = unquote(href)  # decode %20 so word-boundary filters (\bplan\b) actually fire
            if POLICY_RE.search(dec) or not DATA_FILE_RE.search(dec):
                continue
            if cf["include"] and not cf["include"].search(dec):
                continue
            out.append(urljoin(base, href))
        return out

    if html:
        hits = scan(html, cf["listing_url"])
        if not hits:  # one-hop crawl same-host nav links
            host = urlparse(cf["listing_url"]).netloc
            subs, seen = [], set()
            for href in HREF_RE.findall(html):
                full = urljoin(cf["listing_url"], href)
                low = full.lower().split("?")[0]
                if urlparse(full).netloc != host or full == cf["listing_url"]:
                    continue
                if any(low.endswith(e) for e in DATA_EXT) or full in seen:
                    continue
                if NAV_HINT.search(href):
                    seen.add(full)
                    subs.append(full)
            for s in subs[:crawl_cap]:
                sub_html = fetch_text(s)
                if sub_html:
                    hits.extend(scan(sub_html, s))
        found.extend(hits)

    # dedup by basename (same file via two hosts → one entry); newest year first;
    # defer pre-2016 files (old programme-group wording / awkward legacy layouts).
    seen, uniq = set(), []
    for u in sorted(found, key=lambda u: (_url_year(u) or 0, u), reverse=True):
        yr = _url_year(u)
        if yr is not None and yr < 2016:
            continue
        key = u.rsplit("/", 1)[-1].split("?")[0].lower()
        if key not in seen:
            seen.add(key)
            uniq.append(u)
    return uniq


def _url_year(url: str) -> int | None:
    m = re.search(r"20[12]\d", unquote(url.rsplit("/", 1)[-1]))
    return int(m.group()) if m else None


# ============================================================================ readers
def to_eur(token) -> float | None:
    if token is None:
        return None
    if isinstance(token, (int, float)) and not isinstance(token, bool):
        return float(token)
    s = str(token).strip()
    neg = s.startswith("(") and s.endswith(")")
    m = NUM_RE.search(s)
    if not m:
        return None
    with contextlib.suppress(ValueError):
        v = float(m.group().replace(",", ""))
        return -v if neg else v
    return None


def strip_id_prefix(s: str) -> str:
    """Remove a leading PO#/vendor-ID run (Mayo/Donegal/Waterford) so the name is clean
    and distinct-supplier counts don't explode (each PO# is unique)."""
    return DIGIT_PREFIX.sub("", str(s or "")).strip(" -:|,")


# ---- PDF: layout-agnostic largest-x-gap split (proven in probe_procurement_pdf_counties) ----
def cluster_word_rows(page, ytol: float = 3.0) -> list[list]:
    words = page.get_text("words")
    words.sort(key=lambda w: (round(w[1] / ytol), w[0]))
    rows, cur, cur_y = [], [], None
    for w in words:
        y = w[1]
        if cur_y is None or abs(y - cur_y) <= ytol:
            cur.append(w)
            cur_y = y if cur_y is None else cur_y
        else:
            rows.append(cur)
            cur, cur_y = [w], y
    if cur:
        rows.append(cur)
    return rows


def split_row(words: list) -> dict | None:
    """Drop the (rightmost) money token; split the remaining words at their largest
    horizontal gap → left=supplier, right=description. Order-independent."""
    ws = sorted(words, key=lambda w: w[0])
    money_idx = [i for i, w in enumerate(ws) if MONEY_RE.search(w[4])]
    if not money_idx:
        return None
    amt_i = max(money_idx, key=lambda i: ws[i][0])  # rightmost money (Galway amount-last)
    eur = to_eur(ws[amt_i][4])
    rest = [w for i, w in enumerate(ws) if i != amt_i and not MONEY_RE.search(w[4])]
    if eur is None or not rest:
        return None
    if len(rest) >= 2:
        gap, cut = max((rest[i + 1][0] - rest[i][2], i) for i in range(len(rest) - 1))
        if gap < 12:
            cut = len(rest) - 1
    else:
        cut = 0
    supplier = strip_id_prefix(" ".join(w[4] for w in rest[: cut + 1]).strip(" -:|"))
    description = " ".join(w[4] for w in rest[cut + 1 :]).strip(" -:|")
    if len(supplier) < 3:
        return None
    return {"supplier": supplier, "eur": eur, "description": description or None, "po": None, "paid": None}


def read_pdf(b: bytes) -> tuple[list[dict], int]:
    doc = fitz.open(stream=b, filetype="pdf")
    chars = sum(len(doc[i].get_text("text").strip()) for i in range(min(3, doc.page_count)))
    rows: list[dict] = []
    for i in range(doc.page_count):
        for wrow in cluster_word_rows(doc[i]):
            rec = split_row(wrow)
            if rec:
                rec["page"] = i + 1
                rows.append(rec)
    doc.close()
    return rows, chars  # chars>200 on page-sample ⇒ digital


# Offaly GL30 reading-order PDFs: each field on its OWN line (the word-geometry reader above
# under-reads them ~50%). 4 sub-layouts across years — [suppid][supplier][product][ordered][received]
# (5-line), [suppid supplier][product][ordered][received] (4-line), [supplier][product][value]
# (3-line, single amount), and a 3-line variant whose amounts omit decimals ("330360" = €330,360).
# Offaly's files are "Purchase Orders GREATER THAN €20,000", so every amount is ≥€20k while the
# supplier-id is ~5 digits (<€20k) — the threshold cleanly separates the two. Anchor on the amount
# line; the text lines before it are supplier (first, minus any inline suppid) + product.
_OFF_AMT = re.compile(r"^(\d{1,3}(?:,\d{3})*(?:\.\d{2})?|\d+(?:\.\d{2})?)$")
_OFF_HDR = re.compile(
    r"supp\.?id|product|order(ed)?\s+(value|amount)|received amount|purchase orders|greater than|"
    r"^page\b|^screen|^user|^\d{2}/\d{2}/\d{4}",
    re.I,
)
_OFF_NUM = re.compile(r"^\d{1,3}(?:,\d{3})*(?:\.\d{2})?$|^\d+(?:\.\d{2})?$")


def read_pdf_offaly(b: bytes) -> tuple[list[dict], int]:
    doc = fitz.open(stream=b, filetype="pdf")
    chars = sum(len(doc[i].get_text("text").strip()) for i in range(min(3, doc.page_count)))
    lines = [ln.strip() for pi in range(doc.page_count) for ln in doc[pi].get_text().splitlines() if ln.strip()]
    doc.close()

    def amt(ln: str) -> float | None:
        m = _OFF_AMT.match(ln)
        if not m:
            return None
        v = float(m.group(1).replace(",", ""))
        return v if 20000 <= v <= 50_000_000 else None

    recs: list[dict] = []
    buf: list[str] = []
    last_was_amt = False
    for ln in lines:
        if _OFF_HDR.match(ln) or TOTAL_RE.search(ln):
            buf, last_was_amt = [], False
            continue
        a = amt(ln)
        if a is not None:
            if buf:  # close the record on its (first/ordered) amount line
                supplier = re.sub(r"^\d{3,6}\s+", "", buf[0]).strip()  # strip an inline supp-id
                description = " ".join(buf[1:]) or None
                if len(supplier) >= 3 and not _OFF_NUM.match(supplier):
                    recs.append({"supplier": supplier, "eur": a, "description": description, "po": None, "paid": None, "page": 1})
                buf, last_was_amt = [], True
            else:  # a trailing 'received amount' duplicate — ignore
                last_was_amt = True
            continue
        if last_was_amt:  # first line after an amount run starts a new record
            buf = []
        buf.append(ln)
        last_was_amt = False
    return recs, chars


# ---- XLSX / CSV: header-detect + content-fallback (handles odd headers e.g. Ap/Ar ID) ----
def _col_roles(header: list[str], data: list[list]) -> dict[str, int | None]:
    # "money-ness" of a column: how many cells parse to a PLAUSIBLE PO/payment amount
    # (>= €100, < €50m). A PO-number column (9-digit integers ~€400m) scores 0, so it can
    # never be mistaken for the amount even when it is the "most numeric" column — the bug
    # that made South Dublin's 2023/24 files read PO numbers (≈€400m) as the value.
    def money_score(j: int) -> int:
        return sum(1 for r in data[:300] if j < len(r) and (v := to_eur(r[j])) is not None and 100 <= v < 50_000_000)

    roles: dict[str, int | None] = {}
    for role, rx in (
        ("supplier", SUP_RE),
        ("amount", AMT_RE),
        ("description", DESC_RE),
        ("paid", PAID_RE),
        ("po", PO_RE),
    ):
        cands = [j for j, h in enumerate(header) if h and rx.search(h)]
        if role == "amount" and cands:  # prefer the most money-like amount-named column
            cands.sort(key=lambda j: -money_score(j))
        roles[role] = cands[0] if cands else None
    # content-fallback for odd headers: amount = most money-like col, supplier = most-alphabetic
    if (roles["supplier"] is None or roles["amount"] is None) and data:
        ncol = max((len(r) for r in data[:200]), default=0)
        txts = [0] * ncol
        for row in data[:200]:
            for j in range(min(ncol, len(row))):
                v = row[j]
                if isinstance(v, str) and re.search(r"[A-Za-z]", v):
                    txts[j] += 1
        if roles["amount"] is None and any(money_score(j) for j in range(ncol)):
            roles["amount"] = max(range(ncol), key=money_score)
        if roles["supplier"] is None and any(txts):
            roles["supplier"] = max(
                (j for j in range(ncol) if j != roles["amount"]), key=lambda j: txts[j], default=None
            )
    return roles


def read_xlsx(b: bytes) -> tuple[list[dict], int]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return _tabular_rows([[c for c in row] for row in raw])


def read_xls(b: bytes) -> tuple[list[dict], int]:
    """Legacy .xls (Wexford's 2016-2024 quarters). xlrd 2.x reads .xls only."""
    import xlrd

    book = xlrd.open_workbook(file_contents=b)
    sh = book.sheet_by_index(0)
    grid = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    return _tabular_rows([[("" if c == "" else c) for c in row] for row in grid])


def read_csv(b: bytes) -> tuple[list[dict], int]:
    # Robust cp1252/UTF-8 decode before parsing (see shared.text_encoding): the old
    # utf8-lossy mangled council exports' cp1252 apostrophes/fadas into '�' (e.g. Meath
    # "O'Mahony Pike Architects", "Navan O'Mahonys CLG").
    df = pl.read_csv(
        io.BytesIO(decode_table_bytes(b).encode("utf-8")),
        infer_schema_length=0,
        truncate_ragged_lines=True,
        ignore_errors=True,
    )
    df = df.rename({c: c.replace("﻿", "").strip() for c in df.columns})
    return _tabular_rows([df.columns] + [list(r) for r in df.iter_rows()])


def _tabular_rows(grid: list[list]) -> tuple[list[dict], int]:
    """grid[0..] rows; find the header row, map roles, emit records.

    Header = the early row with the MOST non-empty cells that also carries a
    supplier/amount keyword. The non-empty-cell count is what discriminates a real
    multi-column header from a single-cell TITLE row (a Kilkenny title cell holds
    "…Orders Over €20,000…", whose € would otherwise fool a keyword-only scorer)."""

    def nonempty(row):
        return sum(1 for c in (row or []) if c is not None and str(c).strip())

    def role_hits(row):
        return sum(
            bool(SUP_RE.search(str(c)) or AMT_RE.search(str(c)) or DESC_RE.search(str(c)))
            for c in (row or [])
            if c is not None
        )

    cands = [(nonempty(grid[i]), role_hits(grid[i]), -i, i) for i in range(min(8, len(grid)))]
    # require ≥2 columns and ≥1 role keyword; tie-break by role hits then earliest row
    valid = [c for c in cands if c[0] >= 2 and c[1] >= 1]
    hrow = max(valid or cands, key=lambda c: (c[0], c[1], c[2]))[3]
    header = [str(c).strip() if c is not None else "" for c in grid[hrow]]
    data = [r for r in grid[hrow + 1 :] if any(c is not None and str(c).strip() for c in r)]
    roles = _col_roles(header, data)
    si, ai = roles["supplier"], roles["amount"]
    if ai is None:
        return [], len(data)
    di, pi, pdi = roles["description"], roles["po"], roles["paid"]

    def cell(r, j):
        return r[j] if j is not None and j < len(r) else None

    out: list[dict] = []
    for r in data:
        eur = to_eur(cell(r, ai))
        sup = cell(r, si)
        if eur is None or sup is None or not str(sup).strip():
            continue
        out.append(
            {
                "supplier": strip_id_prefix(str(sup).strip()),
                "eur": eur,
                "description": (str(cell(r, di)).strip() or None) if cell(r, di) is not None else None,
                "po": (str(cell(r, pi)).strip() or None) if cell(r, pi) is not None else None,
                "paid": (str(cell(r, pdi)).strip() or None) if cell(r, pdi) is not None else None,
                "page": None,
            }
        )
    return out, len(data)


READERS = {".pdf": read_pdf, ".xlsx": read_xlsx, ".xls": read_xls, ".csv": read_csv}


# ============================================================================ period
def period_from_url(url: str) -> tuple[str | None, int | None, int | None]:
    name = unquote(url.rsplit("/", 1)[-1])  # decode %20 so it doesn't read as a "2020"
    # lenient: a contiguous 20xx token (handles CMS dedup suffixes e.g. "...-2025 1.xlsx");
    # "20-000" never matches (not four contiguous digits).
    y = re.search(r"20[12]\d", name)
    year = int(y.group()) if y else None
    q = QUARTER_RE.search(name)
    quarter = next((int(g) for g in (q.groups() if q else []) if g), None) if q else None
    period = f"{year}-Q{quarter}" if year and quarter else (str(year) if year else None)
    return period, year, quarter


def period_from_text(text: str) -> tuple[str | None, int | None, int | None]:
    """Period from a listing's anchor/link TEXT (not the URL). For councils that serve files
    from opaque, dateless URLs (e.g. Mayo's /getattachment/{GUID}/attachment.aspx) but whose
    link title states the period — 'Quarter 1 2024', 'Quarter-1-2024', 'Q1 2024'. Punctuation-
    tolerant. Falls back to a bare year. The published date (dd/mm/yyyy) is intentionally NOT
    matched — the PO period differs from the upload date."""
    m = re.search(r"Quarter[\s\-]*([1-4])[\s\-,]*?(20[12]\d)", text, re.I) or re.search(
        r"\bQ[\s\-]*([1-4])[\s\-,]*?(20[12]\d)", text, re.I
    )
    if m:
        year, quarter = int(m.group(2)), int(m.group(1))
        return f"{year}-Q{quarter}", year, quarter
    y = re.search(r"\b(20[12]\d)\b", text)
    return (y.group(1), int(y.group(1)), None) if y else (None, None, None)


# ============================================================================ extract
def emit_file(cf: dict, file_url: str, b: bytes, ext: str) -> tuple[list[dict], dict]:
    """Parse one file → fact-schema row dicts (pre-classification) + a per-file stat block."""
    fhash = hashlib.sha256(b).hexdigest()[:16]
    # anchor_listing councils carry a period derived from the listing TEXT (their URL is dateless);
    # use it when present, else fall back to the URL.
    override = cf.get("_period_by_url", {}).get(file_url)
    period, year, quarter = override if override else period_from_url(file_url)
    recs, scanned_hint = READERS[ext](b)
    digital = scanned_hint > 200 if ext == ".pdf" else True

    # debit-sign convention: several councils (Kilkenny, Limerick, …) publish amounts as
    # NEGATIVE. Detect it at file level (majority of non-null amounts negative) and abs the
    # whole file, so a refund-style lone negative elsewhere is left untouched. `neg_abs`
    # forces it where the sample was unanimous.
    amts = [r["eur"] for r in recs if r["supplier"] and r["eur"] is not None]
    neg_share = (sum(a < 0 for a in amts) / len(amts)) if amts else 0.0
    do_abs = cf["neg_abs"] or neg_share > 0.5

    rows_out: list[dict] = []
    for srn, rec in enumerate(recs):
        sup = rec["supplier"]
        eur = rec["eur"]
        if sup is None or eur is None:
            continue
        if TOTAL_RE.search(sup):  # drop printed total / "all suppliers" rows
            continue
        if do_abs:
            eur = abs(eur)
        if eur < MIN_EUR:
            continue
        rows_out.append(
            {
                "publisher_id": f"ie_la_{cf['slug']}",
                "publisher_name": cf["council"],
                "publisher_type": "local_authority",
                "sector": "local_government",
                "region": cf["region"],
                "entity_type": cf["entity"],
                "source_landing_url": cf["listing_url"],
                "source_file_url": file_url,
                "source_file_hash": fhash,
                "period": period,
                "year": year,
                "quarter": quarter,
                "supplier_raw": sup,
                "amount_eur": eur,
                "value_kind": cf["value_kind"],
                "realisation_tier": TIER[cf["value_kind"]],
                "description": rec["description"],
                "po_number": rec["po"],
                "paid_flag": rec["paid"],
                "source_row_number": srn,
                "source_page_number": rec["page"],
                "parser_name": f"la_payments_{ext.lstrip('.')}",
                "parser_version": PARSER_VERSION,
                "source_caveat": cf["caveat"] or None,
                "supplier_is_id_code": cf["supplier_is_id"],
            }
        )

    # content-validity gate: a real PO/payment list has many rows and a healthy share of
    # company / public-body payees. A stray policy / review / heritage doc that slips past
    # the filename filters yields a handful of rows of prose-numbers with almost no
    # company suffixes → reject the whole file so it never pollutes the fact.
    org_share = (
        (
            sum(
                bool(COMPANY_SUFFIX.search(r["supplier_raw"]) or PUBLIC_BODY.search(r["supplier_raw"]))
                for r in rows_out
            )
            / len(rows_out)
        )
        if rows_out
        else 0.0
    )
    looks_like_po = len(rows_out) >= 8 and org_share >= 0.12
    accepted = rows_out if looks_like_po else []
    stat = {
        "file": file_url.rsplit("/", 1)[-1][:48],
        "digital": digital,
        "rows": len(accepted),
        "raw_rows": len(rows_out),
        "period": period,
        "total_eur": sum(r["amount_eur"] for r in accepted),
        "org_share": round(org_share, 2),
        "valid": looks_like_po,
    }
    return accepted, stat


def classify_and_flag(df: pl.DataFrame) -> pl.DataFrame:
    """supplier_normalised + supplier_class + privacy quarantine + value_safe_to_sum.

    Quarantine is APPLIED here (unlike the deferred semistate run): sole-trader /
    individual / bare-ID payees are personal data → public_display=False.
    """
    if df.is_empty():
        return df
    df = (
        df.with_columns(
            name_norm_expr("supplier_raw").alias("supplier_normalised"),
            pl.col("supplier_raw")
            .map_elements(lambda s: bool(PUBLIC_BODY.search(s or "")), return_dtype=pl.Boolean)
            .alias("_pub"),
            pl.col("supplier_raw")
            .map_elements(lambda s: bool(COMPANY_SUFFIX.search(s or "")), return_dtype=pl.Boolean)
            .alias("_co"),
            pl.col("supplier_raw")
            .map_elements(lambda s: bool(FOREIGN_FORM.search(s or "")), return_dtype=pl.Boolean)
            .alias("_for"),
            pl.col("supplier_raw")
            .map_elements(lambda s: bool(ID_CODE_RE.match((s or "").strip())), return_dtype=pl.Boolean)
            .alias("_idc"),
        )
        .with_columns(
            pl.when(pl.col("supplier_is_id_code") | pl.col("_idc"))
            .then(pl.lit("id_code"))
            .when(pl.col("_pub"))
            .then(pl.lit("public_body"))
            .when(pl.col("_co"))
            .then(pl.lit("company"))
            .when(pl.col("_for"))
            .then(pl.lit("foreign_company"))
            .when(pl.col("supplier_raw").is_null() | (pl.col("supplier_raw").str.strip_chars() == ""))
            .then(pl.lit("unknown"))
            .otherwise(pl.lit("sole_trader_or_individual"))
            .alias("supplier_class"),
        )
        .with_columns(
            # quarantine personal-data classes (kept in the fact, hidden from display)
            pl.col("supplier_class").is_in(["sole_trader_or_individual", "id_code", "unknown"]).alias("_quar"),
        )
        .with_columns(
            pl.when(pl.col("_quar")).then(pl.lit("quarantined")).otherwise(pl.lit("public")).alias("privacy_status"),
            (~pl.col("_quar")).alias("public_display"),
            pl.when(pl.col("supplier_class") == "sole_trader_or_individual")
            .then(pl.lit("sole_trader_or_individual"))
            .when(pl.col("supplier_class") == "id_code")
            .then(pl.lit("bare_id_code"))
            .when(pl.col("supplier_class") == "unknown")
            .then(pl.lit("unresolved_name"))
            .otherwise(pl.lit(None))
            .alias("privacy_reason"),
            # po_committed + payment_actual are summable within council+period; abs already applied
            (
                pl.col("value_kind").is_in(["po_committed", "payment_actual"])
                & pl.col("amount_eur").is_not_null()
                & (pl.col("amount_eur") > 0)
            ).alias("value_safe_to_sum"),
        )
        .drop(["_pub", "_co", "_for", "_idc", "_quar"])
    )
    return df


FACT_COLS = [
    "publisher_id",
    "publisher_name",
    "publisher_type",
    "sector",
    "region",
    "entity_type",
    "source_landing_url",
    "source_file_url",
    "source_file_hash",
    "period",
    "year",
    "quarter",
    "supplier_raw",
    "supplier_normalised",
    "supplier_class",
    "supplier_is_id_code",
    "amount_eur",
    "value_kind",
    "realisation_tier",
    "value_safe_to_sum",
    "description",
    "po_number",
    "paid_flag",
    "source_row_number",
    "source_page_number",
    "parser_name",
    "parser_version",
    "privacy_status",
    "public_display",
    "privacy_reason",
    "source_caveat",
]


# ============================================================================ main
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true", help="harvest-only: print candidate files, no parse")
    ap.add_argument("--only", default="", help="comma-separated council slugs")
    # default must cover the full back-catalogue (Cork City publishes 66 files to 2016): the
    # fact is OVERWRITTEN each run, so a low cap silently downgrades silver to a recent slice
    # and the next consolidate would wipe gold history. Pass a low value for smoke tests only.
    ap.add_argument("--max-files", type=int, default=99, help="cap files parsed per council (recent-first)")
    ap.add_argument(
        "--merge",
        action="store_true",
        help="with --only: update JUST those councils inside the existing fact (all other councils "
        "kept) instead of overwriting the whole parquet. Mirrors procurement_public_body_extract's "
        "--merge; lets one council be fixed+re-run without wiping the other 22 (a plain --only would).",
    )
    args = ap.parse_args()
    only = {x.strip() for x in args.only.split(",") if x.strip()} or None
    if args.merge and not only:
        raise SystemExit("--merge requires --only (it updates the named councils in place)")

    councils = [
        c
        for c in SCHEMA_MAP
        if (not only or c["slug"] in only) and c["status"] not in ("NON-PUBLISHER", "NEEDS-RENDER")
    ]
    hr(f"LA PO/PAYMENTS-over-€20k EXTRACT — {len(councils)} councils" + (" (LIST MODE)" if args.list else ""))

    all_rows: list[dict] = []
    per_council: list[dict] = []
    cro = None if args.list else pl.read_parquet(CRO).select(["name_norm", "company_num"])

    for cf in councils:
        files = harvest_files(cf)
        print(f"\n[{cf['slug']:<14}] {cf['council']}  ({cf['value_kind']} → {TIER[cf['value_kind']]})")
        print(f"   files harvested: {len(files)}")
        if not files:
            REPORT.record_zero_harvest(
                publisher_id=cf["slug"], publisher_name=cf["council"], listing_url=cf["listing_url"]
            )
        if args.list:
            for u in files[:10]:
                print(f"     - {u.rsplit('/', 1)[-1][:72]}")
            per_council.append({"slug": cf["slug"], "files": len(files)})
            continue

        c_rows: list[dict] = []
        file_stats: list[dict] = []
        breaker = Breaker()
        file_list = files[: args.max_files]
        for i, u in enumerate(file_list):
            ext = next((e for e in DATA_EXT if u.lower().split("?")[0].endswith(e)), "")
            if ext not in READERS and not u.endswith(".aspx"):
                continue
            b = fetch_to_bronze(cf["slug"], u, ext)
            read_ext = ext if ext in READERS else ".pdf"  # Mayo getattachment.aspx serves a PDF
            bad = classify_body(b, b"%PDF" if read_ext == ".pdf" else None) if b else None
            if not b or bad:
                breaker.record(False)
                err = bad or LAST_ERR.get("error_class", "unknown")
                REPORT.record_failure(
                    publisher_id=cf["slug"],
                    publisher_name=cf["council"],
                    url=u,
                    listing_url=cf["listing_url"],
                    error_class=err,
                    http_status=LAST_ERR.get("http_status"),
                    attempts=2 if not b else 1,
                )
                print(f"     ! download failed ({err}): {u.rsplit('/', 1)[-1][:48]}")
                if breaker.tripped:
                    rest = len(file_list) - i - 1
                    REPORT.record_breaker_trip(
                        publisher_id=cf["slug"], publisher_name=cf["council"], files_skipped=rest
                    )
                    print(
                        f"     !! breaker tripped: {breaker.consecutive} consecutive failures — skipping {rest} remaining files"
                    )
                    break
                continue
            breaker.record(True)
            write_sentinel("la_payments", cf["slug"], u)
            try:
                rows, stat = emit_file(cf, u, b, read_ext)
            except Exception as e:
                print(f"     ! parse error ({type(e).__name__}): {u.rsplit('/', 1)[-1][:44]}")
                continue
            # aggregate_guard: a 1-row "file" is a prompt-pay/total aggregate, not a PO list
            if cf["aggregate_guard"] and stat["rows"] <= 1:
                print(f"     ~ skip (aggregate_guard, {stat['rows']} row): {stat['file']}")
                continue
            if not stat["valid"]:
                print(
                    f"     ~ reject (not a PO list: {stat['raw_rows']} rows, "
                    f"{stat['org_share']:.0%} orgs): {stat['file']}"
                )
                continue
            c_rows.extend(rows)
            file_stats.append(stat)
            print(
                f"     -> {stat['file']:<48} rows={stat['rows']:<5} €{stat['total_eur'] / 1e6:>6.2f}m  {stat['period']}"
            )

        # per-council CRO reconciliation band
        cro_rate = 0.0
        if c_rows:
            cdf = pl.DataFrame({"supplier_raw": [r["supplier_raw"] for r in c_rows]})
            sup = (
                cdf.with_columns(name_norm_expr("supplier_raw").alias("nn"))
                .filter(pl.col("nn").str.len_chars() >= 4)
                .unique(subset=["nn"])
            )
            hit = (
                sup.join(cro, left_on="nn", right_on="name_norm", how="left")
                .filter(pl.col("company_num").is_not_null())
                .select("nn")
                .n_unique()
            )
            cro_rate = hit / max(1, sup.height)
        all_rows.extend(c_rows)
        per_council.append(
            {
                "slug": cf["slug"],
                "council": cf["council"],
                "region": cf["region"],
                "value_kind": cf["value_kind"],
                "files_parsed": len(file_stats),
                "rows": len(c_rows),
                "total_eur": sum(s["total_eur"] for s in file_stats),
                "cro_rate": round(cro_rate, 3),
                "periods": sorted({s["period"] for s in file_stats if s["period"]}),
            }
        )
        if c_rows:
            print(
                f"   == {cf['council']}: {len(c_rows):,} rows  €{sum(s['total_eur'] for s in file_stats) / 1e6:.1f}m  CRO {cro_rate:.0%}"
            )

    if args.list:
        hr("LIST DONE — lock URLs, then run without --list")
        return
    if not all_rows:
        print("\nno rows extracted")
        REPORT.write()
        for line in REPORT.summary_lines():
            print(line)
        return

    df = pl.DataFrame(all_rows, infer_schema_length=None)
    df = classify_and_flag(df)
    df = df.select([c for c in FACT_COLS if c in df.columns]).sort(["publisher_id", "year", "quarter"])

    # ── Cross-period republish de-duplication ──────────────────────────────────────────────
    # Several councils carry forward / re-list the same line (esp. big multi-year "CONTRACT
    # PAYMENTS") in later quarters' files, so a naive sum double-counts it (Mayo's €918m → the
    # same BAM-JV €9.2m line appearing in 2022-Q4 AND 2023-Q2). Rule: for an identical line
    # (publisher+supplier+amount+description+po_number), keep EVERY occurrence in its EARLIEST
    # period and drop any re-listing in a LATER period. Same-period (same-file) repeats are
    # left exactly as published — they are ambiguous, not provably artefacts. Pure row-drop;
    # never alters a value.
    _dkey = ["publisher_id", "supplier_raw", "amount_eur", "description", "po_number"]
    _before = df.height
    df = (
        df.with_columns(pl.col("period").min().over(_dkey).alias("_first_period"))
        .filter(pl.col("period").is_null() | (pl.col("period") == pl.col("_first_period")))
        .drop("_first_period")
    )
    _removed = _before - df.height
    _eur_removed = 0.0  # informational only
    print(f"cross-period republish dedupe: dropped {_removed:,} re-listed lines ({_removed / max(_before,1):.1%})")

    # MERGE MODE: fold the freshly-parsed councils into the existing fact instead of replacing it.
    # Drop any prior rows for the councils we just (re)parsed — idempotent — then concat the
    # untouched councils back on. The kept rows are already classified to the same FACT_COLS, so
    # no re-classification is needed; diagonal_relaxed aligns dtypes defensively.
    if args.merge and OUT_FACT.exists():
        existing = pl.read_parquet(OUT_FACT)
        sel_ids = [f"ie_la_{c['slug']}" for c in councils]
        kept = existing.filter(~pl.col("publisher_id").is_in(sel_ids))
        print(
            f"MERGE: existing fact {existing.height:,} rows / {existing['publisher_id'].n_unique()} councils; "
            f"replacing {existing.height - kept.height:,} rows for {sorted(set(sel_ids) & set(existing['publisher_id'].unique().to_list()))}, "
            f"keeping {kept.height:,} rows for the rest"
        )
        df = pl.concat([kept, df], how="diagonal_relaxed").sort(["publisher_id", "year", "quarter"])

    OUT_FACT.parent.mkdir(parents=True, exist_ok=True)
    save_parquet(df, OUT_FACT, min_rows=MIN_FACT_ROWS)

    hr("SILVER FACT WRITTEN")
    print(f"rows: {df.height:,}  councils: {df['publisher_id'].n_unique()}  ->  {OUT_FACT}")
    with pl.Config(tbl_rows=12):
        print(df.group_by("supplier_class").len().sort("len", descending=True))
    safe = df.filter(pl.col("value_safe_to_sum") & pl.col("public_display"))
    print(f"\npublic + safe-to-sum: {safe.height:,} rows  €{(safe['amount_eur'].sum() or 0) / 1e6:,.1f}m")
    quar = df.filter(~pl.col("public_display"))
    print(f"quarantined (personal data, hidden): {quar.height:,} rows ({quar.height / df.height:.0%})")

    cov = {
        "councils_attempted": len(per_council),
        "councils_with_rows": sum(p["rows"] > 0 for p in per_council),
        "rows_extracted": df.height,
        "rows_public_display": int(df["public_display"].sum()),
        "rows_quarantined": int((~df["public_display"]).sum()),
        "supplier_class_counts": {
            r["supplier_class"]: r["len"] for r in df.group_by("supplier_class").len().iter_rows(named=True)
        },
        "value_kind_counts": {r["value_kind"]: r["len"] for r in df.group_by("value_kind").len().iter_rows(named=True)},
        "public_safe_to_sum_rows": safe.height,
        "public_safe_to_sum_total_eur": float(safe["amount_eur"].sum() or 0),
        "by_council": per_council,
        "privacy_quarantine_applied": True,
        "schema_version": 1,
        "parser_version": PARSER_VERSION,
        "taxonomy": "master value_kind + realisation_tier; unions with public_payments_fact "
        "(rename that fact's amount_semantics→value_kind, derive realisation_tier).",
        "source": "Each LA's own website, Circular Fin 07/2012 publications (CC-BY / attribute).",
        "generated_at": datetime.now(UTC).isoformat(timespec="seconds"),
        "caveat": "One row = one published PO/payment line. value_kind distinguishes "
        "po_committed (ordered/COMMITTED) vs payment_actual (paid/SPENT); only "
        "value_safe_to_sum may be totalled, within a council+period, labelled "
        "'ordered'/'paid', never mixed with award ceilings. Personal-data payees "
        "(sole_trader/id_code/unknown) are quarantined (public_display=False). "
        "A line is a purchase order or payment, NOT evidence of influence.",
    }
    OUT_COV.write_text(json.dumps(cov, indent=2), encoding="utf-8")
    print(f"wrote coverage {OUT_COV}")

    report_path = REPORT.write()
    lines = REPORT.summary_lines()
    if lines:
        print(f"\nfetch-failure report -> {report_path}")
        for line in lines:
            print(line)


if __name__ == "__main__":
    main()
