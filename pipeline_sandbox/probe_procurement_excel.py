"""PROBE (throwaway): SIZE THE TECH DEBT of the EXCEL slice of procurement spend.

The PDF slice is solved (fitz, no OCR). The CSV slice is the eTenders bulk. This
probe assesses the THIRD format — XLSX/XLS — because Excel is where normalisation
debt hides: merged header cells, multi-sheet workbooks, header rows offset below
logos/titles, subtotal/total rows, numbers-stored-as-text, and (worst) every
publisher using a DIFFERENT column schema. The user is rightly cautious: starting
this = assuming tech debt, so MEASURE it before committing.

For each sampled workbook it reports the Excel-specific landmines:
  - sheets (1 vs many), used dims, MERGED-CELL count
  - where the real HEADER row sits (offset below title rows?)
  - detected column headers -> schema SIGNATURE (to count distinct schemas)
  - supplier + amount columns present? amount numeric or text?
  - data rows vs total/subtotal rows
  - .xls legacy (needs xlrd, NOT installed) = a hard gap, flagged not parsed
Then: schema heterogeneity tally + supplier->CRO on the clean ones + debt verdict.

Run:  ./.venv/Scripts/python.exe pipeline_sandbox/probe_procurement_excel.py
Reads CRO silver; downloads sampled workbooks to c:/tmp; writes no repo data.
"""

from __future__ import annotations

import io
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl
import polars as pl
import requests

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from shared.name_norm import name_norm_expr  # noqa: E402

CRO = ROOT / "data/silver/cro/companies.parquet"
TMP = Path("c:/tmp/procurement_xlsx")
H = {"User-Agent": "Mozilla/5.0 (dail-tracker research probe)"}

SUP_RE = re.compile(r"supplier|payee|vendor|beneficiar|name|company|creditor", re.I)
AMT_RE = re.compile(r"amount|value|total|paid|gross|net|cost|€|eur", re.I)
TOTAL_RE = re.compile(r"^\s*(grand\s+)?total|sub-?total", re.I)


def hr(t: str) -> None:
    print(f"\n{'=' * 70}\n{t}\n{'=' * 70}")


def ckan_search(q: str, rows: int = 200) -> list[dict]:
    try:
        r = requests.get("https://data.gov.ie/api/3/action/package_search",
                         params={"q": q, "rows": rows}, headers=H, timeout=60)
        return r.json()["result"]["results"]
    except Exception as e:
        print(f"  ckan ERR {e!r}")
        return []


SPEND_TITLE_RE = re.compile(r"purchase order|payments? over|purchases over|po'?s over|prompt payment", re.I)


def discover_excel() -> list[tuple[str, str, str]]:
    """(publisher, format, url) for xls/xlsx resources in genuine spend datasets."""
    pkgs: dict[str, dict] = {}
    for q in ("Purchase Orders over 20000", "Payments over 20000", "Prompt Payment", "spend over 20000"):
        for d in ckan_search(q):
            if SPEND_TITLE_RE.search(d.get("title", "")):
                pkgs[d["id"]] = d
    out = []
    for d in pkgs.values():
        pub = ((d.get("organization") or {}).get("title") or d.get("title") or "?")[:45]
        for x in d.get("resources", []):
            u = x.get("url", "") or ""
            fmt = (x.get("format", "") or (u.rsplit(".", 1)[-1] if "." in u else "")).lower()
            if fmt in ("xls", "xlsx", "xlsm") or u.lower().endswith((".xls", ".xlsx", ".xlsm")):
                out.append((pub, "xls" if u.lower().endswith(".xls") else "xlsx", u))
    return out


# off-catalog XLSX (South Dublin = the clean template found in the Dublin pilot)
OFFCATALOG = [
    ("South Dublin County Council", "xlsx",
     "https://www.sdcc.ie/en/services/business/payments/purchase-order-over-20-000-quarter-1-2025.xlsx"),
    ("South Dublin County Council", "xlsx",
     "https://www.sdcc.ie/en/services/business/payments/purchase-order-over-20-000-quarter-2-2024.xlsx"),
]


def fetch(url: str, idx: int) -> bytes | None:
    ext = ".xls" if url.lower().endswith(".xls") else ".xlsx"
    dest = TMP / (re.sub(r"[^A-Za-z0-9._-]", "_", url.rsplit("/", 1)[-1])[:60] or f"wb_{idx}{ext}")
    if not dest.name.lower().endswith((".xls", ".xlsx")):
        dest = dest.with_name(f"wb_{idx}{ext}")
    if dest.exists() and dest.stat().st_size > 1000:
        return dest.read_bytes()
    try:
        b = requests.get(url, headers=H, timeout=90, allow_redirects=True).content
    except Exception as e:
        print(f"    fetch ERR {type(e).__name__}")
        return None
    if b[:2] not in (b"PK", b"\xd0\xcf"):  # zip(xlsx) or OLE(xls)
        print(f"    not an Excel file (got {b[:8]!r})")
        return None
    TMP.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(b)
    return b


def find_header(ws, scan: int = 12):
    """The header is the first row (within the top `scan`) where >=2 cells match
    supplier/amount keywords — Excel reports often have title/logo rows above it."""
    best = None
    for r in range(1, min(scan, ws.max_row or 1) + 1):
        vals = [str(c.value).strip() if c.value is not None else "" for c in ws[r]]
        nonempty = [v for v in vals if v]
        hits = sum(bool(SUP_RE.search(v) or AMT_RE.search(v)) for v in nonempty)
        if hits >= 2 and len(nonempty) >= 2:
            return r, vals
        if best is None and len(nonempty) >= 3:
            best = (r, vals)
    return best or (1, [str(c.value or "") for c in ws[1]])


def assess_xlsx(b: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(b), read_only=False, data_only=True)
    info: dict = {"sheets": wb.sheetnames, "n_sheets": len(wb.sheetnames)}
    ws = wb[wb.sheetnames[0]]
    info["dims"] = f"{ws.max_row}x{ws.max_column}"
    info["merged"] = len(ws.merged_cells.ranges)
    hrow, headers = find_header(ws)
    headers = [h for h in headers if h]
    info["header_row"] = hrow
    info["headers"] = headers
    sup_col = next((h for h in headers if SUP_RE.search(h)), None)
    amt_col = next((h for h in headers if AMT_RE.search(h)), None)
    info["supplier_col"] = sup_col
    info["amount_col"] = amt_col
    # data rows + total rows + amount-as-text check
    data_rows = total_rows = amt_text = 0
    suppliers: list[str] = []
    amt_idx = headers.index(amt_col) if amt_col in headers else None
    sup_idx = headers.index(sup_col) if sup_col in headers else None
    for r in range(hrow + 1, (ws.max_row or hrow) + 1):
        cells = [c.value for c in ws[r]]
        nonempty = [c for c in cells if c not in (None, "")]
        if not nonempty:
            continue
        first = str(cells[0]) if cells and cells[0] is not None else ""
        if TOTAL_RE.search(first) or (len(nonempty) <= 2 and any(isinstance(c, (int, float)) for c in nonempty)):
            total_rows += 1
            continue
        data_rows += 1
        if sup_idx is not None and sup_idx < len(cells) and cells[sup_idx]:
            suppliers.append(str(cells[sup_idx]))
        if amt_idx is not None and amt_idx < len(cells):
            v = cells[amt_idx]
            if isinstance(v, str) and re.search(r"\d", v):
                amt_text += 1
    info["data_rows"] = data_rows
    info["total_rows"] = total_rows
    info["amount_as_text"] = amt_text
    info["suppliers"] = suppliers
    wb.close()
    return info


def schema_sig(headers: list[str]) -> str:
    """Canonical-ish signature: sorted lowercased alnum header tokens."""
    norm = sorted({re.sub(r"[^a-z0-9]", "", h.lower()) for h in headers if h})
    return "|".join(t for t in norm if t)[:120]


def main() -> None:
    hr("DISCOVER EXCEL RESOURCES (CKAN spend datasets + off-catalog)")
    cat = discover_excel()
    print(f"CKAN xls/xlsx resources in spend datasets: {len(cat)}")
    print("  by format:", dict(Counter(f for _, f, _ in cat)))
    print("  by publisher:", dict(Counter(p for p, _, _ in cat).most_common(8)))
    targets = cat[:8] + OFFCATALOG
    n_xls = sum(1 for _, f, _ in targets if f == "xls")
    print(f"\nsampling {len(targets)} workbooks ({n_xls} legacy .xls)")

    cro = pl.read_parquet(CRO).select(["name_norm", "company_num"])
    schemas: Counter = Counter()
    assessed = []
    all_suppliers: list[str] = []

    for i, (pub, fmt, url) in enumerate(targets):
        print(f"\n• [{pub}] {url.rsplit('/', 1)[-1][:48]}")
        if fmt == "xls":
            print("    LEGACY .xls -> needs xlrd (NOT installed) — FLAGGED, not parsed. Real gap.")
            assessed.append({"pub": pub, "fmt": "xls", "ok": False})
            continue
        b = fetch(url, i)
        if not b:
            continue
        try:
            info = assess_xlsx(b)
        except Exception as e:
            print(f"    openpyxl ERR {type(e).__name__}: {str(e)[:60]}")
            continue
        sig = schema_sig(info["headers"])
        schemas[sig] += 1
        landmines = []
        if info["n_sheets"] > 1:
            landmines.append(f"{info['n_sheets']} sheets")
        if info["merged"]:
            landmines.append(f"{info['merged']} merged-cells")
        if info["header_row"] > 1:
            landmines.append(f"header@row{info['header_row']}")
        if info["total_rows"]:
            landmines.append(f"{info['total_rows']} total-rows")
        if info["amount_as_text"]:
            landmines.append(f"{info['amount_as_text']} amt-as-text")
        print(f"    dims={info['dims']}  data_rows={info['data_rows']}  "
              f"supplier={info['supplier_col']!r}  amount={info['amount_col']!r}")
        print(f"    headers: {info['headers'][:8]}")
        print(f"    landmines: {landmines or ['none — clean table']}")
        assessed.append({"pub": pub, "fmt": "xlsx", "ok": True, **info})
        all_suppliers += info["suppliers"]

    hr("SCHEMA HETEROGENEITY (the core normalisation cost)")
    print(f"distinct column schemas across sampled xlsx: {len(schemas)}")
    for sig, n in schemas.most_common():
        print(f"  {n}x  {sig[:90]}")

    if all_suppliers:
        hr("SUPPLIER -> CRO (sampled xlsx)")
        sdf = (pl.DataFrame({"raw": all_suppliers})
               .with_columns(name_norm_expr("raw").alias("nn"))
               .filter(pl.col("nn").str.len_chars() >= 4).unique(subset=["nn"]))
        m = sdf.join(cro, left_on="nn", right_on="name_norm", how="left")
        hit = m.filter(pl.col("company_num").is_not_null()).select("nn").n_unique()
        print(f"distinct suppliers: {sdf.height}   CRO exact-name 1:1: {hit} ({hit / max(1, sdf.height):.0%})")

    hr("TECH-DEBT VERDICT")
    ok = [a for a in assessed if a.get("ok")]
    xls_legacy = [a for a in assessed if a["fmt"] == "xls"]
    clean = [a for a in ok if a.get("header_row") == 1 and a.get("n_sheets") == 1 and not a.get("merged")]
    print(f"workbooks assessed: {len(assessed)}  |  parsed xlsx: {len(ok)}  |  legacy .xls (blocked): {len(xls_legacy)}")
    print(f"  clean single-sheet/header@1/no-merge: {len(clean)}/{len(ok)}")
    print(f"  distinct schemas: {len(schemas)}  -> roughly 1 bespoke parser config per schema")
    print("\nDEBT DRIVERS (rank): (1) per-publisher SCHEMA divergence = N column maps to maintain;")
    print(" (2) legacy .xls needs xlrd (add dep) OR convert; (3) Excel report-formatting")
    print(" (title rows / merged cells / total rows) needs header-detection + total-row drop;")
    print(" (4) amount-as-text cells need cleaning. NONE are blockers — all are KNOWN-pattern")
    print(" work. The lever that caps debt: a SCHEMA-MAP table (publisher -> {supplier,amount})")
    print(" + a generic reader (find header, drop totals, coerce amount). Cost scales with #")
    print(" of distinct schemas, NOT # of files. Size THAT before committing.")


if __name__ == "__main__":
    main()
