"""PROBE (throwaway): MEASURE COVERAGE — actually parse Excel vs PDF (vs CSV) across
councils and compare what we really get. Still EVALUATION (no ETL wired).

For a representative council per format it: gathers a couple of real quarterly files
(reusing the seed probe's fetch+curl+one-hop-crawl), fully parses them, and reports
rows / €m / distinct suppliers / CRO 1:1 / time-span. Then it estimates national
coverage = (rows per file) x (files available, from the seed report) and compares the
formats head to head. Finally it sizes the remaining sources (CSV, CKAN-catalogued,
eTenders/TED already in gold).

Run:  ./.venv/Scripts/python.exe pipeline_sandbox/probe_procurement_coverage.py
Reads CRO silver + c:/tmp/procurement_la/seed_report.json; downloads samples to c:/tmp.
"""

from __future__ import annotations

import contextlib
import io
import json
import re
import sys
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
import polars as pl

ROOT = Path(__file__).resolve().parents[1]
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(HERE))
with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8")

from procurement_la_seed import (  # noqa: E402  (reuse robust fetch + crawl + filters)
    DATA_FILE_RE,
    NAV_HINT,
    POLICY_RE,
    extract_data_links,
    fetch_bytes,
    fetch_text,
)

from cro_normalise import name_norm_expr  # noqa: E402

CRO = ROOT / "data/silver/cro/companies.parquet"
REPORT = Path("c:/tmp/procurement_la/seed_report.json")
MONEY_RE = re.compile(r"(?:€|EUR)?\s?\d{1,3}(?:,\d{3})+(?:\.\d{2})?|\d+\.\d{2}")
NUM_RE = re.compile(r"\d[\d,]*(?:\.\d+)?")
SUP_RE = re.compile(r"supplier|payee|vendor|beneficiar|name|company|creditor", re.I)
AMT_RE = re.compile(r"amount|value|total|paid|gross|net|cost|euro|€|eur", re.I)
TOTAL_RE = re.compile(r"^\s*(grand\s+)?total|sub-?total", re.I)
YEAR_RE = re.compile(r"20[12]\d")


def hr(t: str) -> None:
    print(f"\n{'=' * 74}\n{t}\n{'=' * 74}")


def to_eur(tok: str) -> float:
    m = NUM_RE.search(tok or "")
    if not m:
        return 0.0
    with contextlib.suppress(ValueError):
        return float(m.group().replace(",", ""))
    return 0.0


# ---- gather a few real quarterly files per council ----
def gather(landing: str, ext: str, n: int = 2) -> list[str]:
    ext = ext if ext.startswith(".") else "." + ext  # extract_data_links keys with the dot
    html = fetch_text(landing)
    if not html:
        return []
    hits = extract_data_links(html, landing)
    if not hits.get(ext):  # one-hop crawl
        from urllib.parse import urljoin, urlparse
        host = urlparse(landing).netloc
        subs, seen = [], set()
        for href in re.findall(r"""href\s*=\s*["']([^"']+)["']""", html):
            full = urljoin(landing, href)
            is_file = any(full.lower().split("?")[0].endswith(e) for e in (".pdf", ".xlsx", ".xls", ".csv"))
            if (urlparse(full).netloc == host and full != landing and NAV_HINT.search(href)
                    and not is_file and full not in seen):
                seen.add(full)
                subs.append(full)
        for s in subs[:6]:
            sh = fetch_text(s)
            if sh:
                for e, v in extract_data_links(sh, s).items():
                    hits.setdefault(e, []).extend(v)
    good = [u for u in hits.get(ext, []) if DATA_FILE_RE.search(u) and not POLICY_RE.search(u)]
    # de-dupe keep order
    out, seen = [], set()
    for u in (good or hits.get(ext, [])):
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out[:n]


# ---- parsers: each returns list[(supplier, eur)] ----
def parse_xlsx(b: bytes) -> list[tuple[str, float]]:
    wb = openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = list(ws.iter_rows(values_only=True))
    wb.close()
    hrow = 0
    for i, row in enumerate(rows_iter[:12]):
        vals = [str(c).strip() if c is not None else "" for c in row]
        hits = sum(bool(SUP_RE.search(v) or AMT_RE.search(v)) for v in vals if v)
        if hits >= 2:
            hrow = i
            break
    headers = [str(c).strip() if c is not None else "" for c in rows_iter[hrow]]
    sup_i = next((j for j, h in enumerate(headers) if SUP_RE.search(h)), None)
    amt_i = next((j for j, h in enumerate(headers) if AMT_RE.search(h)), None)
    out: list[tuple[str, float]] = []
    data = rows_iter[hrow + 1:]
    # fallback for odd headers (e.g. Kilkenny supplier col = "Ap/Ar ID(T)"): score columns
    # by content — amount = most-numeric column, supplier = most-alphabetic (non-amount).
    if (sup_i is None or amt_i is None) and data:
        ncol = max((len(r) for r in data[:200]), default=0)
        nums = [0] * ncol
        txts = [0] * ncol
        for row in data[:200]:
            for j in range(min(ncol, len(row))):
                v = row[j]
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    nums[j] += 1
                elif isinstance(v, str) and re.search(r"[A-Za-z]", v):
                    txts[j] += 1
        if amt_i is None and any(nums):
            amt_i = max(range(ncol), key=lambda j: nums[j])
        if sup_i is None and any(txts):
            sup_i = max((j for j in range(ncol) if j != amt_i), key=lambda j: txts[j])
    if sup_i is None or amt_i is None:
        return out
    for row in data:
        if sup_i >= len(row) or amt_i >= len(row):
            continue
        sup = row[sup_i]
        amt = row[amt_i]
        if not sup or TOTAL_RE.search(str(sup)):
            continue
        eur = float(amt) if isinstance(amt, (int, float)) else to_eur(str(amt))
        if len(str(sup).strip()) >= 3 and eur >= 1000:
            out.append((str(sup).strip(), eur))
    return out


def parse_csv(b: bytes) -> list[tuple[str, float]]:
    df = pl.read_csv(io.BytesIO(b), infer_schema_length=0, truncate_ragged_lines=True,
                     ignore_errors=True, encoding="utf8-lossy")
    df = df.rename({c: c.replace("﻿", "").strip() for c in df.columns})
    sup = next((c for c in df.columns if SUP_RE.search(c)), None)
    amt = next((c for c in df.columns if AMT_RE.search(c)), None)
    if not sup or not amt:
        return []
    out = []
    for s, a in zip(df[sup].to_list(), df[amt].to_list(), strict=False):
        if s and not TOTAL_RE.search(str(s)):
            eur = to_eur(str(a))
            if len(str(s).strip()) >= 3 and eur >= 1000:
                out.append((str(s).strip(), eur))
    return out


def parse_pdf(b: bytes) -> list[tuple[str, float]]:
    doc = fitz.open(stream=b, filetype="pdf")
    out: list[tuple[str, float]] = []
    for page in doc:
        words = page.get_text("words")
        words.sort(key=lambda w: (round(w[1] / 3.0), w[0]))
        rows, cur, cy = [], [], None
        for w in words:
            if cy is None or abs(w[1] - cy) <= 3.0:
                cur.append(w)
                cy = w[1] if cy is None else cy
            else:
                rows.append(cur)
                cur, cy = [w], w[1]
        if cur:
            rows.append(cur)
        for wr in rows:
            ws = sorted(wr, key=lambda w: w[0])
            mi = [i for i, w in enumerate(ws) if MONEY_RE.search(w[4])]
            if not mi:
                continue
            ai = max(mi, key=lambda i: ws[i][0])
            eur = to_eur(ws[ai][4])
            rest = [w for i, w in enumerate(ws) if i != ai and not MONEY_RE.search(w[4])]
            if not rest or eur < 1000:
                continue
            if len(rest) >= 2:
                gap, cut = max((rest[i + 1][0] - rest[i][2], i) for i in range(len(rest) - 1))
                if gap < 12:
                    cut = len(rest) - 1
            else:
                cut = 0
            sup = " ".join(w[4] for w in rest[:cut + 1]).strip(" -:|")
            sup = re.sub(r"^(?:\d{3,}\s+){1,2}", "", sup).strip(" -:|")  # strip PO#/ID prefix
            if len(sup) >= 3:
                out.append((sup, eur))
    doc.close()
    return out


PARSERS = {"xlsx": parse_xlsx, "csv": parse_csv, "pdf": parse_pdf}

# council, format, mode("harvest"|direct list), ref
TARGETS: list[tuple[str, str, str, object]] = [
    ("South Dublin", "xlsx", "harvest", "https://www.sdcc.ie/en/services/business/payments/"),
    ("Cork City", "xlsx", "harvest", "https://www.corkcity.ie/en/council-services/public-info/spending-and-revenue/"),
    ("Monaghan", "xlsx", "harvest", "https://monaghan.ie/finance/publication-of-purchase-orders/"),
    ("Kilkenny", "xlsx", "harvest", "https://kilkennycoco.ie/eng/services/finance/purchase-orders-over-%E2%82%AC20-000/"),
    ("Wicklow", "xlsx", "harvest", "https://www.wicklow.ie/Living/Your-Council/Finance/Procurement/Purchase-Orders-Over-20-000"),
    ("Wicklow(csv)", "csv", "harvest", "https://www.wicklow.ie/Living/Your-Council/Finance/Procurement/Purchase-Orders-Over-20-000"),
    ("Cork County", "pdf", "harvest", "https://www.corkcoco.ie/en/council/accessibility-maps-and-publications/purchase-orders-in-excess-of-eu20000"),
    ("Kildare", "pdf", "harvest", "https://kildarecoco.ie/YourCouncil/Publications/Finance/PurchaseOrdersover20000/"),
    ("Westmeath", "pdf", "harvest", "https://www.westmeathcoco.ie/en/ourservices/finance/procurement/purchaseorders/"),
    ("Longford", "pdf", "harvest", "https://www.longfordcoco.ie/services/finance/finance-documents/large-purchase-orders/"),
    ("Galway City", "pdf", "direct", [
        "https://www.galwaycity.ie/sites/default/files/2026-05/Qtr%201%202026_Purchase%20Orders%20over%20%E2%82%AC20k_0.pdf",
        "https://www.galwaycity.ie/sites/default/files/2025-10/Qtr%203%202025%20Purchase%20Order%20over%20%E2%82%AC20k.pdf"]),
    ("Galway County", "pdf", "harvest", "https://www.gaillimh.ie/en/finance/financial-publications/purchase-orders"),
    ("Limerick", "pdf", "direct", [
        "https://www.limerick.ie/sites/default/files/media/documents/2026-05/purchase-orders-over-eu20-000-quarter-1-2026.pdf"]),
    ("Mayo", "pdf", "direct", [
        "https://www.mayo.ie/getattachment/6a915b4f-ab78-4b69-8ba6-d07569222e03/attachment.aspx",
        "https://www.mayo.ie/getattachment/00705716-9095-410a-a719-9305e0a37f4f/attachment.aspx"]),
    ("Donegal", "pdf", "direct", [
        "https://www.donegalcoco.ie/media/h0flvm3b/2025.pdf",
        "https://www.donegalcoco.ie/media/b2aopuh2/2024.pdf"]),
    ("Clare", "pdf", "direct", [
        "https://www.clarecoco.ie/sites/default/files/2025-08/purchase-orders-over-20-000-in-the-2nd-quarter-of-2025-58204.pdf"]),
    # ---- the rest (parse the remaining obtainable councils) ----
    ("Wexford", "xlsx", "harvest", "https://www.wexfordcoco.ie/council-and-democracy/procurement-finance-and-credit-control/council-spend"),
    ("Waterford", "pdf", "harvest", "https://waterfordcouncil.ie/openness-transparency/governance-related-financial-information/procurement/purchase-orders-e20000/"),
    ("Offaly", "pdf", "harvest", "https://www.offaly.ie/financial-reports/"),
    ("Kerry", "pdf", "harvest", "https://www.kerrycoco.ie/finance/financial-documents/"),
    ("Meath", "xlsx", "harvest", "https://www.meath.ie/council/your-council/finance-and-procurement/tenders-and-contracts/payments-over-eu20000"),
    ("Sligo", "pdf", "harvest", "https://www.sligococo.ie/YourCouncil/Finance/ProcurementPurchasing/PurchasingActivity/"),
    ("Leitrim", "pdf", "harvest", "https://www.leitrim.ie/council/services/finance/accounts-payable/purchase-to-pay/"),
    ("Laois", "pdf", "harvest", "https://laois.ie/finance/business-and-enterprise-support/procurement-information-and-advice"),
    ("Fingal", "pdf", "harvest", "https://www.fingal.ie/council/service/procurement"),
    ("Louth", "pdf", "harvest", "https://www.louthcoco.ie/en/publications/finance_reports/"),
    ("Tipperary", "pdf", "harvest", "https://www.tipperarycoco.ie/finance/financial-reports"),
    # JS-rendered (expected to skip → documents the Playwright gap)
    ("Carlow", "pdf", "harvest", "https://carlow.ie/information-technology/statistics-and-reports/financial-statistical-reports"),
    ("Cavan", "pdf", "harvest", "https://www.cavancoco.ie/file-library/business/procurement/over-20k/"),
    ("Roscommon", "pdf", "harvest", "https://www.roscommoncoco.ie/en/Download-It/Finance-Publications/"),
]


def main() -> None:
    cro = pl.read_parquet(CRO).select(["name_norm", "company_num"])
    counts = {}
    if REPORT.exists():
        for d in json.loads(REPORT.read_text(encoding="utf-8")):
            counts[d["council"]] = sum((d.get("formats") or {}).values())

    rows_out = []
    for council, fmt, mode, ref in TARGETS:
        urls = ref if mode == "direct" else gather(ref, fmt, n=2)
        recs: list[tuple[str, float]] = []
        got = 0
        years: set[str] = set()
        for u in urls:
            b = fetch_bytes(u)
            if not b:
                continue
            years |= set(YEAR_RE.findall(u))
            with contextlib.suppress(Exception):
                r = PARSERS[fmt](b)
                if r:
                    recs += r
                    got += 1
        if not recs:
            print(f"  [skip] {council:<16} {fmt:<4} — no rows (urls tried: {len(urls)})")
            rows_out.append((council, fmt, 0, 0, 0.0, 0.0, "", 0))
            continue
        cdf = pl.DataFrame(recs, schema=["supplier", "eur"], orient="row")
        total = cdf["eur"].sum()
        sup = (cdf.select("supplier").with_columns(name_norm_expr("supplier").alias("nn"))
               .filter(pl.col("nn").str.len_chars() >= 4).unique(subset=["nn"]))
        m = sup.join(cro, left_on="nn", right_on="name_norm", how="left")
        hit = m.filter(pl.col("company_num").is_not_null()).select("nn").n_unique()
        rate = hit / max(1, sup.height)
        per_file = cdf.height / max(1, got)
        avail = counts.get(council.replace("(csv)", ""), 0)
        est = int(per_file * avail) if avail else 0
        span = f"{min(years)}–{max(years)}" if years else "?"
        rows_out.append((council, fmt, cdf.height, got, total / 1e6, rate, span, est))
        print(f"  {council:<16} {fmt:<4} files={got}  rows={cdf.height:<5} €{total / 1e6:>7.1f}m  "
              f"suppliers={sup.height:<4} CRO={rate:.0%}  span={span}  ~total={est:,}")

    hr("FORMAT COMPARISON (sampled)")
    print(f"{'format':<8}{'councils':>9}{'rows':>8}{'€m':>9}{'avg CRO':>9}")
    for fmt in ("xlsx", "csv", "pdf"):
        g = [r for r in rows_out if r[1] == fmt and r[2] > 0]
        if not g:
            continue
        nr = sum(r[2] for r in g)
        em = sum(r[4] for r in g)
        cro = sum(r[5] for r in g) / len(g)
        print(f"{fmt:<8}{len(g):>9}{nr:>8,}{em:>8.1f}{cro:>9.0%}")

    hr("COVERAGE ESTIMATE (rows/file × files available per council)")
    est_total = sum(r[7] for r in rows_out)
    measured = sum(r[2] for r in rows_out)
    print(f"councils sampled: {sum(1 for r in rows_out if r[2] > 0)}  |  PO rows parsed now: {measured:,}")
    print(f"estimated rows across SAMPLED councils' full archives: ~{est_total:,}")
    print("(national ≈ this × ~2, scaling the ~16 sampled to ~27 obtainable councils)")

    hr("REMAINING DATA SOURCES (already measured elsewhere — not re-parsed here)")
    print("  CSV (council)   : Wicklow ships csv alongside xlsx (measured above); few others.")
    print("  CKAN-catalogued : Kilkenny (67) on data.gov.ie; Dept Housing (29, central) — same")
    print("                    PO/payments grain, tabular, CC-BY (probe_procurement_pdf.py census).")
    print("  eTenders awards : SHIPPED to gold — 100,106 notices / 40,474 awards (CEILINGS).")
    print("  TED API         : 8,230 IE award notices w/ REAL values (zero-auth; not yet pulled).")
    print("  => LA spend (this) is the per-transaction layer; eTenders/TED the award layer.")


if __name__ == "__main__":
    main()
