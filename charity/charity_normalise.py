#!/usr/bin/env python3
"""
Charities Public Register normaliser.

Reads the most recent public_register_*.xlsx in data/bronze/charities/ and
writes four silver parquets at data/silver/charities/.

Implements the charity half of CRO/INTEGRATION_PLAN.md §10 (NORMALISE per source).

INPUTS:
- data/bronze/charities/public_register_<DATE>.xlsx
  (Sheet "Disclaimer", "Public Register", "Annual Reports". Row 1 is the
  effective-date metadata band; row 2 is the column headers.)

OUTPUTS (silver):
- data/silver/charities/register.parquet         (one row per charity, RCN PK)
- data/silver/charities/annual_reports.parquet   (one row per (RCN, period_end))
- data/silver/charities/charity_latest.parquet   (one row per RCN — latest period)
- data/silver/charities/trustees_long.parquet    (one row per parsed trustee token)

WHY OPENPYXL DIRECTLY:
- pl.read_excel(engine="openpyxl") chokes on the Effective-Date metadata in
  row 1 (Series name must be a string). xlsx2csv / fastexcel are not pinned in
  this venv. Reading via openpyxl is deterministic and the file is small.
- read_only=True + values_only=True is fast (~15s for both sheets). The xlsx
  contains many trailing empty rows; we break on the first all-null row.

CLEANS:
- "CRO Number" = 0 → null (placeholder)
- "CRO Number" = RCN value → null (self-reference data-quality issue)
- Strip whitespace from all string columns
- Parse "Charity Classification: Primary [Secondary (Sub)]" into three columns
- Heuristic county = comma-split address, second-from-last token; "Unknown"
  when ambiguous
- Trustees (Start Date) parsed via regex into long table; unparseable tokens
  preserved with parse_quality='raw'

DERIVES:
- name_norm / aka_norm: same rule as cro_normalise.py (intentionally duplicated;
  the project's shared normalise_join_key.py is for TD names, not corporate)
- period_year: from Period End Date
- gov_share: (gov_or_la_income + other_public_bodies_income) / gross_income,
  null when gross_income is null/0
- amount_implausible_flag: within-entity sanity gate on annual_reports. The
  source register carries filer data-entry errors where one return's gross
  income/expenditure is orders of magnitude beyond the same charity's every
  other filing (e.g. RCN 20026691 gross_expenditure = €299bn for 2024). Flagged
  True when, for a charity with >=3 filings, gross_income or gross_expenditure is
  both >=€100m and >=50x that charity's own median for the field. The raw value
  is PRESERVED (faithful extraction); charity_latest excludes flagged filings so
  a single bad cell never becomes the headline figure or skews the income trend.
- trustee_count: count of parsed trustees per RCN, folded onto register.parquet
- charity_latest is a per-RCN profile: a latest-filing snapshot PLUS a
  multi-year trajectory drawn from the whole annual-reports time series.
  TRAJECTORY (all filed years):
    years_filed, first/last_period_year, deficit_years_count
    income_change_pct + income_trend (growing|flat|shrinking|insufficient_data;
      first-vs-last gross income over income-bearing years, ±20% band, ≥3 yrs)
  COMPOSITION (latest filing):
    share_<government|other_public|philanthropic|donations|trading|other|
      bequests>: each income stream as a share of gross income
    dominant_income_source: argmax label over the seven streams
    funding_profile: state_funded | mostly_donations | mostly_trading | mixed
      | undisclosed
  FINANCIAL HEALTH (latest filing):
    reserves_months: net_assets / gross_expenditure × 12, capped to [-24, 120]
    reserves_band: thin (<3m) | adequate (3-12m) | strong (>12m) | unknown
  SCALE (latest filing):
    employees_band_latest / volunteers_band_latest: text band, never numeric;
      nulled if not a recognised band
    employees_ft_latest / employees_pt_latest: numeric, post-2024 filings only
  DESCRIPTIVE (latest filing):
    beneficiary_tags: cleaned list parsed from the Beneficiaries tag string
    report_activity_latest
- state_adjacent_flag: gov_share >= 0.80 AND gross_income >= 100_000_000
- Warning flags surfaced on charity_latest for the lobbyist POC view:
    charity_filing_overdue_flag    period_end_latest < today − 18m
    charity_deficit_latest_flag    surplus_deficit_latest < 0
    charity_insolvent_latest_flag  total_liabilities_latest > total_assets_latest
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook

from config import BRONZE_DIR, SILVER_DIR
from services.parquet_io import save_parquet
from shared.name_norm import name_norm_expr

BRONZE_CHARITY_DIR = BRONZE_DIR / "charities"
DEFAULT_SILVER_DIR = SILVER_DIR / "charities"


def latest_bronze_xlsx() -> Path:
    """Pick the most recent public_register_*.xlsx in the bronze charity dir.

    Files are filename-dated (public_register_YYYYMMDD.xlsx) so lexical sort is
    chronological. Raises SystemExit if nothing matches.
    """
    candidates = sorted(BRONZE_CHARITY_DIR.glob("public_register_*.xlsx"))
    if not candidates:
        raise SystemExit(f"no public_register_*.xlsx files in {BRONZE_CHARITY_DIR}")
    return candidates[-1]



REGISTER_RENAME = {
    "Registered Charity Number": "rcn",
    "Registered Charity Name": "registered_charity_name",
    "Also Known As": "also_known_as",
    "Status": "status",
    "Charity Classification: Primary [Secondary (Sub)]": "classification_raw",
    "Primary Address": "primary_address",
    "Also Operates In": "also_operates_in",
    "Governing Form": "governing_form",
    "CRO Number": "cro_number_raw",
    "Country Established": "country_established",
    "Charitable Purpose": "charitable_purpose",
    "Charitable Objects": "charitable_objects",
    "Trustees (Start Date)": "trustees_raw",
}

ANNUAL_RENAME = {
    "Registered Charity Number (RCN)": "rcn",
    "Registered Charity Name": "registered_charity_name",
    "Period Start Date": "period_start_date",
    "Period End Date": "period_end_date",
    "Report Activity": "report_activity",
    "Activity Description": "activity_description",
    "Beneficiaries": "beneficiaries",
    "Income: Government or Local Authorities": "income_govt_or_la",
    "Income: Other Public Bodies": "income_other_public_bodies",
    "Income: Philantrophic Organisations": "income_philanthropic_orgs",
    "Income: Donations": "income_donations",
    "Income: Trading and Commercial Activities": "income_trading",
    "Income: Other Sources": "income_other",
    "Income: Bequests": "income_bequests",
    "Gross Income": "gross_income",
    "Gross Expenditure": "gross_expenditure",
    "Surplus / (Deficit) for the Period": "surplus_deficit",
    "Cash at Hand and in Bank": "cash_at_hand",
    "Other Assets": "other_assets",
    "Total Assets": "total_assets",
    "Total Liabilities": "total_liabilities",
    "Net Assets / (Liabilities)": "net_assets",
    "Gross Income (Schools)": "gross_income_schools",
    "Gross Expenditure (Schools)": "gross_expenditure_schools",
    "Number of Employees": "employees_band",
    "Number of Full-Time Employees": "employees_full_time",
    "Number of Part-Time Employees": "employees_part_time",
    "Number of Volunteers": "volunteers_band",
}

EMPLOYEES_BAND_ORDER = [
    "NONE",
    "1-9",
    "10-19",
    "20-49",
    "50-249",
    "250-499",
    "500-999",
    "1000-4999",
    "5000+",
]

# Valid employee / volunteer bands seen in the source, including the legacy
# "250+" band the regulator used in early filing years. Anything else (blank,
# stray text) is nulled.
VALID_BANDS = set(EMPLOYEES_BAND_ORDER) | {"250+"}

# Annual-report income streams → short label used for the share_* columns and
# dominant_income_source. Order is the tie-break priority for the argmax.
INCOME_STREAMS: dict[str, str] = {
    "income_govt_or_la": "government",
    "income_other_public_bodies": "other_public",
    "income_philanthropic_orgs": "philanthropic",
    "income_donations": "donations",
    "income_trading": "trading",
    "income_other": "other",
    "income_bequests": "bequests",
}

CLASSIFICATION_PATTERN = re.compile(
    r"^(?P<primary>[^\[]+?)\s*"
    r"(?:\[(?P<secondary>[^\(\]]+?)\s*"
    r"(?:\((?P<sub>[^\)]+)\))?\s*\])?\s*$"
)

TRUSTEE_PATTERN = re.compile(
    r"^(?P<name>.+?)"
    r"(?:\s*\((?P<role>[^)]+?)\))?"
    r"\s*\((?P<dt>\d{1,2}/\d{1,2}/\d{4})\)\s*$"
)


# ---------------------------------------------------------------------------
# IO
# ---------------------------------------------------------------------------


def read_sheet(path: Path, sheet: str) -> dict[str, list[Any]]:
    """Read one sheet via openpyxl, returning a columnar dict ready for pl.DataFrame.

    Row 1 of these files is the effective-date metadata band; row 2 is the
    real column headers. Trailing empty rows (the regulator's xlsx ships with
    many) are cut off when both of the first two cells are null.

    Cell values are coerced per-column so Polars's schema inference doesn't
    trip over xlsx int/float ambiguity in money columns:
    - any column whose only non-null types are int/float is normalised to float
    - everything else is left as-is (strings, datetimes, None)
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"sheet {sheet!r} not in {path.name}: {wb.sheetnames}")
    ws = wb[sheet]
    iterator = ws.iter_rows(min_row=2, values_only=True)
    headers = next(iterator)
    headers = [h.replace("\n", " ").strip() if isinstance(h, str) else None for h in headers]
    columns: dict[str, list[Any]] = {h: [] for h in headers if h}
    keep = [(i, h) for i, h in enumerate(headers) if h]
    for r in iterator:
        if r[0] is None and r[1] is None:
            break
        for i, h in keep:
            columns[h].append(r[i] if i < len(r) else None)
    wb.close()

    for h, values in columns.items():
        non_null_types = {type(v) for v in values if v is not None}
        if non_null_types and non_null_types.issubset({int, float}):
            columns[h] = [float(v) if v is not None else None for v in values]
    return columns


# ---------------------------------------------------------------------------
# Expression helpers
# ---------------------------------------------------------------------------


# name_norm_expr is imported from shared.name_norm (the canonical company-name key).
# The local copy here OMITTED the NFD accent-fold, so "Tirlán" → "TIRL N" instead of
# "TIRLAN" and charities never joined their CRO / supplier record on an accented name
# (doc/ENTITY_CROSSWALK_ORG_DOSSIER_DESIGN.md §2). Using the shared rule fixes that and
# also adds the missing 'AND' connector strip.


def classification_split(col: str) -> list[pl.Expr]:
    pat = CLASSIFICATION_PATTERN.pattern
    return [
        pl.col(col).str.extract(pat, 1).str.strip_chars().alias("classification_primary"),
        pl.col(col).str.extract(pat, 2).str.strip_chars().alias("classification_secondary"),
        pl.col(col).str.extract(pat, 3).str.strip_chars().alias("classification_sub"),
    ]


def county_from_address_expr(col: str) -> pl.Expr:
    parts = pl.col(col).str.split(",").list.eval(pl.element().str.strip_chars())
    second_last = parts.list.get(-2, null_on_oob=True)
    return (
        pl.when(second_last.is_not_null() & (second_last != ""))
        .then(second_last)
        .otherwise(pl.lit("Unknown"))
        .alias("county")
    )


# ---------------------------------------------------------------------------
# Public register
# ---------------------------------------------------------------------------


def normalise_register(path: Path) -> pl.DataFrame:
    cols = read_sheet(path, "Public Register")
    missing = [c for c in REGISTER_RENAME if c not in cols]
    if missing:
        raise SystemExit(f"public register missing columns: {missing}")
    df = pl.DataFrame({REGISTER_RENAME[k]: cols[k] for k in REGISTER_RENAME}, strict=False)

    df = df.with_columns(
        pl.col("rcn").cast(pl.Int64, strict=False),
        pl.col("registered_charity_name").str.strip_chars(),
        pl.col("status").str.strip_chars(),
        pl.col("governing_form").str.strip_chars(),
        pl.col("primary_address").str.strip_chars(),
        pl.col("also_known_as").str.strip_chars(),
        pl.col("country_established").str.strip_chars(),
        pl.col("classification_raw").str.strip_chars(),
        pl.col("cro_number_raw").cast(pl.Int64, strict=False),
    )

    df = df.with_columns(
        pl.when((pl.col("cro_number_raw") == 0) | (pl.col("cro_number_raw") == pl.col("rcn")))
        .then(None)
        .otherwise(pl.col("cro_number_raw"))
        .alias("cro_number"),
        name_norm_expr("registered_charity_name").alias("name_norm"),
        name_norm_expr("also_known_as").alias("aka_norm"),
        *classification_split("classification_raw"),
        county_from_address_expr("primary_address"),
    )
    df = df.with_columns(
        pl.col("cro_number").is_not_null().alias("has_cro_number_flag"),
    )
    return df.drop("cro_number_raw")


# ---------------------------------------------------------------------------
# Annual reports
# ---------------------------------------------------------------------------


def _money(col: str) -> pl.Expr:
    return pl.col(col).cast(pl.Float64, strict=False)


# Within-entity plausibility gate for the two headline money fields. The
# Charities Regulator's public register contains filer data-entry errors where a
# single annual return carries a value many orders of magnitude beyond the same
# charity's every other filing (e.g. RCN 20026691 South West Mayo Development
# Company filed gross_expenditure = €299,304,304,680 for 2024, vs ~€3-5m every
# other year — the same row's gross_income is a correct €5m). These are faithful
# extractions of a wrong source cell, so we never rewrite the value; we only
# flag the row so it can be excluded from the latest-snapshot and trajectory
# metrics that feed gold/UI. A value is implausible when, across a charity with
# >=3 filings, it is both >=€100m in absolute terms AND >=50x that charity's own
# median for the field. HSE/HEA/Pobal-scale bodies grow smoothly (max/median
# ~1.5x) and are never flagged; the genuine garbage sits at 14,000x-2,800,000x.
IMPLAUSIBLE_RATIO = 50.0
IMPLAUSIBLE_FLOOR_EUR = 100_000_000.0

# Income-trend anchor floor. Many returns carry a placeholder near-zero gross
# income (€0.01, €10 etc.) for a dormant/first year; using one as the first
# anchor makes income_change_pct explode to billions of percent and mislabels
# the charity "growing" (e.g. RCN 20084417 read €0.01→€873k = +8.7bn%, when its
# real income went €972k→€873k = flat). A charity reporting <€1,000 gross income
# is effectively a nil return, so we anchor the trend on its first/last filing
# with at least this much income. Filings below the floor are untouched in the
# data — only excluded as trend anchors.
TREND_MIN_INCOME_EUR = 1_000.0


def _implausible(col: str, median_col: str) -> pl.Expr:
    return (
        (pl.col("_entity_filings") >= 3)
        & pl.col(col).is_not_null()
        & (pl.col(median_col) > 0)
        & (pl.col(col) >= IMPLAUSIBLE_FLOOR_EUR)
        & (pl.col(col) >= IMPLAUSIBLE_RATIO * pl.col(median_col))
    )


def add_implausible_flag(df: pl.DataFrame) -> pl.DataFrame:
    """Add the boolean amount_implausible_flag via the within-entity gate.

    Requires rcn, gross_income, gross_expenditure columns. Median is taken over
    each charity's own filing history; a lone extreme barely moves it, so a
    garbage year stands out as a huge multiple of its peers. Raw values are left
    untouched — only the flag column is added. See _implausible for the rule.
    """
    df = df.with_columns(
        pl.len().over("rcn").alias("_entity_filings"),
        pl.col("gross_income").median().over("rcn").alias("_gross_income_median"),
        pl.col("gross_expenditure").median().over("rcn").alias("_gross_expenditure_median"),
    )
    df = df.with_columns(
        (
            _implausible("gross_income", "_gross_income_median")
            | _implausible("gross_expenditure", "_gross_expenditure_median")
        ).alias("amount_implausible_flag")
    )
    return df.drop("_entity_filings", "_gross_income_median", "_gross_expenditure_median")


def normalise_annual_reports(path: Path) -> pl.DataFrame:
    cols = read_sheet(path, "Annual Reports")
    missing = [c for c in ANNUAL_RENAME if c not in cols]
    if missing:
        raise SystemExit(f"annual reports missing columns: {missing}")
    df = pl.DataFrame({ANNUAL_RENAME[k]: cols[k] for k in ANNUAL_RENAME}, strict=False)

    money_cols = [
        "income_govt_or_la",
        "income_other_public_bodies",
        "income_philanthropic_orgs",
        "income_donations",
        "income_trading",
        "income_other",
        "income_bequests",
        "gross_income",
        "gross_expenditure",
        "surplus_deficit",
        "cash_at_hand",
        "other_assets",
        "total_assets",
        "total_liabilities",
        "net_assets",
        "gross_income_schools",
        "gross_expenditure_schools",
    ]
    df = df.with_columns(
        pl.col("rcn").cast(pl.Int64, strict=False),
        *[_money(c).alias(c) for c in money_cols],
        pl.col("employees_band").cast(pl.Utf8, strict=False).str.strip_chars().alias("employees_band"),
        pl.col("volunteers_band").cast(pl.Utf8, strict=False).str.strip_chars().alias("volunteers_band"),
    )

    period_start = (
        pl.col("period_start_date").cast(pl.Date, strict=False)
        if df.schema.get("period_start_date") in (pl.Datetime, pl.Date)
        else pl.col("period_start_date").cast(pl.Utf8).str.to_date(strict=False)
    )
    period_end = (
        pl.col("period_end_date").cast(pl.Date, strict=False)
        if df.schema.get("period_end_date") in (pl.Datetime, pl.Date)
        else pl.col("period_end_date").cast(pl.Utf8).str.to_date(strict=False)
    )
    df = df.with_columns(
        period_start.alias("period_start_date"),
        period_end.alias("period_end_date"),
    )
    df = df.with_columns(
        pl.col("period_end_date").dt.year().alias("period_year"),
        pl.when(pl.col("gross_income").is_null() | (pl.col("gross_income") <= 0))
        .then(None)
        .otherwise(
            (pl.col("income_govt_or_la").fill_null(0) + pl.col("income_other_public_bodies").fill_null(0))
            / pl.col("gross_income")
        )
        .alias("gov_share"),
    )

    # Within-entity plausibility flag — flags filer data-entry errors in the
    # source register without rewriting the raw value (see add_implausible_flag).
    return add_implausible_flag(df)


# ---------------------------------------------------------------------------
# Charity latest snapshot
# ---------------------------------------------------------------------------


def _band_clean(col: str, alias: str) -> pl.Expr:
    """Pass a band column through only if it is a recognised band, else null."""
    return pl.when(pl.col(col).is_in(list(VALID_BANDS))).then(pl.col(col)).otherwise(None).alias(alias)


def compute_income_trend(annual: pl.DataFrame) -> pl.DataFrame:
    """Per-RCN income trend: first vs last gross income across income-bearing years.

    Requires rcn, period_year, gross_income. Returns rcn, income_change_pct,
    income_trend (growing|flat|shrinking|insufficient_data; ±20% band, ≥3 years).
    Anchors only on filings with a substantive income (>= TREND_MIN_INCOME_EUR)
    so a placeholder near-zero first year can't make income_change_pct explode to
    billions of percent and mislabel a flat charity "growing".
    """
    return (
        annual.filter(pl.col("gross_income").is_not_null() & (pl.col("gross_income") >= TREND_MIN_INCOME_EUR))
        .group_by("rcn")
        .agg(
            pl.col("gross_income").sort_by("period_year").first().alias("gi_first"),
            pl.col("gross_income").sort_by("period_year").last().alias("gi_last"),
            pl.col("period_year").drop_nulls().n_unique().alias("income_years"),
        )
        .with_columns(
            (pl.col("gi_last") / pl.col("gi_first") - 1.0).alias("income_change_pct"),
        )
        .with_columns(
            pl.when(pl.col("income_years") < 3)
            .then(pl.lit("insufficient_data"))
            .when(pl.col("income_change_pct") > 0.20)
            .then(pl.lit("growing"))
            .when(pl.col("income_change_pct") < -0.20)
            .then(pl.lit("shrinking"))
            .otherwise(pl.lit("flat"))
            .alias("income_trend"),
        )
        .select(["rcn", "income_change_pct", "income_trend"])
    )


def build_charity_latest(annual: pl.DataFrame) -> pl.DataFrame:
    """Per-RCN charity profile: latest-filing snapshot + multi-year trajectory.

    Trajectory metrics (trend, deficit frequency, filing span) are aggregates
    over the WHOLE annual-reports time series for each charity — not just the
    last filing. Composition, financial-health, scale and descriptive columns
    are taken from the single most-recent filing.
    """
    today = dt.date.today()

    # Exclude filings flagged as implausible (see _implausible) from every
    # derived metric: a single fat-fingered source cell must not become a
    # charity's headline figure in gold/UI, nor skew its income trend. The raw
    # rows remain in annual_reports.parquet; here we simply look past them and
    # fall back to the charity's most recent clean filing.
    annual = annual.filter(pl.col("amount_implausible_flag").not_())

    # ── Trajectory — aggregates over every filed year ───────────────────────
    traj = annual.group_by("rcn").agg(
        pl.col("period_year").drop_nulls().n_unique().alias("years_filed"),
        pl.col("period_year").min().alias("first_period_year"),
        pl.col("period_year").max().alias("last_period_year"),
        (pl.col("surplus_deficit") < 0).sum().cast(pl.Int32).alias("deficit_years_count"),
    )

    inc_traj = compute_income_trend(annual)

    # ── Latest filing per RCN ───────────────────────────────────────────────
    latest = (
        annual.drop_nulls(["period_end_date"])
        .sort(["rcn", "period_end_date"], descending=[False, True])
        .unique(subset=["rcn"], keep="first")
    )

    gi = pl.col("gross_income")

    # 7-way income composition — each stream as a share of gross income.
    share_exprs = [
        pl.when(gi.is_null() | (gi <= 0)).then(None).otherwise(pl.col(stream).fill_null(0) / gi).alias(f"share_{label}")
        for stream, label in INCOME_STREAMS.items()
    ]

    # Dominant income source — argmax label over the seven streams; ties break
    # by INCOME_STREAMS order. Null when the charity reports no positive income.
    max_income = pl.max_horizontal([pl.col(c).fill_null(0) for c in INCOME_STREAMS])
    dominant = pl.when(max_income <= 0).then(None)
    for stream, label in INCOME_STREAMS.items():
        dominant = dominant.when(pl.col(stream).fill_null(0) >= max_income).then(pl.lit(label))
    dominant = dominant.otherwise(None).alias("dominant_income_source")

    # Reserves runway — net assets expressed as months of expenditure. The raw
    # ratio is wildly noisy (max ~39M months); the stored value is capped, the
    # band is taken off the uncapped ratio.
    reserves_raw = (
        pl.when(pl.col("gross_expenditure").is_null() | (pl.col("gross_expenditure") <= 0))
        .then(None)
        .otherwise(pl.col("net_assets") / pl.col("gross_expenditure") * 12.0)
    )
    reserves_band = (
        pl.when(reserves_raw.is_null())
        .then(pl.lit("unknown"))
        .when(reserves_raw < 3)
        .then(pl.lit("thin"))
        .when(reserves_raw <= 12)
        .then(pl.lit("adequate"))
        .otherwise(pl.lit("strong"))
        .alias("reserves_band")
    )

    donation_share = pl.col("income_donations").fill_null(0) / gi
    trading_share = pl.col("income_trading").fill_null(0) / gi
    funding_profile = (
        pl.when(gi.is_null() | (gi <= 0))
        .then(pl.lit("undisclosed"))
        .when(pl.col("gov_share").is_null())
        .then(pl.lit("undisclosed"))
        .when(pl.col("gov_share") >= 0.5)
        .then(pl.lit("state_funded"))
        .when(donation_share >= 0.5)
        .then(pl.lit("mostly_donations"))
        .when(trading_share >= 0.5)
        .then(pl.lit("mostly_trading"))
        .otherwise(pl.lit("mixed"))
        .alias("funding_profile")
    )

    state_adjacent = ((pl.col("gov_share").fill_null(0) >= 0.80) & (gi.fill_null(0) >= 100_000_000)).alias(
        "state_adjacent_flag"
    )

    cutoff_18m = today - dt.timedelta(days=int(365 * 1.5))
    filing_overdue = (
        (pl.col("period_end_date") < pl.lit(cutoff_18m)).fill_null(False).alias("charity_filing_overdue_flag")
    )
    deficit = (pl.col("surplus_deficit") < 0).fill_null(False).alias("charity_deficit_latest_flag")
    insolvent = (
        pl.col("total_liabilities").is_not_null()
        & pl.col("total_assets").is_not_null()
        & (pl.col("total_liabilities") > pl.col("total_assets"))
    ).alias("charity_insolvent_latest_flag")

    # Beneficiaries — semicolon-delimited tag string → cleaned list.
    beneficiary_tags = (
        pl.col("beneficiaries")
        .str.split(";")
        .list.eval(pl.element().str.strip_chars())
        .list.eval(pl.element().filter(pl.element().str.len_chars() > 0))
        .alias("beneficiary_tags")
    )

    latest = latest.with_columns(
        *share_exprs,
        dominant,
        reserves_raw.clip(-24.0, 120.0).alias("reserves_months"),
        reserves_band,
        funding_profile,
        state_adjacent,
        filing_overdue,
        deficit,
        insolvent,
        beneficiary_tags,
        _band_clean("employees_band", "employees_band_latest"),
        _band_clean("volunteers_band", "volunteers_band_latest"),
    ).rename(
        {
            "gov_share": "gov_funded_share_latest",
            "gross_income": "gross_income_latest_eur",
            "gross_expenditure": "gross_expenditure_latest_eur",
            "period_end_date": "period_end_latest",
            "period_year": "period_year_latest",
            "surplus_deficit": "surplus_deficit_latest",
            "total_assets": "total_assets_latest_eur",
            "total_liabilities": "total_liabilities_latest_eur",
            "net_assets": "net_assets_latest_eur",
            "cash_at_hand": "cash_at_hand_latest_eur",
            "employees_full_time": "employees_ft_latest",
            "employees_part_time": "employees_pt_latest",
            "report_activity": "report_activity_latest",
        }
    )

    latest = latest.join(traj, on="rcn", how="left").join(inc_traj, on="rcn", how="left")

    return latest.select(
        [
            "rcn",
            "period_end_latest",
            "period_year_latest",
            "gross_income_latest_eur",
            "gross_expenditure_latest_eur",
            "gov_funded_share_latest",
            "surplus_deficit_latest",
            "total_assets_latest_eur",
            "total_liabilities_latest_eur",
            "net_assets_latest_eur",
            "cash_at_hand_latest_eur",
            "reserves_months",
            "reserves_band",
            "share_government",
            "share_other_public",
            "share_philanthropic",
            "share_donations",
            "share_trading",
            "share_other",
            "share_bequests",
            "dominant_income_source",
            "funding_profile",
            "employees_band_latest",
            "volunteers_band_latest",
            "employees_ft_latest",
            "employees_pt_latest",
            "beneficiary_tags",
            "report_activity_latest",
            "years_filed",
            "first_period_year",
            "last_period_year",
            "deficit_years_count",
            "income_change_pct",
            "income_trend",
            "state_adjacent_flag",
            "charity_filing_overdue_flag",
            "charity_deficit_latest_flag",
            "charity_insolvent_latest_flag",
        ]
    )


# ---------------------------------------------------------------------------
# Trustees long
# ---------------------------------------------------------------------------


def parse_trustees(raw: str | None, rcn: int | None) -> list[dict[str, Any]]:
    if not raw or rcn is None:
        return []
    out: list[dict[str, Any]] = []
    for token in raw.split(";"):
        t = re.sub(r"\s+", " ", token).strip()
        if not t:
            continue
        m = TRUSTEE_PATTERN.match(t)
        if m:
            out.append(
                {
                    "rcn": rcn,
                    "trustee_name": m["name"].strip(),
                    "role": (m["role"] or "").strip() or None,
                    "start_date_raw": m["dt"],
                    "parse_quality": "strict",
                    "raw_token": t,
                }
            )
        else:
            out.append(
                {
                    "rcn": rcn,
                    "trustee_name": None,
                    "role": None,
                    "start_date_raw": None,
                    "parse_quality": "raw",
                    "raw_token": t,
                }
            )
    return out


def build_trustees_long(register: pl.DataFrame) -> pl.DataFrame:
    rows: list[dict[str, Any]] = []
    for r in register.select(["rcn", "trustees_raw"]).iter_rows(named=True):
        rows.extend(parse_trustees(r["trustees_raw"], r["rcn"]))
    if not rows:
        return pl.DataFrame(
            schema={
                "rcn": pl.Int64,
                "trustee_name": pl.Utf8,
                "role": pl.Utf8,
                "start_date_raw": pl.Utf8,
                "parse_quality": pl.Utf8,
                "raw_token": pl.Utf8,
                "trustee_name_norm": pl.Utf8,
                "start_date": pl.Date,
            }
        )
    df = pl.from_dicts(rows)
    return df.with_columns(
        pl.col("start_date_raw").str.to_date("%d/%m/%Y", strict=False).alias("start_date"),
        name_norm_expr("trustee_name").alias("trustee_name_norm"),
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    p = argparse.ArgumentParser(description="Charities Public Register normaliser")
    p.add_argument(
        "--bronze", type=Path, default=None, help="bronze xlsx path; defaults to the most recent public_register_*.xlsx"
    )
    p.add_argument("--silver-dir", type=Path, default=DEFAULT_SILVER_DIR)
    args = p.parse_args()

    bronze_path = args.bronze or latest_bronze_xlsx()
    if not bronze_path.exists():
        raise SystemExit(f"bronze input not found: {bronze_path}")
    args.silver_dir.mkdir(parents=True, exist_ok=True)

    print(f"[charity_normalise] reading {bronze_path}")
    register = normalise_register(bronze_path)
    annual = normalise_annual_reports(bronze_path)
    trustees = build_trustees_long(register)

    # trustee_count is a register-level governance signal; fold it onto the
    # register so non-filing charities keep it too.
    trustee_count = trustees.group_by("rcn").agg(pl.len().alias("trustee_count"))
    register = register.join(trustee_count, on="rcn", how="left").with_columns(
        pl.col("trustee_count").fill_null(0).cast(pl.Int32)
    )

    latest = build_charity_latest(annual)

    out_register = args.silver_dir / "register.parquet"
    out_annual = args.silver_dir / "annual_reports.parquet"
    out_latest = args.silver_dir / "charity_latest.parquet"
    out_trustees = args.silver_dir / "trustees_long.parquet"
    save_parquet(register, out_register)
    save_parquet(annual, out_annual)
    save_parquet(latest, out_latest)
    save_parquet(trustees, out_trustees)

    print(f"[charity_normalise] wrote {out_register}        rows={register.height}  cols={register.width}")
    print(f"[charity_normalise] wrote {out_annual}   rows={annual.height}  cols={annual.width}")
    print(f"[charity_normalise] wrote {out_latest}   rows={latest.height}  cols={latest.width}")
    print(f"[charity_normalise] wrote {out_trustees}    rows={trustees.height}  cols={trustees.width}")

    implausible = int(annual["amount_implausible_flag"].sum())
    cro_present = int(register["has_cro_number_flag"].sum())
    deregistered = int(register.filter(pl.col("status").str.contains("(?i)deregister")).height)
    state_adj = int(latest["state_adjacent_flag"].sum())
    strict_trustees = int(trustees.filter(pl.col("parse_quality") == "strict").height)
    raw_trustees = int(trustees.filter(pl.col("parse_quality") == "raw").height)
    print("  annual_amount_implausible:", f"{implausible:,}")
    print("  register_with_cro_number:", f"{cro_present:,}")
    print("  register_deregistered:   ", f"{deregistered:,}")
    print("  charities_state_adjacent:", f"{state_adj:,}")
    print(f"  trustees_strict:         {strict_trustees:,}")
    print(f"  trustees_raw:            {raw_trustees:,}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
