"""READ-ONLY tractability probe for the two GENUINELY-NEW structured candidates in
doc/dail_tracker_public_record_intelligence_sources_for_claude.md (the "Public Record
Intelligence Sources" brief), as assessed in doc/PUBLIC_RECORD_SOURCES_REVIEW.md.

It does NOT write parquet, touch gold/silver, or modify pipeline.py. It only fetches a
couple of candidate URLs and reports content-type / size / shape so the review's
format-tractability scores rest on real HTTP evidence rather than the brief's prose.

Two candidates probed (the only two that are structured AND join to an existing entity
and are NOT already covered by the publishers_seed / Iris public-appointments work):

  1. NWRA ERDF beneficiaries XLSX  (recipient -> CRO supplier dimension; value_kind=grant)
     URL from doc/dail_tracker_second_pass_tangible_sources.md (concrete .xlsx links).
  2. membership.stateboards.ie      (public-body universe + current board membership;
     complements the appointment-EVENT Iris spine which has no "current roster" view).

Run (offline-safe — prints SKIPPED if no network):
  ./.venv/Scripts/python.exe pipeline_sandbox/probe_review_src1_stateboards_erdf.py
"""

from __future__ import annotations

import contextlib
import io
import sys

with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
except Exception:  # pragma: no cover
    requests = None  # type: ignore[assignment]

H = {"User-Agent": "Mozilla/5.0 (dail-tracker read-only review probe)"}

CANDIDATES = [
    # (label, url, kind)
    ("nwra_erdf_xlsx",
     "https://www.nwra.ie/wp-content/uploads/2025/11/2021-2027-beneficiaries-Oct-2025.xlsx",
     "xlsx"),
    ("stateboards_membership",
     "https://membership.stateboards.ie/",
     "html"),
    ("southern_erdf_landing",
     "https://www.southernassembly.ie/erdf/beneficiaries-21-27",
     "html"),
]


def probe_one(label: str, url: str, kind: str) -> None:
    print(f"\n=== {label} ({kind}) ===\n{url}")
    if requests is None:
        print("  SKIPPED: requests not importable")
        return
    try:
        r = requests.get(url, headers=H, timeout=30, allow_redirects=True)
    except Exception as e:  # network / SSL / timeout
        print(f"  SKIPPED (network): {type(e).__name__}: {e}")
        return
    ct = r.headers.get("Content-Type", "?")
    print(f"  status={r.status_code}  content-type={ct}  bytes={len(r.content)}")
    if r.status_code != 200:
        return
    if kind == "xlsx" and ("sheet" in ct or "excel" in ct or url.endswith(".xlsx")):
        try:
            import openpyxl  # noqa: PLC0415

            wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header = next(ws.iter_rows(values_only=True), None)
            print(f"  sheets={wb.sheetnames}")
            print(f"  first-row (header?) = {header}")
            n = sum(1 for _ in ws.iter_rows(values_only=True))
            print(f"  ~rows on sheet 1 = {n}")
        except Exception as e:
            print(f"  xlsx parse failed: {type(e).__name__}: {e}")
    elif kind == "html":
        body = r.text.lower()
        for kw in ("board", "vacanc", "beneficiar", "operation", "county", "amount",
                   "table", "download", ".xlsx", ".csv"):
            if kw in body:
                print(f"  html contains keyword: {kw!r}")


def main() -> None:
    print("READ-ONLY review probe — no parquet written, gold/silver untouched.")
    for label, url, kind in CANDIDATES:
        probe_one(label, url, kind)
    print("\nDone. Findings feed doc/PUBLIC_RECORD_SOURCES_REVIEW.md format-tractability scores.")


if __name__ == "__main__":
    main()
