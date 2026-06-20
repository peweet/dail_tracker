"""Derelict Sites Levy — per-local-authority enforcement & collection (gold).

The council-administered Derelict Sites Levy (Derelict Sites Act 1990): how much
each local authority LEVIED vs how much it actually COLLECTED, and the cumulative
amount left OUTSTANDING. The enforcement-gap dataset — many councils levy little and
collect less, so ~€26m sits uncollected nationally. Powers v_la_derelict_sites_levy
(the "Who runs your county" page) and v_constituency_council_housing_performance.

Source: Dept of Housing, LG & Heritage annual return (gov.ie), one consolidated
XLSX per year, per-LA. CC-BY-4.0 (Irish PSI / data.gov.ie licence).
  Landing: gov.ie/.../annual-returns-...-derelict-sites-act-1990/
  File   : assets.gov.ie/static/documents/866f4b20/2024_Derelict_Sites_Statistics.xlsx
The gov.ie CDN 403s a bare request — a browser User-Agent + gov.ie Referer are needed.

Reads  : doc/source_pdfs/2024_Derelict_Sites_Statistics.xlsx  (cached source, git-tracked)
Writes : data/gold/parquet/derelict_sites_levy_wide.parquet  (atomic, via services.parquet_io)

Promoted from pipeline_sandbox/housing/ (2026-06-20): routes through save_parquet,
writes by default, gated on the fidelity check (refuses to persist a non-GREEN parse).
Annual source — re-fetch each year with --download as a new year's file appears; the
freshness watch (tools/check_freshness.py → derelict_sites_levy) flags staleness.

Run:
  python extractors/derelict_sites_levy_extract.py              # parse cache + write gold
  python extractors/derelict_sites_levy_extract.py --download   # re-fetch from gov.ie first
  python extractors/derelict_sites_levy_extract.py --dry-run    # parse + report, no write
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from services.logging_setup import setup_standalone_logging  # noqa: E402
from services.parquet_io import save_parquet  # noqa: E402

LOG = logging.getLogger("derelict_sites_levy")
_SRC = ROOT / "doc" / "source_pdfs" / "2024_Derelict_Sites_Statistics.xlsx"
_OUT = ROOT / "data" / "gold" / "parquet" / "derelict_sites_levy_wide.parquet"
_URL = "https://assets.gov.ie/static/documents/866f4b20/2024_Derelict_Sites_Statistics.xlsx"
_YEAR = 2024

# XLSX column index -> output field (header row is row 2 in the sheet).
_COLS = {
    0: "la",
    1: "notices_issued",
    6: "sites_on_register_end",
    9: "sites_levied",
    10: "amount_levied_eur",
    11: "amount_received_levied_eur",
    12: "total_received_eur",
    13: "cumulative_outstanding_eur",
}


def _download() -> None:
    import requests

    h = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)", "Referer": "https://www.gov.ie/"}
    r = requests.get(_URL, headers=h, timeout=60)
    r.raise_for_status()
    _SRC.parent.mkdir(parents=True, exist_ok=True)
    _SRC.write_bytes(r.content)
    LOG.info("downloaded %d bytes -> %s", len(r.content), _SRC.relative_to(ROOT))


def _clean_la(s: str) -> str:
    n = re.sub(r"\s+", " ", str(s or "")).strip()
    n = n.replace("Dún", "Dun").replace(" & ", " and ")
    n = n.replace("Dun Laoghaire Rathdown", "Dun Laoghaire-Rathdown")
    return n


def _num(v):
    if v is None:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    if not s or s in {"-", ".", "-."}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract() -> tuple[pl.DataFrame, dict]:
    wb = load_workbook(_SRC, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    records, totals = [], {}
    for r in rows[2:]:
        if not r or r[0] is None:
            continue
        name = str(r[0]).strip()
        if name.lower() == "total":
            totals = {f: _num(r[i]) for i, f in _COLS.items() if i != 0}
            continue
        rec = {"year": _YEAR}
        for i, f in _COLS.items():
            rec[f] = _clean_la(r[i]) if f == "la" else _num(r[i])
        records.append(rec)
    df = pl.DataFrame(records)
    cols = ["la", "year"] + [f for f in _COLS.values() if f != "la"]
    return df.select(cols), totals


def fidelity_check(df: pl.DataFrame, totals: dict) -> dict:
    rpt = {"checks": {}, "rows": df.height}
    if df.is_empty():
        rpt["green"] = False
        return rpt
    rpt["checks"]["1_la_coverage"] = {"unique_LAs": df["la"].n_unique(), "pass": df["la"].n_unique() == 31}
    # the per-LA sums must reconcile to the file's own Total row (catches drops/dupes)
    recon = {}
    for f in ("amount_levied_eur", "total_received_eur", "cumulative_outstanding_eur"):
        got = df[f].sum()
        want = totals.get(f)
        recon[f] = {"sum": got, "file_total": want, "ok": want is not None and abs(got - want) < 1}
    rpt["checks"]["2_reconciles_to_total"] = {**recon, "pass": all(v["ok"] for v in recon.values())}
    bad = sum(df.filter(pl.col(f) < 0).height for f in ("amount_levied_eur", "cumulative_outstanding_eur"))
    rpt["checks"]["3_non_negative"] = {"negatives": bad, "pass": bad == 0}
    rpt["green"] = all(c.get("pass", False) for c in rpt["checks"].values())
    return rpt


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--download", action="store_true", help="re-fetch the XLSX from gov.ie first")
    ap.add_argument("--dry-run", action="store_true", help="parse + report, do NOT write")
    args = ap.parse_args()
    setup_standalone_logging("derelict_sites_levy")

    if args.download:
        _download()
    if not _SRC.exists():
        LOG.error("source missing: %s (run with --download)", _SRC)
        sys.exit(1)

    df, totals = extract()
    rpt = fidelity_check(df, totals)
    LOG.info("derelict_sites_levy_wide — %d rows", df.height)
    for n, chk in rpt["checks"].items():
        LOG.info("  [%s] %s: %s", "GREEN" if chk.get("pass") else "FAIL", n, chk)
    if not df.is_empty():
        LOG.info(
            "national: levied €%s | received €%s | outstanding €%s",
            f"{df['amount_levied_eur'].sum():,.0f}",
            f"{df['total_received_eur'].sum():,.0f}",
            f"{df['cumulative_outstanding_eur'].sum():,.0f}",
        )

    if args.dry_run:
        LOG.info("dry-run: not writing (parse %s)", "GREEN" if rpt["green"] else "AMBER")
        return
    if not rpt["green"]:
        LOG.error("fidelity AMBER — refusing to overwrite %s", _OUT.name)
        sys.exit(2)
    save_parquet(df, _OUT, min_rows=30)
    LOG.info("wrote %s (%d rows)", _OUT.relative_to(ROOT), df.height)


if __name__ == "__main__":
    main()
