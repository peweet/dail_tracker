"""eTenders tender documents -> a deterministic deep link to the ELIGIBILITY / SELECTION CRITERIA section.

SANDBOX. `pipeline_sandbox/` is the experiments tree: this is deliberately NOT in the
`pipeline.py` sequence and is run by hand. It writes ONE artifact:

    data/sandbox/parquet/etenders_criteria_pointer.parquet   (one row per notice WITH a landing)

WHAT IT DOES. A bidder's first question about a notice is "am I even eligible?" — and the answer
sits in one named section of one of the (often 15+) published documents. This locates that
section: which document, which heading, which line, and how much to trust the hit.

THE LOCATOR (ported from the validated scratchpad prototype `matcher_v2.py`, adversarially
audited over 37 landings: 24 true positive / 8 partial / 5 false positive). Five rules, each
motivated by a real mis-landing:
  1. document ranking   - the specification document (CFT/RFT/ITT) outranks the ESPD form, which
                          is a tick-box cross-reference and is the LAST-resort document class.
  2. pattern preference - selection_criteria > economic_financial_standing >
                          suitability_assessment > part_iv_espd, then the weaker added patterns.
  3. section_3_3 demoted- `3.3 ...` corroborates a landing, it can never BE one (it fires inside
                          award-scoring tables).
  4. added patterns     - technical/professional ability, Part III exclusion grounds, the
                          ampersand and reversed economic/financial variants, standalone Turnover.
  5. TOC avoidance      - a contents-page echo of a heading is never the landing point, and neither
                          is a mid-sentence fragment that merely mentions the words.

CHANGED FROM THE PROTOTYPE: `doc_class` tested the ESPD name pattern FIRST and unconditionally, so
a genuine specification whose filename merely mentions the ESPD ("RFT Volume 2 - ESPD Response.pdf")
scored 5 — below `noise`. The spec check now runs first. Pinned by
test_spec_check_runs_before_the_espd_check.

KNOWN DEFECT (carried across, not fixed): `is_toc_line`'s early-duplicate rule is gated on
`ln.idx / total < 0.20`, so a contents entry past the first fifth of a long line pool is not
recognised as an echo. Kept visible by a strict xfail in the test suite.

CONFIDENCE is a 3-level honest band, not a score — the audit measured reliability per pattern and
does not support a number:
    high   = pattern is `selection_criteria` AND the landing line carries a Word heading style
    medium = exactly one of those holds
    low    = neither

INPUT. Documents are read from a local cache laid out as ``<cache-dir>/<tender_id>/<document_id>__<name>``.
`--fetch` additionally harvests that cache from the anonymous eTenders endpoints first. The
download path MUST keep its `/cft/` segment — without it the server answers HTTP 500 (verified by
curl 2026-08-16). There is no rate limiting in `services.http_engine`; spacing is this caller's
job, hence `--delay-ms` (1500 default, 1000 floor) before EVERY request.

Run:
    ./.venv/Scripts/python.exe pipeline_sandbox/etenders_criteria_pointer.py --cache-dir <dir>
    ./.venv/Scripts/python.exe pipeline_sandbox/etenders_criteria_pointer.py --cache-dir <dir> \
        --listings <listings.jsonl> --dry-run
    ./.venv/Scripts/python.exe pipeline_sandbox/etenders_criteria_pointer.py --cache-dir <dir> \
        --fetch --tender-ids 3509643 6234281
"""

from __future__ import annotations

# isort: off
# Apply native thread caps before Polars/NumPy loads. Ordering is the contract.
import services.runtime_env  # noqa: F401
# isort: on

import argparse
import collections
import contextlib
import io
import json
import logging
import re
import sys
import time
import zipfile
from datetime import UTC, datetime
from pathlib import Path, PurePosixPath

import polars as pl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
with contextlib.suppress(Exception):
    reconfigure_stdout = getattr(sys.stdout, "reconfigure", None)
    if callable(reconfigure_stdout):
        reconfigure_stdout(encoding="utf-8")
from services.extract_runner import run_extractor  # noqa: E402
from services.http_engine import fetch_bytes, fetch_text, polite_headers  # noqa: E402
from services.parquet_io import save_parquet  # noqa: E402

log = logging.getLogger(__name__)

MATCHER_VERSION = "2.1.0"
BASE = "https://www.etenders.gov.ie"
# The `/cft/` segment is load-bearing on BOTH endpoints: dropping it returns HTTP 500,
# not a redirect (verified by curl 2026-08-16).
LIST_URL = f"{BASE}/epps/cft/listContractDocuments.do"
DOWNLOAD_URL = f"{BASE}/epps/cft/downloadContractDocument.do"
OUT_PARQUET = ROOT / "data/sandbox/parquet/etenders_criteria_pointer.parquet"

DELAY_MS_FLOOR = 1000
MAX_DOC_BYTES = 80 * 1024 * 1024
TEXTY_EXTS = {".pdf", ".docx", ".doc", ".zip", ".rtf", ".xlsx", ".xls", ".odt"}
HEADINGISH_MAX = 130


# ── patterns ─────────────────────────────────────────────────────────────────
# (tag, regex, role). role: "land" = may be a link target, "corrob" = confirmation only.
PATTERNS = [
    # --- preference order for landing, best first -------------------------
    ("selection_criteria", re.compile(r"selection\s+criteri(?:a|on)", re.I), "land"),
    (
        "economic_financial_standing",
        # ampersand AND reversed word order, correction 4
        re.compile(r"(?:economic|financial)\s*(?:and|&|\+)\s*(?:financial|economic)\s+standing", re.I),
        "land",
    ),
    ("suitability_assessment", re.compile(r"suitability\s+assessment", re.I), "land"),
    ("part_iv_espd", re.compile(r"\bpart\s+IV\b", re.I), "land"),
    # --- added patterns (correction 4), ranked below the four named above --
    ("technical_professional_ability", re.compile(r"technical\s+(?:and|&)\s+professional\s+abilit", re.I), "land"),
    ("financial_capacity", re.compile(r"financial\s+capacity", re.I), "land"),
    ("part_iii_exclusion", re.compile(r"\bpart\s+III\b\s*[:.\-–]?\s*exclusion|exclusion\s+grounds?\b", re.I), "land"),
    (
        "turnover",
        re.compile(
            r"^\s*(?:[\dIVXivx]+(?:[.\)][\da-z]+)*[.\)]?\s+)?"
            r"(?:minimum\s+|specific\s+|annual\s+|average\s+|evidence\s+of\s+)*turnover\b",
            re.I,
        ),
        "land",
    ),
    # --- corroboration only, never a landing target ------------------------
    ("section_3_3", re.compile(r"^\s*(?:section\s+)?3\.3(?:[a-h])?\b\s*[.\-–:)]?\s*\S", re.I), "corrob"),
    ("evaluation_criteria", re.compile(r"evaluation\s+criteri", re.I), "corrob"),
    ("qualification_and_award_criteria", re.compile(r"qualification\s+and\s+award\s+criteri", re.I), "corrob"),
    ("criteria_for_qualitative_selection", re.compile(r"criteria\s+for\s+qualitative\s+selection", re.I), "corrob"),
]
LAND_ORDER = [t for t, _, role in PATTERNS if role == "land"]
LAND_RANK = {t: i for i, t in enumerate(LAND_ORDER)}
PAT_BY_TAG = {t: rx for t, rx, _ in PATTERNS}

# The SUPERSEDED v1 pattern set and filename heuristic. Kept — and only ever read by the
# regression tests — because they are what the corrections above were measured against:
# v1's `and`-only regex and its top score for `espd` ARE the defects, and a test that can
# no longer state the defect can no longer prove the fix.
V1_STRICT = [
    ("suitability_assessment", re.compile(r"suitability\s+assessment", re.I)),
    ("section_3_3", re.compile(r"^\s*(?:section\s+)?3\.3(?:[a-h])?\b\s*[.\-–:)]?\s*\S", re.I)),
    ("part_iv_espd", re.compile(r"\bpart\s+IV\b", re.I)),
    ("selection_criteria", re.compile(r"selection\s+criteri(?:a|on)", re.I)),
    ("economic_financial_standing", re.compile(r"economic\s+and\s+financial\s+standing", re.I)),
]
V1_FN_PRIORITY = [
    (10, re.compile(r"suitab|selection\s*criteri|espd|qualification", re.I)),
    (9, re.compile(r"\brft\b|request[\s_-]*for[\s_-]*tender", re.I)),
    (9, re.compile(r"\bitt\b|invitation[\s_-]*to[\s_-]*tender", re.I)),
    (9, re.compile(r"\bcft\b|call[\s_-]*for[\s_-]*tender", re.I)),
]


# ── document ranking (correction 1) ──────────────────────────────────────────
ESPD_RX = re.compile(
    r"espd|electronic[\s_-]*european[\s_-]*single[\s_-]*procurement|european[\s_-]*single[\s_-]*procurement",
    re.I,
)
SPEC_RX = re.compile(
    r"\bcft\b|\brft\b|\bitt\b|\bsrft\b|"
    r"request[\s_-]*for[\s_-]*tenders?|call[\s_-]*for[\s_-]*tenders?|"
    r"invitation[\s_-]*to[\s_-]*tender|instructions?[\s_-]*to[\s_-]*tenderers?|"
    r"tender[\s_-]*document",
    re.I,
)
CRIT_RX = re.compile(
    r"suitab|selection\s*criteri|pre[\s_-]*qualification|qualification|evaluation\s*criteri|questionnaire", re.I
)
NOISE_RX = re.compile(
    r"renunciation|property\s+licence|licence\s+template|confidentiality|"
    r"data\s+processing|pricing\s+(?:schedule|document)|price\s+schedule|"
    r"terms\s+(?:and|&)\s+conditions|standard\s+terms|site\s+visit|"
    r"deed\s+of|appendix|annex",
    re.I,
)

DOC_CLASS_SCORE = {"spec": 100, "criteria": 90, "other": 50, "noise": 20, "espd": 5}

CONFIDENCE_PATTERN = "selection_criteria"


def member_name(source_label: str) -> str:
    """`bundle.zip!dir/Inner Doc.docx` -> `Inner Doc.docx`; plain names pass through."""
    s = source_label or ""
    if "!" in s:
        s = s.rsplit("!", 1)[1]
    return PurePosixPath(s.replace("\\", "/")).name


def doc_class(filename: str) -> tuple[str, int]:
    """Return (class_name, score). ESPD is deliberately the lowest class: correction 1.

    Underscores are word characters, so `RFT_Volume_1.pdf` defeats a `\\brft\\b` pattern.
    Normalise separators to spaces before classing.

    SPEC is tested BEFORE ESPD. The prototype tested ESPD first and unconditionally, which
    demoted a genuine specification whose filename merely MENTIONS the form (`RFT Volume 2 -
    ESPD Response.pdf`) to score 5 — below `noise` — i.e. exactly the mis-ranking correction 1
    exists to prevent. A document that is both is a specification.
    """
    n = re.sub(r"[_]+", " ", filename or "")
    if SPEC_RX.search(n):
        return "spec", DOC_CLASS_SCORE["spec"]
    if ESPD_RX.search(n):
        return "espd", DOC_CLASS_SCORE["espd"]
    if CRIT_RX.search(n):
        return "criteria", DOC_CLASS_SCORE["criteria"]
    if NOISE_RX.search(n):
        return "noise", DOC_CLASS_SCORE["noise"]
    return "other", DOC_CLASS_SCORE["other"]


def v1_fn_score(name: str) -> int:
    """The SUPERSEDED v1 filename heuristic — `espd` scored 10, above every specification."""
    for sc, rx in V1_FN_PRIORITY:
        if rx.search(name or ""):
            return sc
    return 0


# ── text extraction ──────────────────────────────────────────────────────────
# `docx` / `fitz` / `openpyxl` are imported INSIDE these functions on purpose: importing the
# module must stay free of the heavy optional stack (pinned by a test).
class Line:
    """One extracted line of a document, with the position and Word style that rank it."""

    __slots__ = ("idx", "kind", "source", "style", "text")

    def __init__(self, text: str, idx: int, style: str, source: str, kind: str) -> None:
        self.text, self.idx, self.style, self.source, self.kind = text, idx, style, source, kind

    def as_dict(self) -> dict:
        return {"text": self.text, "idx": self.idx, "style": self.style, "source": self.source, "kind": self.kind}


def _docx_lines(data: bytes, source: str) -> list[Line]:
    import docx
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    d = docx.Document(io.BytesIO(data))
    out: list[tuple[str, str, str]] = []

    def _style_of(p) -> str:
        try:
            return p.style.name or ""
        except Exception:  # noqa: BLE001 — a broken style reference must not lose the line
            return ""

    def walk(parent, elm) -> None:
        for child in elm.iterchildren():
            tag = child.tag.split("}")[-1]
            if tag == "p":
                text = Paragraph(child, parent).text.strip()
                if text:
                    out.append((text, _style_of(Paragraph(child, parent)), "para"))
            elif tag == "tbl":
                for row in Table(child, parent).rows:
                    seen: set[int] = set()
                    for cell in row.cells:
                        if id(cell._tc) in seen:  # merged cells repeat across the row
                            continue
                        seen.add(id(cell._tc))
                        for p in cell.paragraphs:
                            text = p.text.strip()
                            if text:
                                out.append((text, _style_of(p), "cell"))

    walk(d, d.element.body)
    return [Line(t, i, st, source, kind) for i, (t, st, kind) in enumerate(out)]


def _pdf_lines(data: bytes, source: str) -> tuple[list[Line], str]:
    import fitz

    doc = fitz.open(stream=data, filetype="pdf")
    out: list[str] = []
    nchars = 0
    for page in doc:
        try:
            text = page.get_text("text")
        except Exception:  # noqa: BLE001 — one unreadable page must not lose the document
            continue
        nchars += len(text)
        out.extend(ln.strip() for ln in text.split("\n") if ln.strip())
    npages = doc.page_count
    doc.close()
    # A scanned PDF yields a handful of chars per page; say so rather than report "no criteria".
    status = "scanned_or_empty" if npages and nchars / max(npages, 1) < 60 else "ok"
    return [Line(t, i, "", source, "pdf") for i, t in enumerate(out)], status


def _xlsx_lines(data: bytes, source: str) -> list[Line]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(str(ws.title))
        for n, row in enumerate(ws.iter_rows(values_only=True)):
            out.extend(v.strip() for v in row if isinstance(v, str) and v.strip())
            if n > 4000:
                break
    wb.close()
    return [Line(t, i, "", source, "cell") for i, t in enumerate(out)]


def _salvage_lines(data: bytes, source: str) -> list[Line]:
    """Legacy .doc / .rtf: pull printable runs out of the raw bytes. Lossy by construction."""
    raw = data.decode("latin-1", "ignore")
    chunks = re.findall(r"[\x20-\x7e]{6,}", raw)
    return [Line(c.strip(), i, "", source, "salvage") for i, c in enumerate(chunks) if c.strip()]


def extract_bytes(name: str, ext: str, data: bytes, depth: int = 0) -> list[tuple[str, list[Line], str]]:
    """-> [(source_label, lines, status)]. A .zip expands to one entry per text-bearing member."""
    if ext == ".zip":
        if depth > 2:
            return [(name, [], "zip_too_deep")]
        try:
            zf = zipfile.ZipFile(io.BytesIO(data))
        except Exception as exc:  # noqa: BLE001 — a truncated bundle is a coverage stat, not an abort
            return [(name, [], f"zip_err {type(exc).__name__}")]
        out: list[tuple[str, list[Line], str]] = []
        for member in [x for x in zf.namelist() if not x.endswith("/")][:60]:
            mext = Path(member).suffix.lower()
            if mext not in TEXTY_EXTS:
                continue
            try:
                inner = zf.read(member)
            except Exception:  # noqa: BLE001
                continue
            if len(inner) > MAX_DOC_BYTES:
                continue
            out.extend(extract_bytes(f"{name}!{member}", mext, inner, depth + 1))
        zf.close()
        return out or [(name, [], "zip_no_text_members")]
    try:
        if ext in (".docx", ".odt"):
            return [(name, _docx_lines(data, name), "ok")]
        if ext == ".pdf":
            lines, status = _pdf_lines(data, name)
            return [(name, lines, status)]
        if ext in (".xlsx", ".xls"):
            return [(name, _xlsx_lines(data, name), "ok")]
        if ext == ".doc":
            if data[:2] == b"PK":  # a .docx mis-named .doc
                return [(name, _docx_lines(data, name), "ok")]
            return [(name, _salvage_lines(data, name), "legacy_doc_salvage")]
        if ext == ".rtf":
            return [(name, _salvage_lines(data, name), "rtf_salvage")]
    except Exception as exc:  # noqa: BLE001 — one unparseable document must not lose the notice
        return [(name, [], f"err {type(exc).__name__}")]
    return [(name, [], "unsupported_ext")]


# ── matching helpers ─────────────────────────────────────────────────────────
NUMPREFIX = re.compile(
    r"^\s*(?:(?:appendix|annex|section|part|schedule)\s+)?[\dIVXivx]+(?:[.\)][\da-z]+)*[.\)]?\s+", re.I
)
DOTLEADER = re.compile(r"\.{2,}\s*\d{1,4}\s*$|\s\.{2,}\s|\t\d{1,4}\s*$")
TRAILING_PAGENO = re.compile(r"[a-z\)]\s{1,}\d{1,3}\s*$", re.I)
PROSE_START = re.compile(r"^[a-z]")


def headingish(line: str) -> bool:
    """The cheap pre-filter: short enough to be a heading, not an obvious sentence."""
    if len(line) > HEADINGISH_MAX:
        return False
    return not (line.endswith((".", ";")) and len(line.split()) > 14)


def norm_head(s: str) -> str:
    s = NUMPREFIX.sub("", s or "").strip()
    s = re.sub(r"\.{2,}\s*\d{1,4}\s*$", "", s).strip()
    s = re.sub(r"\s+", " ", s).lower()
    return s.strip(" .:–-\t")


def is_toc_line(ln: Line, total: int, dup_positions: dict) -> tuple[bool, str]:
    """Correction 5. Returns (is_contents_echo, reason).

    KNOWN DEFECT: the early_duplicate branch is gated on `ln.idx / total < 0.20`, so a contents
    entry past the first fifth of a long line pool is not recognised. PDF extraction routinely
    puts the page number on its own line, leaving no dot leader and no trailing number, so this
    positional gate is the only defence left. Pinned by a strict xfail, deliberately unfixed here.
    """
    st = (ln.style or "").lower()
    if st.startswith("toc") or "table of contents" in st or st.startswith("contents"):
        return True, "toc_style"
    if DOTLEADER.search(ln.text):
        return True, "dot_leader"
    key = norm_head(ln.text)
    later = [p for p in dup_positions.get(key, []) if p > ln.idx]
    if later and total and (ln.idx / total) < 0.20:
        return True, "early_duplicate"
    if total and (ln.idx / total) < 0.12 and TRAILING_PAGENO.search(ln.text) and len(ln.text) < 90:
        return True, "early_pageno"
    return False, ""


def heading_styled(ln: Line) -> bool:
    """True when Word itself called this line a heading — the strongest signal available."""
    st = (ln.style or "").lower()
    return st.startswith("heading") or st.startswith("title")


def heading_shaped(text: str) -> bool:
    """A real heading, not a sentence that happens to contain the words."""
    core = NUMPREFIX.sub("", text or "").strip()
    words = core.split()
    if not words or len(words) > 12:
        return False
    alpha = [w for w in words if w[:1].isalpha()]
    if not alpha:
        return False
    titleish = sum(1 for w in alpha if w[0].isupper()) >= max(1, int(0.6 * len(alpha)))
    return titleish or bool(NUMPREFIX.match(text or ""))


def scan(lines: list[Line]) -> dict:
    """-> {tag: [Line, ...]} for every pattern, landing and corroboration alike."""
    hits: dict = collections.OrderedDict()
    for tag, rx, _role in PATTERNS:
        for ln in lines:
            if headingish(ln.text) and rx.search(ln.text):
                hits.setdefault(tag, []).append(ln)
    return hits


def pick_line(cands: list[Line], total: int) -> tuple[Line, str, dict]:
    """Correction 5: prefer a real heading over a table-of-contents echo -- and, for the
    same reason (land on the section, not on prose that mentions it), over a mid-sentence
    fragment."""
    dup = collections.defaultdict(list)
    for ln in cands:
        dup[norm_head(ln.text)].append(ln.idx)
    scored = []
    n_toc = n_prose = 0
    for ln in cands:
        toc, why = is_toc_line(ln, total, dup)
        score = 0
        if toc:
            score -= 100
            n_toc += 1
        if heading_styled(ln):
            score += 10
        if heading_shaped(ln.text):
            score += 6
        # mid-sentence prose starts with an ASCII lowercase letter. A Greek/bullet
        # marker such as `α:` is a heading prefix, not prose - do not penalise it.
        if PROSE_START.match((ln.text or "").strip()):
            score -= 60
            n_prose += 1
            why = why or "prose_fragment"
        scored.append((-score, ln.idx, ln, why))
    scored.sort(key=lambda x: (x[0], x[1]))
    _, _, ln, why = scored[0]
    picked_toc, _ = is_toc_line(ln, total, dup)
    stats = {
        "toc_rejected": n_toc - (1 if picked_toc else 0),
        "prose_rejected": n_prose - (1 if PROSE_START.match((ln.text or "").strip()) else 0),
        "candidates": len(cands),
    }
    return ln, why, stats


def find_landing(docs: list[dict]) -> dict | None:
    """Correction 1+2+3: the best available pattern in the best available document.

    Ranking is hierarchical and document-dominant: published-document class first, then (for
    .zip bundles) the class of the member inside it, then published listing order. Pattern
    preference only breaks ties inside one unit. A `land`-role pattern is required, so
    `section_3_3` can never be a target (correction 3).
    """
    units = []
    for d in docs:
        dcls, dscore = doc_class(d["filename"])
        for pi, (src, lines, _status) in enumerate(d["parts"]):
            mcls, mscore = doc_class(member_name(src))
            units.append((-dscore, -mscore, d.get("orig_idx", 0), pi, d, src, lines, dcls, mcls))
    units.sort(key=lambda u: u[:4])
    fallback = None
    for *_rank, d, src, lines, dcls, mcls in units:
        hits = scan(lines)
        land_tags = sorted([t for t in hits if t in LAND_RANK], key=lambda t: LAND_RANK[t])
        if not land_tags:
            continue
        total = len(lines)
        for tag in land_tags:
            ln, why, stats = pick_line(hits[tag], total)
            cand = {
                "doc": d["filename"],
                "document_id": d.get("document_id"),
                "part": src,
                "pattern": tag,
                "heading": ln.text,
                "line_idx": ln.idx,
                "n_lines": total,
                "doc_class": dcls,
                "member_class": mcls,
                "toc_skipped": stats["toc_rejected"],
                "prose_skipped": stats["prose_rejected"],
                "candidates": stats["candidates"],
                "reason_of_pick": why,
                "style": ln.style,
                "has_style": heading_styled(ln),
                "corroboration": sorted(t for t in hits if t not in LAND_RANK),
            }
            # A pattern that only ever occurs mid-sentence is not a section to link to;
            # try the next-best pattern, then the next document, before settling for it.
            if heading_shaped(ln.text) and not PROSE_START.match((ln.text or "").strip()):
                return cand
            if fallback is None:
                cand["degraded"] = "no_heading_shaped_occurrence"
                fallback = cand
    return fallback


def confidence_band(pattern: str, has_style: bool) -> str:
    """Three honest levels. The audit measured reliability per pattern; it does not support
    a numeric score, so none is invented."""
    strong = int(pattern == CONFIDENCE_PATTERN) + int(bool(has_style))
    return {2: "high", 1: "medium"}.get(strong, "low")


def document_url(tender_id: str, document_id: str) -> str:
    """The anonymous deep link. `/cft/` is load-bearing — dropping it returns HTTP 500."""
    return f"{DOWNLOAD_URL}?documentId={document_id}&resourceId={tender_id}"


# ── cache loading ────────────────────────────────────────────────────────────
def load_cached_notice(notice_dir: Path, listing_docs: list[dict] | None = None) -> list[dict]:
    """Read one notice's cached documents as `find_landing` units.

    Files are `<document_id>__<filename>`. `listing_docs` (from a listings JSONL) restores the
    PUBLISHED order, which is the last tiebreak in the ranking; without it, on-disk name order
    is used so a run stays deterministic but the tiebreak is arbitrary.
    """
    if not notice_dir.is_dir():
        return []
    on_disk: dict[str, Path] = {}
    for p in sorted(notice_dir.iterdir()):
        if p.is_file():
            on_disk.setdefault(p.name.split("__", 1)[0], p)
    if listing_docs:
        ordered = [
            (str(d.get("document_id")), d.get("filename") or "")
            for d in listing_docs
            if str(d.get("document_id")) in on_disk
        ]
    else:
        ordered = [(docid, p.name.split("__", 1)[-1]) for docid, p in on_disk.items()]
    out = []
    for i, (docid, published_name) in enumerate(ordered):
        path = on_disk[docid]
        parts = extract_bytes(path.name, path.suffix.lower(), path.read_bytes())
        out.append(
            {
                "filename": published_name or path.name.split("__", 1)[-1],
                "document_id": docid,
                "orig_idx": i,
                "parts": parts,
                "n_lines": sum(len(lines) for _, lines, _ in parts),
            }
        )
    return out


def load_listings(path: Path | None) -> dict[str, list[dict]]:
    """{tender_id: [{document_id, filename}, ...]} from a JSONL of listing records."""
    if not path:
        return {}
    out: dict[str, list[dict]] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        rec = json.loads(line)
        out[str(rec.get("tender_id"))] = rec.get("docs") or []
    return out


# ── network lane (opt-in) ────────────────────────────────────────────────────
DOC_ANCHOR = re.compile(r"onclick=\"downloadDocForAnonymous\('([^']+)'\)\"[^>]*>(.*?)</a>", re.S)
TAGS = re.compile(r"<[^>]+>")
HEADERS = polite_headers(browser=True, extra={"Referer": BASE})


def _sleep(delay_ms: int) -> None:
    time.sleep(delay_ms / 1000.0)


def fetch_listing(tender_id: str, delay_ms: int) -> list[dict]:
    """First page of a notice's document listing. No rate limiting exists in http_engine,
    so the spacing is applied here, before the request."""
    import html as html_mod

    _sleep(delay_ms)
    page, _status = fetch_text(LIST_URL, headers=HEADERS, params={"resourceId": tender_id}, attempts=2)
    docs = []
    for docid, label in DOC_ANCHOR.findall(page):
        docs.append({"document_id": docid, "filename": html_mod.unescape(TAGS.sub("", label)).strip()})
    return docs


def harvest_notice(tender_id: str, docs: list[dict], notice_dir: Path, delay_ms: int, max_docs: int) -> int:
    """Download up to `max_docs` text-bearing documents into the cache. Returns the count kept."""
    notice_dir.mkdir(parents=True, exist_ok=True)
    kept = 0
    for d in docs:
        if kept >= max_docs:
            break
        name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", d["filename"])[:120]
        if Path(name).suffix.lower() not in TEXTY_EXTS:
            continue
        dest = notice_dir / f"{d['document_id']}__{name}"
        if dest.exists() and dest.stat().st_size > 0:
            kept += 1
            continue
        _sleep(delay_ms)
        blob = fetch_bytes(document_url(tender_id, d["document_id"]), headers=HEADERS, timeout=180)
        if not blob or blob[:15].lstrip()[:5].lower() == b"<html":
            log.warning("%s/%s: download failed or returned HTML", tender_id, d["document_id"])
            continue
        if len(blob) > MAX_DOC_BYTES:
            log.warning("%s/%s: %d bytes over cap, skipped", tender_id, d["document_id"], len(blob))
            continue
        dest.write_bytes(blob)
        kept += 1
    return kept


# ── driver ───────────────────────────────────────────────────────────────────
ROW_SCHEMA = {
    "tender_id": pl.Utf8,
    "document_id": pl.Utf8,
    "document_filename": pl.Utf8,
    "section_heading": pl.Utf8,
    "pattern": pl.Utf8,
    "doc_class": pl.Utf8,
    "line_idx": pl.Int64,
    "n_lines": pl.Int64,
    "has_style": pl.Boolean,
    "confidence": pl.Utf8,
    "document_url": pl.Utf8,
    "extracted_utc": pl.Datetime(time_unit="us", time_zone="UTC"),
    "matcher_version": pl.Utf8,
}


def landing_row(tender_id: str, landing: dict, now: datetime) -> dict:
    """One parquet row from one landing. `document_filename` is the PUBLISHED document — for a
    landing inside a .zip bundle that is the bundle, since that is what `document_url` fetches."""
    return {
        "tender_id": str(tender_id),
        "document_id": str(landing.get("document_id") or ""),
        "document_filename": landing["doc"],
        "section_heading": landing["heading"],
        "pattern": landing["pattern"],
        "doc_class": landing["doc_class"],
        "line_idx": int(landing["line_idx"]),
        "n_lines": int(landing["n_lines"]),
        "has_style": bool(landing["has_style"]),
        "confidence": confidence_band(landing["pattern"], landing["has_style"]),
        "document_url": document_url(str(tender_id), str(landing.get("document_id") or "")),
        "extracted_utc": now,
        "matcher_version": MATCHER_VERSION,
    }


def run(args: argparse.Namespace) -> int:
    cache_dir = Path(args.cache_dir)
    if not cache_dir.is_dir() and not args.fetch:
        log.error("cache dir %s does not exist and --fetch was not given", cache_dir)
        return 2
    listings = load_listings(Path(args.listings) if args.listings else None)

    if args.tender_ids:
        tender_ids = [str(t) for t in args.tender_ids]
    elif listings:
        tender_ids = list(listings)
    else:
        tender_ids = sorted(p.name for p in cache_dir.iterdir() if p.is_dir())

    if args.fetch:
        for tid in tender_ids:
            docs = listings.get(tid) or fetch_listing(tid, args.delay_ms)
            listings.setdefault(tid, docs)
            n = harvest_notice(tid, docs, cache_dir / tid, args.delay_ms, args.max_docs)
            log.info("%s: %d document(s) cached", tid, n)

    now = datetime.now(UTC)
    rows: list[dict] = []
    no_landing: list[str] = []
    for tid in tender_ids:
        docs = load_cached_notice(cache_dir / tid, listings.get(tid))
        if not docs:
            no_landing.append(tid)
            continue
        landing = find_landing(docs)
        if landing is None:
            no_landing.append(tid)
            continue
        rows.append(landing_row(tid, landing, now))

    df = pl.DataFrame(rows, schema=ROW_SCHEMA)
    bands = dict(df.group_by("confidence").len().sort("confidence").iter_rows()) if df.height else {}
    log.info(
        "%d notice(s) scanned, %d landing(s), %d without one; confidence %s",
        len(tender_ids),
        df.height,
        len(no_landing),
        bands,
    )
    if args.dry_run:
        log.info("--dry-run: not writing %s", args.out)
        return 0
    save_parquet(df, Path(args.out), min_rows=1)
    log.info("wrote %s (%d rows)", args.out, df.height)
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Locate the eligibility/selection-criteria section in eTenders documents.")
    ap.add_argument("--cache-dir", required=True, help="<dir>/<tender_id>/<document_id>__<name> document cache")
    ap.add_argument("--listings", help="JSONL of {tender_id, docs:[{document_id, filename}]} — restores listing order")
    ap.add_argument("--tender-ids", nargs="*", help="restrict to these notices (default: every notice in the cache)")
    ap.add_argument("--fetch", action="store_true", help="harvest missing documents into the cache first (network)")
    ap.add_argument("--delay-ms", type=int, default=1500, help="spacing before EVERY request; floor 1000")
    ap.add_argument("--max-docs", type=int, default=14, help="documents to download per notice under --fetch")
    ap.add_argument("--out", default=str(OUT_PARQUET))
    ap.add_argument("--dry-run", action="store_true", help="scan and report, write no parquet")
    args = ap.parse_args()
    if args.delay_ms < DELAY_MS_FLOOR:
        log.error("--delay-ms %d below the %dms politeness floor; refusing.", args.delay_ms, DELAY_MS_FLOOR)
        return 2
    return run(args)


if __name__ == "__main__":
    run_extractor(main)
