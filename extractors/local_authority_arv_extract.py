"""Local-authority Annual Rate on Valuation (ARV) — the commercial-rates multiplier, per LA per year.

A business's rates bill is `NAV x ARV` (NAV = the Tailte valuation of the premises). This ingests
the national ARV table so the app can show every council's multiplier and its trajectory — the
one commercial-rates figure that IS published machine-readably. The Tailte valuation list API
(opendata.tailte.ie GetProperties) was probed 2026-07-24: correct endpoint, but every query
returns zero rows (outage or quiet decommissioning) — NAV ingestion is deferred until it serves
data again, so no per-premises bill can be computed yet.

Source: DHLGH "Local Authority Annual Rates on Valuation (ARV)" publication (gov.ie), one XLSX,
sheet `ARVs`, grain LA x year 2017-2026 (31 rating authorities + footnote rows). The gov.ie CDN
403s bare requests — browser headers required (same as derelict_sites_levy_extract).

Reads  : doc/source_pdfs/2026_Annual_Rates_on_Valuation.xlsx  (cached source, git-tracked)
Writes : data/silver/parquet/local_authority_arv.parquet      (long: la, year, arv)
         data/_meta/local_authority_arv_coverage.json

Run:
  python extractors/local_authority_arv_extract.py              # parse cache + write
  python extractors/local_authority_arv_extract.py --download   # re-fetch from gov.ie first
  python extractors/local_authority_arv_extract.py --dry-run    # parse + report, no write
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from services.coverage_io import save_coverage  # noqa: E402
from services.extract_runner import run_extractor  # noqa: E402
from services.http_engine import fetch_bytes, polite_headers  # noqa: E402
from services.parquet_io import save_parquet  # noqa: E402

LOG = logging.getLogger("local_authority_arv")
_SRC = ROOT / "doc" / "source_pdfs" / "2026_Annual_Rates_on_Valuation.xlsx"
_OUT = ROOT / "data" / "silver" / "parquet" / "local_authority_arv.parquet"
_COV = ROOT / "data" / "_meta" / "local_authority_arv_coverage.json"
_URL = "https://assets.gov.ie/static/documents/09fe3ad4/2026_Annual_Rates_on_Valuation.xlsx"
# 31 rating authorities; fewer rows means the sheet layout changed and the parse must not ship.
_ROW_FLOOR = 31


def _download() -> None:
    data = fetch_bytes(
        _URL,
        headers=polite_headers(browser=True, extra={"Referer": "https://www.gov.ie/"}),
        validate=lambda b: b[:4] == b"PK\x03\x04",  # xlsx = zip container
    )
    if not data:
        raise SystemExit(f"download failed: {_URL}")
    _SRC.parent.mkdir(parents=True, exist_ok=True)
    _SRC.write_bytes(data)
    LOG.info("downloaded %s (%d bytes)", _SRC.name, len(data))


def parse(src: Path) -> pl.DataFrame:
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb["ARVs"] if "ARVs" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # the header row must carry year columns — the sheet's TITLE row also starts with
    # "Local Authority", which is exactly how the first parse matched the wrong row
    header_i = next(
        i
        for i, r in enumerate(rows)
        if r
        and str(r[0] or "").strip().lower().startswith("local authority")
        and any(str(c or "").strip().isdigit() for c in r[1:])
    )
    header = rows[header_i]
    years = [(j, int(c)) for j, c in enumerate(header[1:], start=1) if str(c or "").strip().isdigit()]
    out: list[dict] = []
    for r in rows[header_i + 1 :]:
        la = str(r[0] or "").strip()
        # footnote/blank rows end the table; a real LA row always prices at least one year
        if not la or la.lower().startswith(("note", "*", "source")):
            continue
        for j, year in years:
            v = r[j] if j < len(r) else None
            if isinstance(v, str):
                v = v.strip().replace(",", "")
            try:
                f = float(v)
            except (TypeError, ValueError):
                continue
            # BASIS BREAK, not noise: pre-revaluation years carry the OLD rate-on-valuation basis
            # (e.g. Carlow 69.56 in 2017 vs 0.2571 from 2018; Clare switches at 2023). The two
            # bases apply to DIFFERENT NAV scales and must never be plotted as one series —
            # flagged per row so the UI can split them.
            out.append({"la": la, "year": year, "arv": f, "pre_revaluation_basis": f > 10})
    df = pl.DataFrame(
        out, schema={"la": pl.Utf8, "year": pl.Int32, "arv": pl.Float64, "pre_revaluation_basis": pl.Boolean}
    )
    n_las = df["la"].n_unique()
    if n_las < _ROW_FLOOR:
        raise SystemExit(f"only {n_las} local authorities parsed (< {_ROW_FLOOR}) — layout changed?")
    return df


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--download", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    if args.download or not _SRC.exists():
        _download()
    df = parse(_SRC)
    span = (int(df["year"].min()), int(df["year"].max()))
    LOG.info("parsed %d rows | %d LAs | years %s-%s", df.height, df["la"].n_unique(), *span)
    if args.dry_run:
        print(df.group_by("year").agg(n=pl.len(), mean_arv=pl.col("arv").mean().round(4)).sort("year"))
        return
    save_parquet(df, _OUT)
    save_coverage(
        {
            "source": _URL,
            "cached_file": _SRC.name,
            "rows": df.height,
            "local_authorities": df["la"].n_unique(),
            "year_min": span[0],
            "year_max": span[1],
            "grain": "la x year (long); arv = annual rate on valuation multiplier",
            "basis_break": "pre_revaluation_basis=true rows are the OLD rate basis (values ~50-80 vs ~0.2); never chart the two bases as one series",
            "note": "rates bill = NAV x ARV; NAV per premises awaits the Tailte GetProperties API serving data again",
        },
        _COV,
    )
    print(f"OK {_OUT.name}: {df.height} rows, {df['la'].n_unique()} LAs, {span[0]}-{span[1]}")


if __name__ == "__main__":
    run_extractor(main)
