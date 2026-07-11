"""AHBRA — Register of Approved Housing Bodies (SANDBOX).

Source: https://www.ahbregulator.ie/registration/the-register/
The statutory register (Housing (Regulation of Approved Housing Bodies) Act
2019, Part 4) is published as a single XLSX ("AHB Register <Month> <Year>").
Sheets: 'Register' (current AHBs, ~425) and 'Removed From Register' (~448
cancelled/removed entries). Both are captured into one silver frame,
distinguished by ``register_section``.

Join hooks for the wider project:
  * ``ahb_name``       -> social-housing construction-status "Approved Housing Body" col
  * ``charity_rcn``    -> charities register (Registered Charity Number, where applicable)
  * ``cro_number``     -> CRO companies register (where applicable)

Licence note: site footer is "(c) Approved Housing Bodies Regulatory Authority.
All rights reserved." — no explicit re-use licence; statutory public register
(s.32-34 of the 2019 Act). Flag for the licence gate before any promotion.

Run (repo root):  .venv/Scripts/python pipeline_sandbox/new_sources/ahbra_register.py
"""
from __future__ import annotations

import re
import sys
import time
import unicodedata
from datetime import date, datetime
from pathlib import Path

import polars as pl
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _common import POLITE_DELAY_S, cache_raw, now_iso, sha256_bytes, write_silver  # noqa: E402

REGISTER_PAGE = "https://www.ahbregulator.ie/registration/the-register/"
FALLBACK_XLSX = "https://www.ahbregulator.ie/app/uploads/2026/06/Copy-of-AHB-Register-June-2026.xlsx"

# Browser-spoof headers (same pattern as extractors/procurement_etenders_extract.py
# GOVIE_HEADERS): ahbregulator.ie serves fine today, but gov.ie-family CDNs 403
# bot user-agents, so we send a browser UA defensively.
BROWSER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Referer": "https://www.ahbregulator.ie/",
    "Accept-Language": "en-IE,en;q=0.9",
}


def fetch_b(url: str, binary: bool = False, timeout: int = 60):
    """GET with browser headers + polite delay; same meta shape as _common.fetch."""
    time.sleep(POLITE_DELAY_S)
    r = requests.get(url, headers=BROWSER_HEADERS, timeout=timeout)
    r.raise_for_status()
    meta = {
        "source_url": r.url,
        "status": r.status_code,
        "content_type": r.headers.get("content-type", ""),
        "source_last_modified": r.headers.get("last-modified"),
        "source_document_hash": sha256_bytes(r.content),
        "fetched_at": now_iso(),
        "bytes": len(r.content),
    }
    return (r.content if binary else r.text), meta


COUNTIES = [
    "Dublin", "Cork", "Galway", "Limerick", "Waterford", "Kerry", "Clare",
    "Tipperary", "Kilkenny", "Wexford", "Carlow", "Kildare", "Laois",
    "Longford", "Louth", "Meath", "Offaly", "Westmeath", "Wicklow", "Cavan",
    "Donegal", "Monaghan", "Leitrim", "Mayo", "Roscommon", "Sligo",
]


def _fold(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower()


def county_from_address(addr: str | None) -> str | None:
    """Best-effort: last county name appearing in the address (addresses end
    with the county). Derived column — never a source value."""
    if not addr:
        return None
    a = _fold(addr)
    best, best_pos = None, -1
    for c in COUNTIES:
        for m in re.finditer(rf"\b{c.lower()}\b", a):
            if m.start() > best_pos:
                best_pos, best = m.start(), c
    return best


def _s(v) -> str | None:
    """Cell -> clean string (ints like CRO/RCN numbers kept un-floated)."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return s or None


def _d(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def resolve_xlsx_url() -> str:
    """The XLSX filename carries the publication month, so resolve it live."""
    try:
        html, _ = fetch_b(REGISTER_PAGE)
        s = BeautifulSoup(html, "html.parser")
        for a in s.find_all("a", href=True):
            if a["href"].lower().endswith(".xlsx"):
                return a["href"]
    except Exception as e:  # noqa: BLE001 — fall back to last-known URL
        print(f"  register page fetch failed ({type(e).__name__}: {e}); using fallback URL")
    return FALLBACK_XLSX


def register_version(xlsx_url: str) -> tuple[str | None, str | None]:
    """('June 2026', '2026-06') from the filename — month precision, kept honest."""
    m = re.search(r"([A-Z][a-z]+)[-_ ](\d{4})", Path(xlsx_url).name)
    if not m:
        return None, None
    try:
        month = datetime.strptime(m.group(1), "%B").month
    except ValueError:
        return f"{m.group(1)} {m.group(2)}", None
    return f"{m.group(1)} {m.group(2)}", f"{m.group(2)}-{month:02d}"


def run() -> None:
    xlsx_url = resolve_xlsx_url()
    print(f"register XLSX: {xlsx_url}")
    blob, meta = fetch_b(xlsx_url, binary=True)
    bronze_path, _ = cache_raw("ahbra", Path(xlsx_url).name, blob)
    version, pub_month = register_version(xlsx_url)
    print(f"  cached {meta['bytes']:,} bytes -> {bronze_path}  (version: {version})")

    import io
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)

    prov = {
        "source_url": meta["source_url"],
        "source_document_hash": meta["source_document_hash"],
        "fetched_at": meta["fetched_at"],
        "source_published_date": pub_month,          # month precision from filename
        "source_last_modified": meta["source_last_modified"],
        "extraction_method": "xlsx_openpyxl",
        "confidence": "high",
        "privacy_tier": "public",                    # statutory public register
        "register_version": version,
    }

    rows: list[dict] = []
    ws = wb["Register"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = _s(r[1]) if len(r) > 1 else None
        if not name:
            continue
        addr = _s(r[8])
        rows.append({
            "register_section": "registered",
            "ahb_name": name,
            "ahb_registration_number": _s(r[2]),
            "registration_date": _d(r[3]),
            "entity_type": _s(r[4]),
            "cro_number": _s(r[5]),
            "charity_rcn": _s(r[6]),
            "eircode": _s(r[7]),
            "principal_place_of_business": addr,
            "county_derived": county_from_address(addr),
            "non_compliance": _s(r[9]),
            "non_implementation_of_compliance_plan": _s(r[10]),
            "compliance_plan": _s(r[11]),
            "governance_standard": _s(r[12]),
            "financial_standard": _s(r[13]),
            "property_asset_standard": _s(r[14]),
            "tenancy_standard": _s(r[15]),
            "status": _s(r[16]),
            "governing_body_members": _s(r[17]) if len(r) > 17 else None,
            "removal_reason": None,
            "removed_date": None,
            **prov,
        })
    n_registered = len(rows)

    ws = wb["Removed From Register"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        name = _s(r[1]) if len(r) > 1 else None
        if not name:
            continue
        rows.append({
            "register_section": "removed",
            "ahb_name": name,
            "ahb_registration_number": _s(r[2]),
            "registration_date": None,
            "entity_type": None,
            "cro_number": None,
            "charity_rcn": None,
            "eircode": None,
            "principal_place_of_business": None,
            "county_derived": None,
            "non_compliance": None,
            "non_implementation_of_compliance_plan": None,
            "compliance_plan": None,
            "governance_standard": None,
            "financial_standard": None,
            "property_asset_standard": None,
            "tenancy_standard": None,
            "status": _s(r[3]),
            "governing_body_members": None,
            "removal_reason": _s(r[4]),
            "removed_date": _d(r[5]),
            **prov,
        })
    n_removed = len(rows) - n_registered

    df = pl.DataFrame(rows, schema_overrides={"registration_date": pl.Date, "removed_date": pl.Date},
                      infer_schema_length=None)
    out = write_silver("ahbra_register", df)

    # ---- profile (compact; no raw dumps) ----
    reg = df.filter(pl.col("register_section") == "registered")
    print(f"\nSILVER: {out}  rows={df.height} (registered={n_registered}, removed={n_removed})")
    for col in ("ahb_registration_number", "registration_date", "cro_number",
                "charity_rcn", "eircode", "principal_place_of_business", "county_derived"):
        nn = reg[col].null_count()
        print(f"  registered.{col:<32} null {nn:>3}/{reg.height} ({100*nn/reg.height:.0f}%)")
    print("  status counts (registered):", dict(reg.group_by("status").len().iter_rows()))
    top_c = reg.drop_nulls("county_derived").group_by("county_derived").len().sort("len", descending=True)
    print(f"  county_derived: {top_c.height} distinct; top5 "
          f"{[(r[0], r[1]) for r in top_c.head(5).iter_rows()]}")
    # join-readiness signals
    names = reg["ahb_name"].to_list()
    clg = sum(1 for n in names if n.upper().rstrip(". ").endswith("CLG"))
    accents = sum(1 for n in names if _fold(n) != n.lower())
    dupes = reg.group_by("ahb_name").len().filter(pl.col("len") > 1).height
    print(f"  name join-readiness: CLG-suffixed={clg}, accented={accents}, dup names={dupes}")
    print(f"  charity_rcn present: {reg.height - reg['charity_rcn'].null_count()}, "
          f"cro_number present: {reg.height - reg['cro_number'].null_count()}")
    dr = reg.drop_nulls("registration_date")["registration_date"]
    if dr.len():
        print(f"  registration_date range: {dr.min()} .. {dr.max()}")


if __name__ == "__main__":
    run()
