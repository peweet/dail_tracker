"""16 read_* publisher-layout parsers + their small role/column-detection
helpers, for the public-body payments extractor package (split out of
extractors/procurement_public_body_extract.py, doc/REFACTORING_CANDIDATES.md
C7, pure move-function -- no logic changes).

Leaf module: no dependency on config.py/harvest.py/emit.py in this package.
"""

from __future__ import annotations

import contextlib
import io
import re

import fitz  # PyMuPDF
import polars as pl

from shared.pdf_layout import cluster_word_rows
from shared.text_encoding import decode_table_bytes

# ----------------------------------------------------------------------------- regexes
MONEY_RE = re.compile(r"(?:€|EUR)?\s?\d{1,3}(?:,\d{3})+(?:\.\d{2})?|\d+\.\d{2}")
NUM_RE = re.compile(r"\d[\d,]*(?:\.\d+)?")
DIGIT_PREFIX = re.compile(r"^(?:\d{3,}\s+){1,3}")  # strips a leading PO/vendor-ID run
PERIOD_RE = re.compile(r"(?:^|[^0-9])(20[12]\d)(?:[^0-9]|$)")
QUARTER_RE = re.compile(r"q\s?([1-4])|quarter[\s_-]?([1-4])|qtr[\s_-]?([1-4])", re.I)

ROLE_RE = {
    # English + Irish (as Gaeilge) — some bodies (e.g. An Bord Pleanála) publish bilingual
    # headers: Soláthraí=supplier, Glanmhéid/Méid Comhlán=net/gross amount, Cur Síos=description,
    # Tagairt=reference, Dáta=date. English-only role regexes leave supplier=None on those.
    "supplier": re.compile(r"supplier|payee|vendor|provider|customer|recipient|\bname\b|soláthr", re.I),
    "amount": re.compile(
        r"amount|total|value|gross|\beuro\b|€|\bpaid\b|\bvat\b|ledger|payment\b|m[ée]id|mhéid|luach|comhlán", re.I
    ),
    "description": re.compile(
        r"descript|\bdesc\b|detail|categor|service|goods|nature|\bgl\b|main gl|cur síos|tuairisc|seirbhís|earraí", re.I
    ),
    "po": re.compile(
        r"\border\b|\bpo\b|\bpor\b|referen|\bref\b|\bnumber\b|invoice|\bdoc\b|transaction|tagairt|uimhir|ordú", re.I
    ),
    "period": re.compile(r"period|quarter|\bqtr\b|\bdate\b|\byear\b|posting|month|dáta|ráithe|bliain", re.I),
    "paid": re.compile(r"\bpaid\b|payment type|status|no\.? of payments|íoctha", re.I),
}
MERGE_GAP = 22.0


# ============================================================================ readers
def to_eur(token) -> float | None:
    if token is None:
        return None
    if isinstance(token, (int, float)):
        return float(token)
    s = str(token).strip()
    neg = s.startswith("(") and s.endswith(")")
    m = NUM_RE.search(s)
    if not m:
        return None
    with contextlib.suppress(ValueError):
        v = float(m.group().replace(",", ""))
        return -v if neg else v
    return None


def clean_supplier(s) -> str:
    return DIGIT_PREFIX.sub("", str(s or "")).strip(" -:|,")


# Month-range filenames (e.g. "jul-sep-2016.pdf") encode a quarter the Q\d patterns miss; many
# older Dept of Defence PO lists are named this way, so their period otherwise collapses to the
# bare year and recurring quarterly payments look like cross-file duplicates. DQ audit 2026-06-06.
MONTH_RANGE_Q = [
    (re.compile(r"jan\w*[-_ .]+mar", re.I), 1),
    (re.compile(r"apr\w*[-_ .]+jun", re.I), 2),
    (re.compile(r"jul\w*[-_ .]+sep", re.I), 3),
    (re.compile(r"oct\w*[-_ .]+dec", re.I), 4),
]


def quarter_from_name(name: str) -> int | None:
    q = QUARTER_RE.search(name)
    if q:
        return next((int(g) for g in q.groups() if g), None)
    for rx, qn in MONTH_RANGE_Q:
        if rx.search(name):
            return qn
    return None


def period_from_url(url: str) -> tuple[str | None, int | None, int | None]:
    name = url.rsplit("/", 1)[-1]
    y = PERIOD_RE.search(name)
    year = int(y.group(1)) if y else None
    quarter = quarter_from_name(name)
    period = f"{year}-Q{quarter}" if year and quarter else (str(year) if year else None)
    return period, year, quarter


# ---- PDF (header-anchored) ----
def find_header(rows: list[list]):
    best, best_hits = None, 1
    for r in rows[:18]:
        text = " ".join(w[4] for w in r)
        hits = sum(bool(rx.search(text)) for rx in ROLE_RE.values())
        has_anchor = ROLE_RE["supplier"].search(text) or ROLE_RE["amount"].search(text)
        if hits >= 2 and has_anchor and hits > best_hits:
            best, best_hits = r, hits
    return best


def header_columns(header: list) -> list[dict]:
    ws = sorted(header, key=lambda w: w[0])
    cols: list[dict] = []
    for w in ws:
        if cols and w[0] - cols[-1]["x1"] < MERGE_GAP:
            cols[-1]["label"] += " " + w[4]
            cols[-1]["x1"] = max(cols[-1]["x1"], w[2])
        else:
            cols.append({"label": w[4], "x0": w[0], "x1": w[2]})
    for c in cols:
        c["center"] = (c["x0"] + c["x1"]) / 2
    return cols


def assign_role(cols: list[dict]) -> dict[str, int]:
    roles: dict[str, int] = {}
    for role, rx in ROLE_RE.items():
        cands = [i for i, c in enumerate(cols) if rx.search(c["label"])]
        if role == "amount":
            # Drop identifier/count columns that carry a money keyword (DETE PDF "Payment No."
            # holds PO refs ~60,002,179 that dwarf the real "Total" amount). Same guard as the
            # tabular reader; see NON_AMOUNT_HDR.
            cands = [i for i in cands if not NON_AMOUNT_HDR.search(cols[i]["label"])] or cands
        if cands:
            roles[role] = cands[-1] if role == "amount" else cands[0]
    return roles


def row_to_cols(words: list, cols: list[dict]) -> list[str]:
    bounds = [(cols[i]["center"] + cols[i + 1]["center"]) / 2 for i in range(len(cols) - 1)]
    buckets: list[list] = [[] for _ in cols]
    for w in sorted(words, key=lambda w: w[0]):
        c = (w[0] + w[2]) / 2
        idx = 0
        while idx < len(bounds) and c > bounds[idx]:
            idx += 1
        buckets[idx].append(w[4])
    return [" ".join(b).strip(" -:|") for b in buckets]


def refine_roles(cols, roles, records):
    if not records:
        return roles

    def numfrac(i):
        vals = [r[i] for r in records if i < len(r) and r[i]]
        return sum(to_eur(v) is not None for v in vals) / len(vals) if vals else 0.0

    def moneyfrac(i):  # fraction of cells that look like MONEY (thousands/decimals) — a bare
        vals = [r[i] for r in records if i < len(r) and r[i]]  # year col like "2023" won't match
        return sum(bool(MONEY_RE.search(str(v))) for v in vals) / len(vals) if vals else 0.0

    amt_cands = [i for i, c in enumerate(cols) if ROLE_RE["amount"].search(c["label"])]
    # Exclude identifier/count columns before ranking by numeric density — a "Payment No." /
    # "Order No." column is 100% numeric and would otherwise win the numfrac tie-break.
    amt_cands = [i for i in amt_cands if not NON_AMOUNT_HDR.search(cols[i]["label"])] or amt_cands
    if amt_cands:
        roles["amount"] = max(amt_cands, key=numfrac)
    elif "amount" not in roles and cols:
        # No header word matches "amount" (e.g. NTMA's amount column is headed "Q4"/"Q3",
        # the quarter, not a money word). Fall back to the most money-like column by content.
        best = max(range(len(cols)), key=moneyfrac)
        if moneyfrac(best) >= 0.5:
            roles["amount"] = best
    sup_cands = [i for i, c in enumerate(cols) if ROLE_RE["supplier"].search(c["label"])]
    if sup_cands:
        roles["supplier"] = min(sup_cands, key=numfrac)
    return roles


def read_pdf(b: bytes, max_pages: int | None) -> dict:
    doc = fitz.open(stream=b, filetype="pdf")
    npages = doc.page_count
    limit = min(npages, max_pages) if max_pages else npages
    cols, header_label, page0 = [], "", ""
    for i in range(min(npages, 3)):
        rows = cluster_word_rows(doc[i])
        if i == 0:
            page0 = doc[i].get_text("text")
        h = find_header(rows)
        if h:
            cols = header_columns(h)
            header_label = " | ".join(c["label"] for c in cols)
            break
    roles = assign_role(cols) if cols else {}
    out_rows, digital_chars = [], 0
    for i in range(limit):
        page = doc[i]
        digital_chars += len(page.get_text("text").strip())
        if not cols:
            continue
        for wrow in cluster_word_rows(page):
            xs = [w[4] for w in wrow]
            if not any(MONEY_RE.search(t) for t in xs):
                continue
            out_rows.append((i + 1, row_to_cols(wrow, cols)))
    doc.close()
    roles = refine_roles(cols, roles, [r for _, r in out_rows]) if cols else roles
    return {
        "digital": digital_chars > 200,
        "cols": cols,
        "header_label": header_label,
        "roles": roles,
        "rows": out_rows,
        "page0": page0,
        "pages": npages,
    }


# ---- reading-order PDF (DCEDIY / dept_children) ----
_RO_REF = re.compile(r"^(\d{4,7})\s+(.+\S)\s*$")
_RO_REF_ALONE = re.compile(r"^\d{4,7}$")  # Layout B: the reference sits on its OWN line
_RO_DATE = re.compile(r"^(\d{2})/(\d{2})/(\d{4})(?:\s+(.*))?$")
_RO_AMT = re.compile(r"^\D{0,3}([\d,]+\.\d{2})\s*(.*)$")
_RO_HEADER = {
    "reference", "name", "supplier name", "amount", "total paid",
    "description", "payment", "payment date", "date",
}  # fmt: skip
# Boilerplate / footnote lines that bleed into the text layer between records (the DCEDIY
# "Please note: … inclusive of VAT …" footer block). Excluded so they never become a row or
# contaminate a description (user req: never ingest headers/footers).
_RO_BOILER = re.compile(r"please note|inclusive of vat|purchase orders? (are|for)|may have been|would appear", re.I)


def read_reading_order(b: bytes, max_pages: int | None) -> list[dict]:
    """Line-pair reader for the DCEDIY PO layout the generic word-geometry reader can't parse.

    Two published sub-layouts, handled together:
      A. 'ref supplier' on ONE line, then date + '€amount description' (in either order); and
      B. the reference ALONE on its own line, then supplier on the next line, then amount/desc/date
         — the layout the original reader silently dropped (e.g. 36d03592 yielded 1 of 252 rows,
         and several quarters yielded zero, ~8k rows absent from the fact: coverage_qa caught it).
    Order-agnostic: classify each post-ref line by pattern, flush on the next ref. Carries a
    PER-ROW payment date (the value here — provider spend over time), unlike the file-period schema.
    """
    doc = fitz.open(stream=b, filetype="pdf")
    limit = min(doc.page_count, max_pages) if max_pages else doc.page_count
    recs: list[dict] = []
    cur: dict | None = None

    def flush():
        nonlocal cur
        if cur and cur.get("amount") is not None and cur.get("date"):
            recs.append(cur)
        cur = None

    for pi in range(limit):
        for raw in doc[pi].get_text().splitlines():
            line = raw.strip()
            if not line or line.lower() in _RO_HEADER or _RO_BOILER.search(line):
                continue
            m = _RO_REF.match(line)
            if m:  # Layout A: 'ref supplier' inline
                flush()
                cur = {
                    "ref": m.group(1),
                    "supplier": m.group(2).strip(),
                    "date": None,
                    "amount": None,
                    "desc": "",
                    "page": pi + 1,
                }
                continue
            if _RO_REF_ALONE.match(line):  # Layout B: reference alone -> supplier on the next line
                flush()
                cur = {"ref": line, "supplier": None, "date": None, "amount": None, "desc": "", "page": pi + 1}
                continue
            if cur is None:
                continue
            if cur["supplier"] is None and not _RO_AMT.match(line) and not _RO_DATE.match(line):
                cur["supplier"] = line  # Layout B: first line after the bare ref is the supplier
                continue
            dm = _RO_DATE.match(line)
            if dm and cur["date"] is None:
                cur["date"] = f"{dm.group(3)}-{dm.group(2)}-{dm.group(1)}"
                if dm.group(4):  # date + amount + desc on one line
                    am = _RO_AMT.match(dm.group(4).strip())
                    if am and cur["amount"] is None:
                        cur["amount"], cur["desc"] = float(am.group(1).replace(",", "")), am.group(2).strip()
                continue
            am = _RO_AMT.match(line)
            if am and cur["amount"] is None:
                cur["amount"], cur["desc"] = float(am.group(1).replace(",", "")), am.group(2).strip()
                continue
            if cur["amount"] is not None:  # description continuation
                cur["desc"] = (cur["desc"] + " " + line).strip()
    flush()
    doc.close()
    return recs


# Courts Service "PO analysis report" reading-order reader. The generic word-geometry parser
# mis-columns these PDFs: for 2016+ quarters the PO number and supplier name are merged onto one
# text line, and x-coordinate bucketing then splits the name (body → PO column, legal suffix →
# supplier column), leaving ~850 empty-supplier rows AND ingesting period-total lines as payments
# (DQ audit P1, 2026-06-23; validated in pipeline_sandbox/courts_reader/). The files are a regular
# 5-field record anchored on the amount line; reading them in order recovers supplier+PO+amount+
# description+paid and naturally skips totals/boilerplate.
_CT_AMT = re.compile(r"^([\d,]+\.\d{2})\b\s*(.*)$")  # amount alone OR amount + desc on one line
_CT_PAID = re.compile(r"^(Y|N|Yes|No|Paid|Not Paid)$", re.I)
_CT_PO_NAME = re.compile(r"^(\d{4,7})\s+(.+)$")  # merged 'PO Supplier Name'
_CT_PO_ONLY = re.compile(r"^\d{4,7}$")
_CT_BOILER = {
    "po no.",
    "po no",
    "supplier name",
    "amount",
    "description",
    "paid yes/no",
    "paid",
    "the courts services",
    "the courts service",
    "yes/no",
}


def read_courts(b: bytes, max_pages: int | None) -> list[dict]:
    """Reading-order reader for the Courts Service PO PDFs → list of
    {ref, supplier, amount, desc, paid} dicts (period comes from the file URL)."""
    doc = fitz.open(stream=b, filetype="pdf")
    limit = min(doc.page_count, max_pages) if max_pages else doc.page_count
    lines: list[str] = []
    for pi in range(limit):
        for raw in doc[pi].get_text().splitlines():
            s = re.sub(
                r"\s+", " ", raw.replace("\xa0", " ").replace(" ", " ").replace(" ", " ").replace("�", " ")
            ).strip()
            if not s:
                continue
            low = s.lower()
            if low in _CT_BOILER or low.startswith("purchase orders") or "analysis report" in low:
                continue
            # end-of-page TOTAL summary + the multi-clause "Please Note:" footnote run into the next
            # page's first record otherwise — strip them. Keep real companies ("Total ICT Services").
            if re.match(r"(?i)^total\b\s*[:€]?\s*[\d,]", s) or "please note" in low:
                continue
            if re.match(r"^\d\)\s", s):  # footnote numbered clauses ("1) Withholding Tax …")
                continue
            if re.fullmatch(r"page \d+( of \d+)?", low):
                continue
            lines.append(s)
    doc.close()

    n = len(lines)
    recs: list[dict] = []
    rec_start = 0
    i = 0
    while i < n:
        am = _CT_AMT.match(lines[i])
        if not am:
            i += 1
            continue
        amount = float(am.group(1).replace(",", ""))
        head = lines[rec_start:i]
        j = i + 1
        desc: list[str] = []
        if am.group(2).strip():
            desc.append(am.group(2).strip())
        paid = None
        while j < n and not _CT_AMT.match(lines[j]):
            if _CT_PAID.match(lines[j]):
                paid = lines[j]
                j += 1
                break
            if _CT_PO_NAME.match(lines[j]) or _CT_PO_ONLY.match(lines[j]):
                break  # next record's head terminates this description
            desc.append(lines[j])
            j += 1
        po = None
        headtext = " ".join(head).strip()
        # Take the LAST 'PO-digits NAME' in the head. A page-break TOTAL line + multi-clause
        # footnote can bleed in front of the next record's header; the real PO+supplier is always
        # the final 'NNNNN NAME' run, so anchoring on the last one strips the noise. A pure 'Total'
        # head has no such run → supplier stays 'Total' and is dropped by the summary filter below.
        last = None
        for mm in re.finditer(r"(\d{4,7})\s+", headtext):
            if re.search(r"[A-Za-z]", headtext[mm.end() :]):
                last = mm
        if last is not None:
            po, supplier = last.group(1), headtext[last.end() :].strip()
        else:
            supplier = headtext
        # keep real suppliers; skip period TOTAL/sub-total summary lines ('Total', 'Total:',
        # 'Total €11,700,436.17'). A real company keeps a letter after 'Total' ('Total ICT Services
        # Ltd'), so the trailing class is punctuation/€/digits only — those rows survive.
        if (
            supplier
            and re.search(r"[A-Za-z]", supplier)
            and not re.fullmatch(r"(?i)\s*(grand\s+)?(sub[-\s]*)?totals?\b[\s:.€\d,]*", supplier)
        ):
            recs.append(
                {
                    "ref": po,
                    "supplier": supplier,
                    "amount": amount,
                    "desc": " ".join(desc).strip() or None,
                    "paid": paid,
                }
            )
        rec_start = j
        i = j
    return recs


# Department of Defence PO reading-order reader. The PDFs are a 5-column record
#   NUMBER (PO) | CATEGORY (unit) | SUPPLIER | CURRENCY (EUR/USD…) | AMOUNT
# published in TWO header orderings (Number,Category,Supplier vs Number,Supplier,Category). The
# generic word-geometry parser scrambles them — category→po_number, PO→description, the supplier's
# legal suffix split off — leaving 1,191 empty-supplier rows whose real name lands in po_number
# (DQ audit 2026-06; validated in pipeline_sandbox/courts_reader/). Anchoring on the CURRENCY line
# (a lone 3-letter code) with the AMOUNT right after it, then walking back to the leading PO digit
# line, recovers PO + supplier + amount cleanly and is robust to the two header orderings.
# Currency code on its own line — the anchor. Older Defence quarters (e.g. jan-mar-2019)
# print the WORD form ("EURO"/"US DOLLAR"/"STERLING") instead of the ISO code, so a code-only
# regex anchored 0 of those records (jan-mar-2019 fell to ~14% yield). Accept both forms.
_DD_CUR = re.compile(
    r"^(EUR|EURO|USD|US DOLLAR|GBP|STERLING|POUND STERLING|CHF|SEK|NOK|DKK|JPY|CAD|AUD|PLN|CZK|HUF|ZAR|SGD|HKD|NZD)$",
    re.I,
)
_DD_AMT = re.compile(r"^[€$£]?\s*([\d,]+\.\d{2})$")
_DD_PO = re.compile(r"^\d{4,7}$")
_DD_PO_NAME = re.compile(r"^(\d{4,7})\s+(.+)$")  # merged 'PO NAME' (4-field layout)
_DD_BOILER = {"number", "category", "supplier", "vendor", "currency", "amount", "po number", "po no", "po no.", "po"}


def _dd_supplier_first(doc) -> bool:
    """Header order: does SUPPLIER/VENDOR come before CATEGORY on the first page?"""
    head = " ".join(line.strip().lower() for line in doc[0].get_text().splitlines()[:8])
    si = min([head.find(k) for k in ("supplier", "vendor") if head.find(k) >= 0] or [9999])
    ci = head.find("category")
    return si < ci if (si < 9999 and ci >= 0) else False


def read_defence(b: bytes, max_pages: int | None) -> list[dict]:
    """Reading-order reader for the Department of Defence PO PDFs → list of
    {ref, supplier, category, amount} dicts (period comes from the file URL)."""
    doc = fitz.open(stream=b, filetype="pdf")
    limit = min(doc.page_count, max_pages) if max_pages else doc.page_count
    sup_first = _dd_supplier_first(doc)
    lines: list[str] = []
    for pi in range(limit):
        for raw in doc[pi].get_text().splitlines():
            s = re.sub(r"\s+", " ", raw.replace("\xa0", " ").replace("€", "")).strip()
            low = s.lower()
            if not s or low in _DD_BOILER or "purchase order" in low or "department of defence" in low:
                continue
            lines.append(s)
    doc.close()

    n = len(lines)
    recs: list[dict] = []
    i = 0
    while i < n:
        # anchor on a currency line immediately followed by an amount line
        if _DD_CUR.match(lines[i]) and i + 1 < n and _DD_AMT.match(lines[i + 1]):
            amount = float(_DD_AMT.match(lines[i + 1]).group(1).replace(",", ""))
            # walk back: PO is the nearest preceding pure-digit line; the 1-2 lines between PO and
            # the currency are CATEGORY + SUPPLIER (assigned by the file's header order).
            k = i - 1
            mid: list[str] = []
            po = None
            while k >= 0:
                if _DD_PO.match(lines[k]):  # pure-digit PO line (5-field layout)
                    po = lines[k]
                    break
                m = _DD_PO_NAME.match(lines[k])  # merged 'PO NAME' line (4-field layout)
                if m:
                    po = m.group(1)
                    mid.insert(0, m.group(2).strip())
                    break
                if _DD_CUR.match(lines[k]) or _DD_AMT.match(lines[k]):  # ran into the previous record
                    break
                mid.insert(0, lines[k])
                k -= 1
            category = supplier = None
            if len(mid) == 1:
                supplier = mid[0]
            elif len(mid) >= 2:
                if sup_first:
                    category, supplier = mid[-1], " ".join(mid[:-1])
                else:
                    category, supplier = mid[0], " ".join(mid[1:])
            if supplier and re.search(r"[A-Za-z]", supplier):
                recs.append({"ref": po, "supplier": supplier.strip(), "category": category, "amount": amount})
            i += 2
            continue
        i += 1
    return recs


# ---- Revenue Commissioners reading-order reader. Each record is the reference, supplier and
# description on their own lines, then the amount line (the amount carries an inline paid flag
# in some quarters: "102,534.41 Y"; others print the amount alone with the € and Y on separate
# lines, or "€2,372,503.95 Y" with a leading €). The generic word-geometry reader mis-columned
# these → ~726 rows lost (coverage_qa: quarter4-2025 at 0%). Reference-led: split on the bare
# reference line, classify the amount by its 2-decimals, supplier=first text line, rest=description.
_REV_REF = re.compile(r"^\d{4,7}$")
_REV_AMT = re.compile(r"^€?\s*([\d,]+\.\d{2})\s*(Y|N|Yes|No)?\s*$", re.I)
_REV_HDR = {
    "reference", "supplier", "supplier ", "description", "amount", "paid", "ledger amount",
    "customer/supplier description",
}  # fmt: skip


def read_revenue(b: bytes, max_pages: int | None) -> list[dict]:
    """Reading-order reader for the Revenue Commissioners payments PDFs → list of
    {ref, supplier, amount, desc, paid} dicts (period from the file URL)."""
    doc = fitz.open(stream=b, filetype="pdf")
    limit = min(doc.page_count, max_pages) if max_pages else doc.page_count
    lines: list[str] = []
    for pi in range(limit):
        for raw in doc[pi].get_text().splitlines():
            s = raw.replace("\xa0", " ").strip()
            if s:
                lines.append(s)
    doc.close()

    recs: list[dict] = []
    n = len(lines)
    i = 0
    while i < n:
        if _REV_REF.match(lines[i]) and not _REV_AMT.match(lines[i]):
            ref = lines[i]
            j = i + 1
            sup = None
            desc: list[str] = []
            amt = paid = None
            while j < n and not (_REV_REF.match(lines[j]) and not _REV_AMT.match(lines[j])):
                line = lines[j]
                if line.lower() in _REV_HDR:
                    j += 1
                    continue
                m = _REV_AMT.match(line)
                if m:
                    amt, paid = float(m.group(1).replace(",", "")), m.group(2)
                    j += 1
                    break
                if sup is None:
                    sup = line
                else:
                    desc.append(line)
                j += 1
            if amt is not None and amt >= 20000 and sup:
                recs.append({"ref": ref, "supplier": sup, "amount": amt, "desc": " ".join(desc) or None, "paid": paid})
            i = j
        else:
            i += 1
    return recs


# ---- Tailte Éireann reading-order reader. Same shape as the Courts reader (amount-anchored,
# 'PO supplier' merged on one line, supplier sometimes wraps to a 2nd line) BUT with 13-digit PO
# numbers (Courts uses 4-7) and a leading € on the amount — so it needs its own widened regexes.
# The generic reader recovered ~6 of 163 rows.
_TT_AMT = re.compile(r"^€?\s*([\d,]+\.\d{2})\b\s*(.*)$")
_TT_PO_NAME = re.compile(r"^(\d{4,13})\s+(.+)$")
_TT_PO_ONLY = re.compile(r"^\d{4,13}$")
_TT_PAID = re.compile(r"^(Y|N|Yes|No)$", re.I)


def read_tailte(b: bytes, max_pages: int | None) -> list[dict]:
    """Reading-order reader for the Tailte Éireann PO PDFs → list of
    {ref, supplier, amount, desc, paid} dicts (period from the file URL)."""
    doc = fitz.open(stream=b, filetype="pdf")
    limit = min(doc.page_count, max_pages) if max_pages else doc.page_count
    lines: list[str] = []
    for pi in range(limit):
        for raw in doc[pi].get_text().splitlines():
            s = re.sub(r"\s+", " ", raw.replace("\xa0", " ").replace("�", " ")).strip()
            if not s:
                continue
            low = s.lower()
            if low.startswith("purchase orders") or "incl vat" in low or "please note" in low:
                continue
            if low in {"po", "supplier", "total", "description", "paid", "(incl vat)"}:
                continue
            if re.match(r"(?i)^total\b\s*[:€]?\s*[\d,]", s):
                continue
            lines.append(s)
    doc.close()

    n = len(lines)
    recs: list[dict] = []
    rec_start = 0
    i = 0
    while i < n:
        am = _TT_AMT.match(lines[i])
        if not am:
            i += 1
            continue
        amount = float(am.group(1).replace(",", ""))
        head = lines[rec_start:i]
        j = i + 1
        desc: list[str] = []
        if am.group(2).strip():
            desc.append(am.group(2).strip())
        paid = None
        while j < n and not _TT_AMT.match(lines[j]):
            if _TT_PAID.match(lines[j]):
                paid = lines[j]
                j += 1
                break
            if _TT_PO_NAME.match(lines[j]) or _TT_PO_ONLY.match(lines[j]):
                break
            desc.append(lines[j])
            j += 1
        headtext = " ".join(head).strip()
        last = None
        for mm in re.finditer(r"(\d{4,13})\s+", headtext):
            if re.search(r"[A-Za-z]", headtext[mm.end() :]):
                last = mm
        if last is not None:
            po, supplier = last.group(1), headtext[last.end() :].strip()
        else:
            po, supplier = None, headtext
        if (
            supplier
            and re.search(r"[A-Za-z]", supplier)
            and amount >= 20000
            and not re.fullmatch(r"(?i)\s*(grand\s+)?(sub[-\s]*)?totals?\b[\s:.€\d,]*", supplier)
        ):
            recs.append(
                {
                    "ref": po,
                    "supplier": supplier,
                    "amount": amount,
                    "desc": " ".join(desc).strip() or None,
                    "paid": paid,
                }
            )
        rec_start = j
        i = j
    return recs


# ---- DPER / OGCIO reading-order reader. Two layout families under one publisher:
#   OGCIO   : PO(13-digit) / supplier / PO-date / amount / description / paid   (6 lines)
#   DPENDR  : PO(13-digit) / supplier / '€amount description' / paid            (no date, amount inline)
# plus 'PO supplier' merged on one line in some quarters. Amount-anchored (Courts/Tailte family)
# with a 13-digit PO and a trailing-date strip; naturally drops the file's grand-total line (no
# supplier) that the generic word-geometry reader ingested as a phantom payment. coverage_qa: OGCIO
# was 58.8% (191 of 325).
_DP_AMT = re.compile(r"^€?\s*([\d,]+\.\d{2})\b\s*(.*)$")
_DP_PO_NAME = re.compile(r"^(\d{10,})\s+(.+)$")
_DP_PO_ONLY = re.compile(r"^\d{10,}$")
_DP_DATE = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{4})\b")
_DP_PAID = re.compile(r"^(Y|N|Yes|No)$", re.I)
_DP_HDR = re.compile(
    r"^(po(\s|$)|po no|po number|po total|po amount|po date|total po value|supplier|supplier name|"
    r"description|paid|total$|purchase orders? for)",
    re.I,
)
_DP_BOILER = re.compile(r"please note|inclusive of vat|€20,000 or|for dpendr|for ogcio|for the dep", re.I)


def read_dper(b: bytes, max_pages: int | None) -> list[dict]:
    """Reading-order reader for the DPER/OGCIO PO PDFs → {ref, supplier, amount, desc, paid}
    dicts (period from the file URL). Handles the dated (OGCIO) and undated (DPENDR) layouts."""
    doc = fitz.open(stream=b, filetype="pdf")
    limit = min(doc.page_count, max_pages) if max_pages else doc.page_count
    lines: list[str] = []
    for pi in range(limit):
        for raw in doc[pi].get_text().splitlines():
            s = re.sub(r"\s+", " ", raw.replace("\xa0", " ").replace("�", " ")).strip()
            if not s or _DP_HDR.match(s) or _DP_BOILER.search(s.lower()):
                continue
            if re.match(r"(?i)^total\b\s*[:€]?\s*[\d,]", s):
                continue
            lines.append(s)
    doc.close()

    n = len(lines)
    recs: list[dict] = []
    rec_start = 0
    i = 0
    while i < n:
        am = _DP_AMT.match(lines[i])
        if not am or float(am.group(1).replace(",", "")) < 20000:
            i += 1
            continue
        amount = float(am.group(1).replace(",", ""))
        head = lines[rec_start:i]
        j = i + 1
        desc: list[str] = []
        if am.group(2).strip():
            desc.append(am.group(2).strip())
        paid = None
        while j < n and not _DP_AMT.match(lines[j]):
            if _DP_PAID.match(lines[j]):
                paid = lines[j]
                j += 1
                break
            if _DP_PO_NAME.match(lines[j]) or _DP_PO_ONLY.match(lines[j]):
                break
            desc.append(lines[j])
            j += 1
        headtext = " ".join(head).strip()
        po = None
        last = None
        for mm in re.finditer(r"(\d{10,})\s+", headtext):
            if re.search(r"[A-Za-z]", headtext[mm.end() :]):
                last = mm
        if last is not None:
            po, supplier = last.group(1), headtext[last.end() :].strip()
        elif head and _DP_PO_ONLY.match(head[-1]):
            po, supplier = head[-1], None
        else:
            supplier = headtext
        if supplier:  # strip a trailing PO-date that bled into the supplier (OGCIO layout)
            dm = _DP_DATE.search(supplier)
            if dm:
                supplier = supplier[: dm.start()].strip()
        if (
            supplier
            and re.search(r"[A-Za-z]", supplier)
            and not re.fullmatch(r"(?i)\s*(grand\s+)?(sub[-\s]*)?totals?\b[\s:.€\d,]*", supplier)
        ):
            recs.append(
                {
                    "ref": po,
                    "supplier": supplier,
                    "amount": amount,
                    "desc": " ".join(desc).strip() or None,
                    "paid": paid,
                }
            )
        rec_start = j
        i = j
    return recs


# ---- Dept of Culture reading-order reader. NO reference/PO column: each record is a supplier
# line then an amount line that carries the description inline ("42,903.39 National Archives"),
# on the next line, or not at all (3 sub-layouts across quarters). Amount-anchored: supplier is
# the line immediately above the amount; description is the inline tail plus any lines up to the
# next record's supplier. The generic word-geometry reader managed ~86% and lost whole files.
_CU_AMT = re.compile(r"^€?\s*([\d,]+\.\d{2})(?:\s+(.*\S))?\s*$")
_CU_HDR = re.compile(r"^(supplier|total\s*€?|description|paid|po\b|purchase orders?|payments?)", re.I)
_CU_BOILER = re.compile(r"please note|inclusive of vat|€20,000|for the dep|for 20|quarter", re.I)


def read_culture(b: bytes, max_pages: int | None) -> list[dict]:
    """Reading-order reader for the Dept of Culture PO PDFs (no PO column) →
    {ref, supplier, amount, desc} dicts (period from the file URL)."""
    doc = fitz.open(stream=b, filetype="pdf")
    limit = min(doc.page_count, max_pages) if max_pages else doc.page_count
    lines: list[str] = []
    for pi in range(limit):
        for raw in doc[pi].get_text().splitlines():
            s = re.sub(r"\s+", " ", raw.replace("\xa0", " ").replace("�", " ")).strip()
            if not s or _CU_HDR.match(s) or _CU_BOILER.search(s.lower()):
                continue
            if re.match(r"(?i)^total\b\s*[:€]?\s*[\d,]", s):
                continue
            lines.append(s)
    doc.close()

    amts = [
        i
        for i, ln in enumerate(lines)
        if _CU_AMT.match(ln) and float(_CU_AMT.match(ln).group(1).replace(",", "")) >= 20000
    ]
    recs: list[dict] = []
    for k, i in enumerate(amts):
        m = _CU_AMT.match(lines[i])
        amount = float(m.group(1).replace(",", ""))
        supplier = lines[i - 1] if i > 0 and not _CU_AMT.match(lines[i - 1]) else None
        desc: list[str] = []
        if m.group(2):
            desc.append(m.group(2).strip())
        nxt = amts[k + 1] if k + 1 < len(amts) else len(lines)
        for q in range(i + 1, max(i + 1, nxt - 1)):  # stop before the next record's supplier line
            if not _CU_AMT.match(lines[q]):
                desc.append(lines[q])
        if (
            supplier
            and re.search(r"[A-Za-z]", supplier)
            and not re.fullmatch(r"(?i)\s*(grand\s+)?(sub[-\s]*)?totals?\b[\s:.€\d,]*", supplier)
        ):
            recs.append({"ref": None, "supplier": supplier, "amount": amount, "desc": " ".join(desc).strip() or None})
    return recs


# DHLGH "procurement-related-payments-over-€20,000" PDFs publish in ≥5 sub-layouts across the
# years that the generic word-geometry reader cannot read (it yields 0 / scrambled rows because
# the "Payment Amount" header wraps to two lines and the column ORDER varies): (A) Date|Supplier|
# Description|Amount with the amount far right; (B) the amount column sits BETWEEN supplier and
# description (2014-15); (C) amount+description merged in one cell ("20,595.33 ICT Helpdesk");
# (D) date+supplier merged in one cell ("17/07/2024 BDO EATON SQUARE LTD"); and (E) a 2024
# split-column layout where the Amount column is rendered on its OWN pages after the text pages.
# Rather than hard-code each, derive the columns from the DATA line x0s (gap clustering), label
# them by content (dates -> date col, leading-money -> amount col; the remaining columns split
# left=supplier / right=description, unless the supplier rides in the date cell), peel money
# merged into a cell, and zip amount-only pages onto the text records in order.
_HS_DATE = re.compile(r"\b\d{2}/\d{2}/\d{4}\b")
_HS_MONEY = re.compile(r"-?(?:€\s*)?\d{1,3}(?:,\d{3})*\.\d{2}\b")
_HS_MONEY0 = re.compile(r"^\s*-?(?:€\s*)?\d{1,3}(?:,\d{3})*\.\d{2}\b")
_HS_TITLE = re.compile(
    r"payment date|supplier|description|amount|payments? (?:of|for) €|department of|for the period",
    re.I,
)


def read_housing(b: bytes, max_pages: int | None) -> list[dict]:
    """Data-derived-column reader for the DHLGH payments PDFs → {ref, supplier, amount, desc,
    date} dicts (per-row Payment Date). Handles the five sub-layouts described above."""
    doc = fitz.open(stream=b, filetype="pdf")
    limit = min(doc.page_count, max_pages) if max_pages else doc.page_count

    def page_lines(page):
        d = page.get_text("dict")
        out = []
        for blk in d.get("blocks", []):
            for ln in blk.get("lines", []):
                txt = "".join(s["text"] for s in ln.get("spans", [])).strip()
                if txt:
                    out.append((ln["bbox"][1], ln["bbox"][0], txt))
        out.sort(key=lambda t: (round(t[0], 1), t[1]))
        return out  # (y0, x0, text)

    def columns(x0s, gap=22):
        xs = sorted(x0s)
        if not xs:
            return []
        cols = [[xs[0]]]
        for x in xs[1:]:
            (cols.append([x]) if x - cols[-1][-1] > gap else cols[-1].append(x))
        return [sum(c) / len(c) for c in cols]

    def col_of(centers, x):  # index of the nearest column centre
        return min(range(len(centers)), key=lambda i: abs(centers[i] - x))

    out: list[dict] = []
    orphan_amounts: list[str] = []  # amounts from amount-only pages (split-column layout)
    for pi in range(limit):
        lines = [
            (y, x, t)
            for (y, x, t) in page_lines(doc[pi])
            if not (_HS_TITLE.search(t) and not _HS_DATE.search(t) and not _HS_MONEY.search(t))
        ]
        if not lines:
            continue
        has_date = any(_HS_DATE.match(t.strip()) for _y, _x, t in lines)
        has_money = any(_HS_MONEY0.match(t) for _y, _x, t in lines)
        if has_money and not has_date:  # amount-only page: collect in reading order
            orphan_amounts += [_HS_MONEY.search(t).group(0) for _y, _x, t in lines if _HS_MONEY0.match(t)]
            continue
        centers = columns([x for _y, x, _t in lines])
        if len(centers) < 2:
            continue
        ndate = [0] * len(centers)
        nmoney = [0] * len(centers)
        for _y, x, t in lines:
            ci = col_of(centers, x)
            if _HS_DATE.match(t.strip()):
                ndate[ci] += 1
            if _HS_MONEY0.match(t):
                nmoney[ci] += 1
        date_col = max(range(len(centers)), key=lambda i: ndate[i]) if any(ndate) else 0
        amt_cands = [i for i in range(len(centers)) if i != date_col]
        amt_col = max(amt_cands, key=lambda i: nmoney[i]) if amt_cands and any(nmoney) else None
        rest = sorted(i for i in range(len(centers)) if i not in (date_col, amt_col))
        # supplier in the date cell ("DD/MM/YYYY SUPPLIER")? then no separate supplier column.
        n_date = n_rem = 0
        for _y, x, t in lines:
            md = _HS_DATE.match(t.strip())
            if md and col_of(centers, x) == date_col:
                n_date += 1
                if t.strip()[md.end() :].strip():
                    n_rem += 1
        if n_date and n_rem > 0.5 * n_date:
            sup_col, desc_cols = None, set(rest)
        else:
            sup_col, desc_cols = (rest[0] if rest else None), set(rest[1:])

        cur = None
        for _y, x, t in lines:
            ci = col_of(centers, x)
            md = _HS_DATE.match(t.strip())
            if ci == date_col and md:
                if cur and (cur["amt"] is not None or cur["sup"] or cur["desc"]):
                    out.append(cur)
                cur = {"date": md.group(0), "sup": [], "desc": [], "amt": None}
                rem = t.strip()[md.end() :].strip()
                if rem:
                    cur["sup"].append(rem)
            elif cur is not None:
                m = _HS_MONEY.search(t)
                if ci == amt_col and m:
                    cur["amt"] = m.group(0)
                    tail = (t[: m.start()] + " " + t[m.end() :]).strip()
                    if tail:
                        cur["desc"].append(tail)
                elif ci == sup_col:
                    cur["sup"].append(t)
                elif ci in desc_cols:
                    cur["desc"].append(t)
                elif m and cur["amt"] is None:
                    cur["amt"] = m.group(0)
        if cur and (cur["amt"] is not None or cur["sup"] or cur["desc"]):
            out.append(cur)
    doc.close()

    if orphan_amounts:  # split-column layout: fill stub records from the amount-only pages
        # non-strict: an amount-only page may carry one extra grand-total beyond the stub count
        for r, a in zip([r for r in out if r["amt"] is None], orphan_amounts, strict=False):
            r["amt"] = a
    recs: list[dict] = []
    for r in out:
        if r["amt"] is None:
            continue
        try:
            amount = float(r["amt"].replace("€", "").replace(",", "").strip())
        except ValueError:
            continue
        recs.append(
            {
                "ref": None,
                "supplier": " ".join(r["sup"]).strip() or None,
                "amount": amount,
                "desc": " ".join(r["desc"]).strip() or None,
                "date": r["date"],
            }
        )
    return recs


# Generic amount-anchored reading-order FALLBACK for publishers whose PDFs put each field on its
# OWN line (so the word-geometry reader finds no column grid and yields 0 rows): ie_bim, ie_lmetb,
# the ie_atu q2-2022 quarter, and any future look-alike. Two record shapes, detected per amount:
#   amount-LAST   : [ref/code?] [supplier] [description?] AMOUNT   (bim "Supplier/Desc/Amount", lmetb)
#   amount-MIDDLE : [PO supplier] then "AMOUNT description"        (atu q2-2022)
# Wired as a FALLBACK (only when the geometry reader got nothing), so it can never change a file
# that already parses — no per-publisher reader flag and no regression risk.
_ROF_HDR = re.compile(
    r"^(supplier|description|amount|por\b|po number|account|code|base amount|date|vat|"
    r"payments?\s+(?:over|of)|excluding|vendor name|total paid|supplier id|order date|order no|"
    r"vat inc|name)\b",
    re.I,
)
_ROF_REF = re.compile(r"^(?:POR[-\s]?\d+|[A-Z]{1,4}\d{3,}|\d{4,})$")  # POR-x / account codes / PO ints
_ROF_BOILER = re.compile(r"^payments?\s+(?:over|of)\b|^\s*page\b|^\s*(grand\s+)?total\b", re.I)
_ROF_DATE = re.compile(r"^\d{1,2}/\d{1,2}/\d{2,4}\b")
_ROF_PO_LEAD = re.compile(r"^\d{4,}\s+(\S.*)$")  # "178186 Kedington Ltd"
_ROF_MONEY_PURE = re.compile(r"^€?\s*\d{1,3}(?:,\d{3})*\.\d{2}$|^€?\s*\d+\.\d{2}$")
_ROF_MONEY_LEAD = re.compile(r"^€?\s*(\d{1,3}(?:,\d{3})*\.\d{2}|\d+\.\d{2})\s+(\S.*)$")


def read_pdf_reading_order_fallback(b: bytes, max_pages: int | None, min_eur: float = 20000.0) -> list[dict]:
    """Amount-anchored reading-order reader → {ref, supplier, amount, desc} dicts. See the comment
    above for the two layouts handled. Returns [] for image/columnar PDFs (no usable text lines)."""
    doc = fitz.open(stream=b, filetype="pdf")
    limit = min(doc.page_count, max_pages) if max_pages else doc.page_count
    lines = [
        re.sub(r"\s+", " ", ln.replace("\xa0", " ").replace(" ", " ").replace("�", " ")).strip()
        for pi in range(limit)
        for ln in doc[pi].get_text().splitlines()
    ]
    doc.close()
    lines = [ln for ln in lines if ln]
    recs: list[dict] = []
    prev_amt = -1
    for i, ln in enumerate(lines):
        mp = _ROF_MONEY_PURE.match(ln)
        ml = None if mp else _ROF_MONEY_LEAD.match(ln)
        if mp:
            amt = float(re.sub(r"[^\d.]", "", ln))
            if amt >= min_eur:
                txt = []
                for x in lines[prev_amt + 1 : i]:  # text between the previous amount and this one
                    if _ROF_REF.match(x) or _ROF_BOILER.search(x) or (_ROF_HDR.match(x) and len(x) < 30):
                        continue
                    md = _ROF_DATE.match(x)
                    x = x[md.end() :].strip() if md else x  # strip a leading "DD/MM/YYYY supplier" date
                    if x:
                        txt.append(x)
                if txt:
                    recs.append({"ref": None, "supplier": txt[0], "amount": amt, "desc": " ".join(txt[1:]) or None})
            prev_amt = i
        elif ml:
            amt = float(ml.group(1).replace("€", "").replace(",", "").strip())
            if amt >= min_eur:
                sup = lines[i - 1] if i > 0 else None
                if sup and not (_ROF_HDR.match(sup) and len(sup) < 30):
                    pm = _ROF_PO_LEAD.match(sup)
                    recs.append(
                        {
                            "ref": None,
                            "supplier": pm.group(1) if pm else sup,
                            "amount": amt,
                            "desc": ml.group(2).strip() or None,
                        }
                    )
            prev_amt = i

    # Garble guard: a column-split / OCR-mangled PDF (LMETB's "CGAO36" quarters: codes, suppliers and
    # amounts each clustered in their own block) yields rows whose "suppliers" are numbers — the line
    # picked before each amount is another amount. OCR also swaps digits for letters ("L95,75r.40"),
    # so test the DIGIT RATIO, not letter-presence: a real name is not 40%+ digits. If most suppliers
    # are digit-heavy, the file is unreadable from its text layer (needs re-OCR); emit nothing.
    def _digit_heavy(s: str) -> bool:
        return bool(s) and sum(c.isdigit() for c in s) > 0.4 * len(s)

    named = [r["supplier"] for r in recs if r["supplier"]]
    if len(named) >= 5 and sum(_digit_heavy(s) for s in named) > 0.6 * len(named):
        return []
    return recs


# ---- XLSX / XLS / CSV ----
def _tabular_from_raw(raw: list[list]):
    """Shared header-pick + body-trim for any 2D cell grid (openpyxl or xlrd)."""
    full = " ".join(str(c) for row in raw[:6] for c in row if c is not None)

    def score(row):
        # Count SHORT role-matching cells only: a real header is short labels ("Vendor Name",
        # "Amount"); a TITLE row above it is one long sentence with embedded keywords
        # ("CHI Vendor payments >25K (incl VAT)") that should NOT win header detection.
        return sum(
            1
            for c in (row or [])
            if c is not None and len(str(c).strip()) <= 30 and any(rx.search(str(c)) for rx in ROLE_RE.values())
        )

    # Tie-break toward the LATER row — a title/banner row precedes the real header.
    hi = max(range(min(8, len(raw))), key=lambda i: (score(raw[i]), i), default=0)
    header = [str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(raw[hi])]
    rows = [r for r in raw[hi + 1 :] if any(c is not None and str(c).strip() for c in r)]
    return header, rows, full


def read_xlsx(b: bytes):
    import openpyxl

    ws = openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True).active
    raw = [list(r) for r in ws.iter_rows(values_only=True)]
    return _tabular_from_raw(raw)


def read_xls(b: bytes):
    import xlrd  # legacy binary .xls (pre-2021 quarterlies); openpyxl is .xlsx-only

    sh = xlrd.open_workbook(file_contents=b).sheet_by_index(0)
    raw = [sh.row_values(i) for i in range(sh.nrows)]
    return _tabular_from_raw(raw)


def read_csv(b: bytes):
    # gov.ie CSVs frequently carry a TITLE + blank preamble before the real header row, e.g. Social
    # Protection "Purchase Orders over €20,000 - Quarter 1, 2026" / blank / "Supplier name,Order
    # type,Order No,Order amount", or Health "Payments €20,000 or above…" / blank / blank /
    # "Reference,Payee Name,Tran Value,Description,Paid Y/N". polars' default has_header=True takes
    # the title as the header and the real columns get lost. Read header-less and reuse the shared
    # title-skipping header detector (same as xlsx/xls).
    # Decode cp1252/UTF-8 robustly BEFORE Polars sees it (see shared.text_encoding): the old
    # encoding="utf8-lossy" turned every cp1252 '€'/fada/apostrophe into '�' (e.g. TII
    # "Éamonn Conlon SC", "Signs Programme – works"). Hand Polars clean UTF-8.
    df = pl.read_csv(
        io.BytesIO(decode_table_bytes(b).encode("utf-8")),
        has_header=False,
        infer_schema_length=0,
        truncate_ragged_lines=True,
        ignore_errors=True,
    )
    raw = [list(r) for r in df.iter_rows()]
    return _tabular_from_raw(raw)


# A header that carries a money-ish KEYWORD but is really an identifier or a count, never the
# amount: a COUNT column (Beaumont "No. of Payments > €20,000", values 1..300) OR an ID/reference
# column (DETE "Payment Number" — matches the amount regex via 'payment', but holds PO refs like
# 137014941 that look like €137m). Excluded from amount detection so the real money column
# ("Total"/"Value"/"Amount") wins; a true amount column is never titled "...Number/Reference".
NON_AMOUNT_HDR = re.compile(
    r"no\.?\s*of\b|number\b|\bnumbers\b|\bcount\b|\bqty\b|quantit|reference\b|\bref\b|\bid\b|"
    # an "<X> No." identifier column (DETE PDF "Payment No.", "PO No.", "Order No.", "Invoice No.")
    r"(?:payment|order|invoice|po|p\.?o\.?|doc|transaction)\s*no\.?\b",
    re.I,
)


def detect_roles_tab(header, rows):
    roles = {k: None for k in ROLE_RE}
    for role, rx in ROLE_RE.items():
        cands = [i for i, h in enumerate(header) if rx.search(h or "")]
        if not cands:
            continue
        if role == "amount":
            # Drop count/identifier columns that carry money keywords in their label (see
            # NON_AMOUNT_HDR) before ranking by numeric density.
            strong = [i for i in cands if not NON_AMOUNT_HDR.search(header[i] or "")]
            cands = strong or cands
            # Drop date/period columns that also match the amount regex. TII's header is
            # "Period Paid,PAYMENT,…": 'Period Paid' matches amount via 'paid', and its
            # 'Jan-21' cells parse (to_eur) to 21, which TIES the numeric-density test below
            # and stole the amount role from the real 'PAYMENT' column on column order — every
            # row's amount became €21 and the real €-millions were discarded. A column better
            # classified as period is never the money column when a non-period candidate exists.
            non_period = [i for i in cands if not ROLE_RE["period"].search(header[i] or "")]
            cands = non_period or cands
            cands.sort(
                key=lambda i: (
                    -(sum(to_eur(r[i]) is not None for r in rows[:200] if i < len(r)) / max(1, len(rows[:200])))
                )
            )
        roles[role] = cands[0]

    # A category/description column is often headed "Payment Type" / "Order Type" (OPW, several
    # gov.ie bodies): the `paid` regex claims it (its pattern includes "payment type") and the
    # `description` regex — which has no "type" keyword — leaves description empty, so the useful
    # category text (Software, Roofworks, Building Maintenance) is discarded. When the column the
    # `paid` role grabbed actually holds free-text categories rather than a Y/N flag or a payment
    # count, promote it to description (and clear `paid`, so "Software" never lands in paid_flag).
    def _looks_descriptive(i: int) -> bool:
        vals = [str(r[i]).strip() for r in rows[:200] if i < len(r) and r[i] not in (None, "")]
        if len(vals) < 3:
            return False
        if sum(to_eur(v) is not None for v in vals) / len(vals) > 0.3:  # a numeric flag/count, not text
            return False
        has_letters = sum(any(c.isalpha() for c in v) for v in vals) / len(vals)
        distinct = len({v.lower() for v in vals})
        # free-text categories vary (Software/Roofworks/…); a Paid/Not-Paid flag has ≤2 values.
        return has_letters > 0.7 and distinct > 2

    if roles["description"] is None and roles["paid"] is not None and _looks_descriptive(roles["paid"]):
        roles["description"], roles["paid"] = roles["paid"], None
    return roles
