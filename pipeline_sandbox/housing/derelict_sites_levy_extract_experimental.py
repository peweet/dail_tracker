"""Derelict Sites Levy — per-local-authority enforcement & collection (2024).

The council-administered Derelict Sites Levy (Derelict Sites Act 1990): how much
each local authority LEVIED vs how much it actually COLLECTED, and the cumulative
amount left OUTSTANDING. The enforcement-gap dataset — many councils levy little and
collect less, so €26m+ sits uncollected nationally.

Source: Dept of Housing, LG & Heritage annual return (gov.ie), one consolidated
XLSX per year, per-LA. CC-BY-4.0 (Irish PSI / data.gov.ie licence).
  Landing: gov.ie/.../annual-returns-for-2024-received-...-derelict-sites-act-1990/
  File   : assets.gov.ie/static/documents/866f4b20/2024_Derelict_Sites_Statistics.xlsx
The gov.ie CDN 403s a bare request — a browser User-Agent + gov.ie Referer are needed.

Reads  : doc/source_pdfs/2024_Derelict_Sites_Statistics.xlsx  (cached source, git-tracked)
Writes : data/gold/parquet/derelict_sites_levy_wide.parquet   (with --write [--gold])

NOTE: sandbox/experimental — sandbox->vet->promote. 2024 only (the Dept publishes
one-shot annual XLSXs; earlier years are PQ/PDF-only — see doc note). Re-fetch each
Q2 with --download as a new year's file appears.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
except Exception:
    pass

_ROOT = Path(__file__).resolve().parents[2]
_SRC = _ROOT / "doc" / "source_pdfs" / "2024_Derelict_Sites_Statistics.xlsx"
_OUT = _ROOT / "data" / "gold" / "parquet"
_URL = "https://assets.gov.ie/static/documents/866f4b20/2024_Derelict_Sites_Statistics.xlsx"
_YEAR = 2024

# XLSX column index -> output field (header row is row 2 in the sheet).
_COLS = {
    0: "la",
    1: "notices_issued",
    6: "sites_on_register_end",
    9: "sites_levied",
    10: "amount_levied_eur",
    11: "amount_received_levied_eur",
    12: "total_received_eur",
    13: "cumulative_outstanding_eur",
}


def _download() -> None:
    import requests

    h = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)", "Referer": "https://www.gov.ie/"}
    r = requests.get(_URL, headers=h, timeout=60)
    r.raise_for_status()
    _SRC.parent.mkdir(parents=True, exist_ok=True)
    _SRC.write_bytes(r.content)
    print(f"Downloaded {len(r.content):,} bytes -> {_SRC.relative_to(_ROOT)}")


def _clean_la(s: str) -> str:
    n = re.sub(r"\s+", " ", str(s or "")).strip()
    n = n.replace("Dún", "Dun").replace(" & ", " and ")
    n = n.replace("Dun Laoghaire Rathdown", "Dun Laoghaire-Rathdown")
    return n


def _num(v):
    if v is None:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    if not s or s in {"-", ".", "-."}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract() -> tuple[pl.DataFrame, dict]:
    wb = load_workbook(_SRC, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    records, totals = [], {}
    for r in rows[2:]:
        if not r or r[0] is None:
            continue
        name = str(r[0]).strip()
        if name.lower() == "total":
            totals = {f: _num(r[i]) for i, f in _COLS.items() if i != 0}
            continue
        rec = {"year": _YEAR}
        for i, f in _COLS.items():
            rec[f] = _clean_la(r[i]) if f == "la" else _num(r[i])
        records.append(rec)
    df = pl.DataFrame(records)
    # column order: la, year, then the rest
    cols = ["la", "year"] + [f for f in _COLS.values() if f != "la"]
    return df.select(cols), totals


def fidelity_check(df: pl.DataFrame, totals: dict) -> dict:
    rpt = {"checks": {}, "rows": df.height}
    if df.is_empty():
        rpt["green"] = False
        return rpt
    rpt["checks"]["1_la_coverage"] = {"unique_LAs": df["la"].n_unique(), "pass": df["la"].n_unique() == 31}
    # the per-LA sums must reconcile to the file's own Total row (catches drops/dupes)
    recon = {}
    for f in ("amount_levied_eur", "total_received_eur", "cumulative_outstanding_eur"):
        got = df[f].sum()
        want = totals.get(f)
        recon[f] = {"sum": got, "file_total": want, "ok": want is not None and abs(got - want) < 1}
    rpt["checks"]["2_reconciles_to_total"] = {**recon, "pass": all(v["ok"] for v in recon.values())}
    # non-negative money
    bad = sum(df.filter(pl.col(f) < 0).height for f in ("amount_levied_eur", "cumulative_outstanding_eur"))
    rpt["checks"]["3_non_negative"] = {"negatives": bad, "pass": bad == 0}
    rpt["green"] = all(c.get("pass", False) for c in rpt["checks"].values())
    return rpt


def _write_parquet(df: pl.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(path, compression="zstd", compression_level=3, statistics=True)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--download", action="store_true", help="re-fetch the XLSX from gov.ie")
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--gold", action="store_true", help="write to data/gold/parquet (else _noac_eval)")
    args = ap.parse_args()

    if args.download:
        _download()
    if not _SRC.exists():
        print(f"ERROR: source missing: {_SRC} (run with --download)")
        sys.exit(1)

    df, totals = extract()
    rpt = fidelity_check(df, totals)
    print(f"=== derelict_sites_levy_wide — {df.height} rows ===")
    for n, chk in rpt["checks"].items():
        print(f"  [{'GREEN' if chk.get('pass') else 'FAIL'}] {n}: {chk}")
    print(f"  >>> {'GREEN' if rpt['green'] else 'AMBER'}")
    if not df.is_empty():
        nat_lev = df["amount_levied_eur"].sum()
        nat_out = df["cumulative_outstanding_eur"].sum()
        print(f"  national: levied €{nat_lev:,.0f} | outstanding €{nat_out:,.0f}")
        with pl.Config(tbl_rows=5, tbl_width_chars=160):
            print(df.sort("cumulative_outstanding_eur", descending=True)
                  .select(["la", "amount_levied_eur", "total_received_eur", "cumulative_outstanding_eur"]).head(5))
    if args.write and rpt["green"]:
        out = (_ROOT / "data" / "gold" / "parquet") if args.gold else (_OUT / "_noac_eval")
        path = out / "derelict_sites_levy_wide.parquet"
        _write_parquet(df, path)
        print(f"  Wrote {path.relative_to(_ROOT)}")
    elif not args.write:
        print("\n(dry-run — pass --write [--gold] to persist)")


if __name__ == "__main__":
    main()
