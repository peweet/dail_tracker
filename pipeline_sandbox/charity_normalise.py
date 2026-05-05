#!/usr/bin/env python3
"""
Charities Public Register normaliser (sandbox).

STATUS: SANDBOX. Self-contained — does not import from pipeline.py / enrich.py /
normalise_join_key.py. Reads bronze xlsx, writes four silver parquets.

Implements the charity half of CRO/INTEGRATION_PLAN.md §10 (NORMALISE per source).

INPUTS:
- data/bronze/charities/public_register_<DATE>.xlsx
  (Sheet "Disclaimer", "Public Register", "Annual Reports". Row 1 is the
  effective-date metadata band; row 2 is the column headers.)

OUTPUTS (silver):
- data/silver/charities/register.parquet         (one row per charity, RCN PK)
- data/silver/charities/annual_reports.parquet   (one row per (RCN, period_end))
- data/silver/charities/charity_latest.parquet   (one row per RCN — latest period)
- data/silver/charities/trustees_long.parquet    (one row per parsed trustee token)

WHY OPENPYXL DIRECTLY:
- pl.read_excel(engine="openpyxl") chokes on the Effective-Date metadata in
  row 1 (Series name must be a string). xlsx2csv / fastexcel are not pinned in
  this venv. Reading via openpyxl is deterministic and the file is small.
- read_only=True + values_only=True is fast (~15s for both sheets). The xlsx
  contains many trailing empty rows; we break on the first all-null row.

CLEANS:
- "CRO Number" = 0 → null (placeholder)
- "CRO Number" = RCN value → null (self-reference data-quality issue)
- Strip whitespace from all string columns
- Parse "Charity Classification: Primary [Secondary (Sub)]" into three columns
- Heuristic county = comma-split address, second-from-last token; "Unknown"
  when ambiguous
- Trustees (Start Date) parsed via regex into long table; unparseable tokens
  preserved with parse_quality='raw'

DERIVES:
- name_norm / aka_norm: same rule as cro_normalise.py (sandbox-isolated copy)
- period_year: from Period End Date
- gov_share: (gov_or_la_income + other_public_bodies_income) / gross_income,
  null when gross_income is null/0
- employees_band: text passthrough — never numeric
- funding_profile per charity_latest: state_funded | mostly_donations |
  mostly_trading | mixed | undisclosed
- state_adjacent_flag: gov_share >= 0.80 AND gross_income >= 100_000_000
- Warning flags surfaced on charity_latest for the lobbyist POC view:
    charity_filing_overdue_flag    period_end_latest < today − 18m
    charity_deficit_latest_flag    surplus_deficit_latest < 0
    charity_insolvent_latest_flag  total_liabilities_latest > total_assets_latest

USAGE:
    python pipeline_sandbox/charity_normalise.py
    python pipeline_sandbox/charity_normalise.py --bronze data/bronze/charities/public_register_20260426.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook

DEFAULT_BRONZE = Path("data/bronze/charities/public_register_20260426.xlsx")
DEFAULT_SILVER_DIR = Path("data/silver/charities")

LEGAL_SUFFIX_PATTERN = (
    r"\b(?:THE|LIMITED|LTD|DAC|PLC|CLG|UC|COMPANY|"
    r"DESIGNATED ACTIVITY COMPANY|"
    r"COMPANY LIMITED BY GUARANTEE|"
    r"UNLIMITED COMPANY|GROUP|HOLDINGS|IRELAND|IRL|OF)\b"
)

REGISTER_RENAME = {
    "Registered Charity Number": "rcn",
    "Registered Charity Name": "registered_charity_name",
    "Also Known As": "also_known_as",
    "Status": "status",
    "Charity Classification: Primary [Secondary (Sub)]": "classification_raw",
    "Primary Address": "primary_address",
    "Also Operates In": "also_operates_in",
    "Governing Form": "governing_form",
    "CRO Number": "cro_number_raw",
    "Country Established": "country_established",
    "Charitable Purpose": "charitable_purpose",
    "Charitable Objects": "charitable_objects",
    "Trustees (Start Date)": "trustees_raw",
}

ANNUAL_RENAME = {
    "Registered Charity Number (RCN)": "rcn",
    "Registered Charity Name": "registered_charity_name",
    "Period Start Date": "period_start_date",
    "Period End Date": "period_end_date",
    "Report Activity": "report_activity",
    "Activity Description": "activity_description",
    "Beneficiaries": "beneficiaries",
    "Income: Government or Local Authorities": "income_govt_or_la",
    "Income: Other Public Bodies": "income_other_public_bodies",
    "Income: Philantrophic Organisations": "income_philanthropic_orgs",
    "Income: Donations": "income_donations",
    "Income: Trading and Commercial Activities": "income_trading",
    "Income: Other Sources": "income_other",
    "Income: Bequests": "income_bequests",
    "Gross Income": "gross_income",
    "Gross Expenditure": "gross_expenditure",
    "Surplus / (Deficit) for the Period": "surplus_deficit",
    "Cash at Hand and in Bank": "cash_at_hand",
    "Other Assets": "other_assets",
    "Total Assets": "total_assets",
    "Total Liabilities": "total_liabilities",
    "Net Assets / (Liabilities)": "net_assets",
    "Gross Income (Schools)": "gross_income_schools",
    "Gross Expenditure (Schools)": "gross_expenditure_schools",
    "Number of Employees": "employees_band",
    "Number of Full-Time Employees": "employees_full_time",
    "Number of Part-Time Employees": "employees_part_time",
    "Number of Volunteers": "volunteers_band",
}

EMPLOYEES_BAND_ORDER = [
    "NONE", "1-9", "10-19", "20-49", "50-249",
    "250-499", "500-999", "1000-4999", "5000+",
]

CLASSIFICATION_PATTERN = re.compile(
    r"^(?P<primary>[^\[]+?)\s*"
    r"(?:\[(?P<secondary>[^\(\]]+?)\s*"
    r"(?:\((?P<sub>[^\)]+)\))?\s*\])?\s*$"
)

TRUSTEE_PATTERN = re.compile(
    r"^(?P<name>.+?)"
    r"(?:\s*\((?P<role>[^)]+?)\))?"
    r"\s*\((?P<dt>\d{1,2}/\d{1,2}/\d{4})\)\s*$"
)


# ---------------------------------------------------------------------------
# IO
# ---------------------------------------------------------------------------

def read_sheet(path: Path, sheet: str) -> dict[str, list[Any]]:
    """Read one sheet via openpyxl, returning a columnar dict ready for pl.DataFrame.

    Row 1 of these files is the effective-date metadata band; row 2 is the
    real column headers. Trailing empty rows (the regulator's xlsx ships with
    many) are cut off when both of the first two cells are null.

    Cell values are coerced per-column so Polars's schema inference doesn't
    trip over xlsx int/float ambiguity in money columns:
    - any column whose only non-null types are int/float is normalised to float
    - everything else is left as-is (strings, datetimes, None)
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"sheet {sheet!r} not in {path.name}: {wb.sheetnames}")
    ws = wb[sheet]
    iterator = ws.iter_rows(min_row=2, values_only=True)
    headers = next(iterator)
    headers = [h.replace("\n", " ").strip() if isinstance(h, str) else None for h in headers]
    columns: dict[str, list[Any]] = {h: [] for h in headers if h}
    keep = [(i, h) for i, h in enumerate(headers) if h]
    for r in iterator:
        if r[0] is None and r[1] is None:
            break
        for i, h in keep:
            columns[h].append(r[i] if i < len(r) else None)
    wb.close()

    for h, values in columns.items():
        non_null_types = {type(v) for v in values if v is not None}
        if non_null_types and non_null_types.issubset({int, float}):
            columns[h] = [float(v) if v is not None else None for v in values]
    return columns


# ---------------------------------------------------------------------------
# Expression helpers
# ---------------------------------------------------------------------------

def name_norm_expr(col: str) -> pl.Expr:
    return (
        pl.col(col)
        .str.to_uppercase()
        .str.replace_all(r"[\.,&'\"]", " ")
        .str.replace_all(LEGAL_SUFFIX_PATTERN, " ")
        .str.replace_all(r"[^A-Z0-9 ]", " ")
        .str.replace_all(r"\s+", " ")
        .str.strip_chars()
    )


def classification_split(col: str) -> list[pl.Expr]:
    pat = CLASSIFICATION_PATTERN.pattern
    return [
        pl.col(col).str.extract(pat, 1).str.strip_chars().alias("classification_primary"),
        pl.col(col).str.extract(pat, 2).str.strip_chars().alias("classification_secondary"),
        pl.col(col).str.extract(pat, 3).str.strip_chars().alias("classification_sub"),
    ]


def county_from_address_expr(col: str) -> pl.Expr:
    parts = (
        pl.col(col)
        .str.split(",")
        .list.eval(pl.element().str.strip_chars())
    )
    second_last = parts.list.get(-2, null_on_oob=True)
    return (
        pl.when(second_last.is_not_null() & (second_last != ""))
        .then(second_last)
        .otherwise(pl.lit("Unknown"))
        .alias("county")
    )


# ---------------------------------------------------------------------------
# Public register
# ---------------------------------------------------------------------------

def normalise_register(path: Path) -> pl.DataFrame:
    cols = read_sheet(path, "Public Register")
    missing = [c for c in REGISTER_RENAME if c not in cols]
    if missing:
        raise SystemExit(f"public register missing columns: {missing}")
    df = pl.DataFrame({REGISTER_RENAME[k]: cols[k] for k in REGISTER_RENAME}, strict=False)

    df = df.with_columns(
        pl.col("rcn").cast(pl.Int64, strict=False),
        pl.col("registered_charity_name").str.strip_chars(),
        pl.col("status").str.strip_chars(),
        pl.col("governing_form").str.strip_chars(),
        pl.col("primary_address").str.strip_chars(),
        pl.col("also_known_as").str.strip_chars(),
        pl.col("country_established").str.strip_chars(),
        pl.col("classification_raw").str.strip_chars(),
        pl.col("cro_number_raw").cast(pl.Int64, strict=False),
    )

    df = df.with_columns(
        pl.when((pl.col("cro_number_raw") == 0) | (pl.col("cro_number_raw") == pl.col("rcn")))
        .then(None)
        .otherwise(pl.col("cro_number_raw"))
        .alias("cro_number"),
        name_norm_expr("registered_charity_name").alias("name_norm"),
        name_norm_expr("also_known_as").alias("aka_norm"),
        *classification_split("classification_raw"),
        county_from_address_expr("primary_address"),
    )
    df = df.with_columns(
        pl.col("cro_number").is_not_null().alias("has_cro_number_flag"),
    )
    return df.drop("cro_number_raw")


# ---------------------------------------------------------------------------
# Annual reports
# ---------------------------------------------------------------------------

def _money(col: str) -> pl.Expr:
    return pl.col(col).cast(pl.Float64, strict=False)


def normalise_annual_reports(path: Path) -> pl.DataFrame:
    cols = read_sheet(path, "Annual Reports")
    missing = [c for c in ANNUAL_RENAME if c not in cols]
    if missing:
        raise SystemExit(f"annual reports missing columns: {missing}")
    df = pl.DataFrame({ANNUAL_RENAME[k]: cols[k] for k in ANNUAL_RENAME}, strict=False)

    money_cols = [
        "income_govt_or_la", "income_other_public_bodies", "income_philanthropic_orgs",
        "income_donations", "income_trading", "income_other", "income_bequests",
        "gross_income", "gross_expenditure", "surplus_deficit", "cash_at_hand",
        "other_assets", "total_assets", "total_liabilities", "net_assets",
        "gross_income_schools", "gross_expenditure_schools",
    ]
    df = df.with_columns(
        pl.col("rcn").cast(pl.Int64, strict=False),
        *[_money(c).alias(c) for c in money_cols],
        pl.col("employees_band").cast(pl.Utf8, strict=False).str.strip_chars().alias("employees_band"),
        pl.col("volunteers_band").cast(pl.Utf8, strict=False).str.strip_chars().alias("volunteers_band"),
    )

    period_start = (
        pl.col("period_start_date").cast(pl.Date, strict=False)
        if df.schema.get("period_start_date") in (pl.Datetime, pl.Date)
        else pl.col("period_start_date").cast(pl.Utf8).str.to_date(strict=False)
    )
    period_end = (
        pl.col("period_end_date").cast(pl.Date, strict=False)
        if df.schema.get("period_end_date") in (pl.Datetime, pl.Date)
        else pl.col("period_end_date").cast(pl.Utf8).str.to_date(strict=False)
    )
    df = df.with_columns(
        period_start.alias("period_start_date"),
        period_end.alias("period_end_date"),
    )
    df = df.with_columns(
        pl.col("period_end_date").dt.year().alias("period_year"),
        pl.when(pl.col("gross_income").is_null() | (pl.col("gross_income") <= 0))
        .then(None)
        .otherwise(
            (pl.col("income_govt_or_la").fill_null(0) + pl.col("income_other_public_bodies").fill_null(0))
            / pl.col("gross_income")
        )
        .alias("gov_share"),
    )
    return df


# ---------------------------------------------------------------------------
# Charity latest snapshot
# ---------------------------------------------------------------------------

def build_charity_latest(annual: pl.DataFrame) -> pl.DataFrame:
    latest = (
        annual.drop_nulls(["period_end_date"])
        .sort(["rcn", "period_end_date"], descending=[False, True])
        .unique(subset=["rcn"], keep="first")
        .rename({
            "gov_share": "gov_funded_share_latest",
            "gross_income": "gross_income_latest_eur",
            "employees_band": "employees_band_latest",
            "period_end_date": "period_end_latest",
            "period_year": "period_year_latest",
            "surplus_deficit": "surplus_deficit_latest",
            "total_assets": "total_assets_latest_eur",
            "total_liabilities": "total_liabilities_latest_eur",
        })
    )

    donation_share = pl.col("income_donations").fill_null(0) / pl.col("gross_income_latest_eur")
    trading_share = pl.col("income_trading").fill_null(0) / pl.col("gross_income_latest_eur")

    funding_profile = (
        pl.when(pl.col("gross_income_latest_eur").is_null() | (pl.col("gross_income_latest_eur") <= 0))
        .then(pl.lit("undisclosed"))
        .when(pl.col("gov_funded_share_latest").is_null())
        .then(pl.lit("undisclosed"))
        .when(pl.col("gov_funded_share_latest") >= 0.5)
        .then(pl.lit("state_funded"))
        .when(donation_share >= 0.5)
        .then(pl.lit("mostly_donations"))
        .when(trading_share >= 0.5)
        .then(pl.lit("mostly_trading"))
        .otherwise(pl.lit("mixed"))
    )

    state_adjacent = (
        pl.col("gov_funded_share_latest").fill_null(0) >= 0.80
    ) & (
        pl.col("gross_income_latest_eur").fill_null(0) >= 100_000_000
    )

    cutoff_18m = dt.date.today() - dt.timedelta(days=int(365 * 1.5))
    filing_overdue = (pl.col("period_end_latest") < pl.lit(cutoff_18m)).fill_null(False)
    deficit = (pl.col("surplus_deficit_latest") < 0).fill_null(False)
    insolvent = (
        pl.col("total_liabilities_latest_eur").is_not_null()
        & pl.col("total_assets_latest_eur").is_not_null()
        & (pl.col("total_liabilities_latest_eur") > pl.col("total_assets_latest_eur"))
    )

    return latest.with_columns(
        funding_profile.alias("funding_profile"),
        state_adjacent.alias("state_adjacent_flag"),
        filing_overdue.alias("charity_filing_overdue_flag"),
        deficit.alias("charity_deficit_latest_flag"),
        insolvent.alias("charity_insolvent_latest_flag"),
    ).select([
        "rcn",
        "period_end_latest",
        "period_year_latest",
        "gross_income_latest_eur",
        "gov_funded_share_latest",
        "employees_band_latest",
        "surplus_deficit_latest",
        "total_assets_latest_eur",
        "total_liabilities_latest_eur",
        "funding_profile",
        "state_adjacent_flag",
        "charity_filing_overdue_flag",
        "charity_deficit_latest_flag",
        "charity_insolvent_latest_flag",
    ])


# ---------------------------------------------------------------------------
# Trustees long
# ---------------------------------------------------------------------------

def parse_trustees(raw: str | None, rcn: int | None) -> list[dict[str, Any]]:
    if not raw or rcn is None:
        return []
    out: list[dict[str, Any]] = []
    for token in raw.split(";"):
        t = re.sub(r"\s+", " ", token).strip()
        if not t:
            continue
        m = TRUSTEE_PATTERN.match(t)
        if m:
            out.append({
                "rcn": rcn,
                "trustee_name": m["name"].strip(),
                "role": (m["role"] or "").strip() or None,
                "start_date_raw": m["dt"],
                "parse_quality": "strict",
                "raw_token": t,
            })
        else:
            out.append({
                "rcn": rcn,
                "trustee_name": None,
                "role": None,
                "start_date_raw": None,
                "parse_quality": "raw",
                "raw_token": t,
            })
    return out


def build_trustees_long(register: pl.DataFrame) -> pl.DataFrame:
    rows: list[dict[str, Any]] = []
    for r in register.select(["rcn", "trustees_raw"]).iter_rows(named=True):
        rows.extend(parse_trustees(r["trustees_raw"], r["rcn"]))
    if not rows:
        return pl.DataFrame(schema={
            "rcn": pl.Int64, "trustee_name": pl.Utf8, "role": pl.Utf8,
            "start_date_raw": pl.Utf8, "parse_quality": pl.Utf8, "raw_token": pl.Utf8,
            "trustee_name_norm": pl.Utf8, "start_date": pl.Date,
        })
    df = pl.from_dicts(rows)
    return df.with_columns(
        pl.col("start_date_raw").str.to_date("%d/%m/%Y", strict=False).alias("start_date"),
        name_norm_expr("trustee_name").alias("trustee_name_norm"),
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    p = argparse.ArgumentParser(description="Charities Public Register normaliser (sandbox)")
    p.add_argument("--bronze", type=Path, default=DEFAULT_BRONZE)
    p.add_argument("--silver-dir", type=Path, default=DEFAULT_SILVER_DIR)
    args = p.parse_args()

    if not args.bronze.exists():
        raise SystemExit(f"bronze input not found: {args.bronze}")
    args.silver_dir.mkdir(parents=True, exist_ok=True)

    print(f"[charity_normalise] reading {args.bronze}")
    register = normalise_register(args.bronze)
    annual = normalise_annual_reports(args.bronze)
    latest = build_charity_latest(annual)
    trustees = build_trustees_long(register)

    out_register = args.silver_dir / "register.parquet"
    out_annual = args.silver_dir / "annual_reports.parquet"
    out_latest = args.silver_dir / "charity_latest.parquet"
    out_trustees = args.silver_dir / "trustees_long.parquet"
    register.write_parquet(out_register, compression="zstd")
    annual.write_parquet(out_annual, compression="zstd")
    latest.write_parquet(out_latest, compression="zstd")
    trustees.write_parquet(out_trustees, compression="zstd")

    print(f"[charity_normalise] wrote {out_register}        rows={register.height}  cols={register.width}")
    print(f"[charity_normalise] wrote {out_annual}   rows={annual.height}  cols={annual.width}")
    print(f"[charity_normalise] wrote {out_latest}   rows={latest.height}  cols={latest.width}")
    print(f"[charity_normalise] wrote {out_trustees}    rows={trustees.height}  cols={trustees.width}")

    cro_present = int(register["has_cro_number_flag"].sum())
    deregistered = int(register.filter(pl.col("status").str.contains("(?i)deregister")).height)
    state_adj = int(latest["state_adjacent_flag"].sum())
    strict_trustees = int(trustees.filter(pl.col("parse_quality") == "strict").height)
    raw_trustees = int(trustees.filter(pl.col("parse_quality") == "raw").height)
    print("  register_with_cro_number:", f"{cro_present:,}")
    print("  register_deregistered:   ", f"{deregistered:,}")
    print("  charities_state_adjacent:", f"{state_adj:,}")
    print(f"  trustees_strict:         {strict_trustees:,}")
    print(f"  trustees_raw:            {raw_trustees:,}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
