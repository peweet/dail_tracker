"""Public-body PO/payment extractor -> silver public_payments_fact.

Promotes the two PRE-ETL sample readers into ONE config-driven extractor:
  - PDF header-anchored reader  <- sample_extract_procurement_pdf.py
  - XLSX/CSV tabular reader      <- sample_extract_procurement.py
  - gold conventions (name_norm, supplier_class, value-safe-to-sum, coverage JSON,
    zstd parquet)                <- procurement_etenders_extract.py

Wired into pipeline.py as the ``public_body_payments`` chain. Writes the SILVER fact
data/silver/parquet/public_payments_fact.parquet (lifted out of data/sandbox/ 2026-06-12);
the ``procurement_consolidate`` chain folds it into gold procurement_payments_fact. One row
per source line, with full provenance (plan PROCUREMENT_SEMISTATE_EXPANSION_PLAN.md Phase 5).

SCOPE / OWNERSHIP (multi-context split, 2026-06-03): this extractor owns the publishers
the GENERIC reader handles cleanly (Tier A + the corrected Tier C). HSE + Tusla are owned
by procurement_hse_tusla_parser.py (bespoke column-x specs — the generic reader misparses
them); local-authority POs and the BUDGET tier are owned by other context windows. All
emit THIS same schema so the layers union at promotion time.

PRIVACY: supplier_class -> privacy_status, and the quarantine IS APPLIED. Any row whose
supplier looks like a sole trader / individual (privacy_status=review_personal_data) is
marked public_display=False so it can never surface in a UI or be promoted. Rows are
RETAINED (nothing is dropped) for analysis/coverage — the gate is the display flag — and a
runtime invariant in main() refuses to write if any personal row is left displayable.
Classification errs toward over-quarantine (an org without a recognised company suffix is
treated as personal): the safe direction. coverage records privacy_quarantine_applied=true
with the suppressed-row count.

Run:
  ./.venv/Scripts/python.exe extractors/procurement_public_body_extract.py --list            # harvest-only (lock URLs)
  ./.venv/Scripts/python.exe extractors/procurement_public_body_extract.py --list --only ie_opw,ie_tii
  ./.venv/Scripts/python.exe extractors/procurement_public_body_extract.py                   # full ingest
  ./.venv/Scripts/python.exe extractors/procurement_public_body_extract.py --only ie_hse --max-files 2
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import io
import json
import re
import subprocess
import sys
import time
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import quote, unquote, urljoin, urlparse

import fitz  # PyMuPDF
import polars as pl
import requests

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from services.fetch_report import Breaker, FetchReport, classify_body, classify_exception, write_sentinel  # noqa: E402
from services.parquet_io import save_parquet  # noqa: E402

with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8")

from shared.name_norm import name_norm_expr  # noqa: E402

H = {"User-Agent": "Mozilla/5.0 (dail-tracker research probe)"}
TMP = Path("c:/tmp/procurement_publishers")
BRONZE = ROOT / "data/bronze/pdfs/public_body_procurement"
REPORT = FetchReport("public_body")
LAST_ERR: dict = {}  # set by fetch_bytes on failure, read by the download loop
OUT_FACT = ROOT / "data/silver/parquet/public_payments_fact.parquet"
OUT_COV = ROOT / "data/_meta/public_payments_coverage.json"
PARSER_VERSION = "0.1.0"

DATA_EXT = (".pdf", ".xlsx", ".xls", ".csv")

# ----------------------------------------------------------------------------- regexes
MONEY_RE = re.compile(r"(?:€|EUR)?\s?\d{1,3}(?:,\d{3})+(?:\.\d{2})?|\d+\.\d{2}")
NUM_RE = re.compile(r"\d[\d,]*(?:\.\d+)?")
HREF_RE = re.compile(r"""href\s*=\s*["']([^"']+)["']""", re.I)
DIGIT_PREFIX = re.compile(r"^(?:\d{3,}\s+){1,3}")  # strips a leading PO/vendor-ID run
# A multi-word alphabetic string (e.g. "AN POST", "AIRNAV IRELAND") — used to recover a supplier
# name that a mis-mapped column put into po_number while the supplier cell came out blank.
NAME_LIKE = re.compile(r"[A-Za-z]{2,}[A-Za-z .,&'/-]*\s[A-Za-z&]", re.I)
NUMERIC_NOISE = re.compile(r"\d{4,}|\d,\d{3}")  # a big/grouped number => category total, not a name
PERIOD_RE = re.compile(r"(?:^|[^0-9])(20[12]\d)(?:[^0-9]|$)")
QUARTER_RE = re.compile(r"q\s?([1-4])|quarter[\s_-]?([1-4])|qtr[\s_-]?([1-4])", re.I)

ROLE_RE = {
    # English + Irish (as Gaeilge) — some bodies (e.g. An Bord Pleanála) publish bilingual
    # headers: Soláthraí=supplier, Glanmhéid/Méid Comhlán=net/gross amount, Cur Síos=description,
    # Tagairt=reference, Dáta=date. English-only role regexes leave supplier=None on those.
    "supplier": re.compile(r"supplier|payee|vendor|provider|customer|recipient|\bname\b|soláthr", re.I),
    "amount": re.compile(
        r"amount|total|value|gross|\beuro\b|€|\bpaid\b|\bvat\b|ledger|payment\b|m[ée]id|mhéid|luach|comhlán", re.I
    ),
    "description": re.compile(
        r"descript|\bdesc\b|detail|categor|service|goods|nature|\bgl\b|main gl|cur síos|tuairisc|seirbhís|earraí", re.I
    ),
    "po": re.compile(
        r"\border\b|\bpo\b|\bpor\b|referen|\bref\b|\bnumber\b|invoice|\bdoc\b|transaction|tagairt|uimhir|ordú", re.I
    ),
    "period": re.compile(r"period|quarter|\bqtr\b|\bdate\b|\byear\b|posting|month|dáta|ráithe|bliain", re.I),
    "paid": re.compile(r"\bpaid\b|payment type|status|no\.? of payments|íoctha", re.I),
}
CAVEAT_RE = re.compile(r"\bvat\b|exclud|inclus|indicativ|not (a )?payment|net of|estimate|note:|please note", re.I)
# A company / organisation indicator: a legal form, a plurality word, or a business-activity
# stem that a lone private individual's name never carries. Two arms because the matching differs:
#   • WHOLE words (\b…\b) for short tokens that would over-match as prefixes ("uc" inside "UCD",
#     "co" inside "Connolly"): the legal forms and bare plurality words.
#   • STEMS (leading \b only, NO trailing \b) so an inflection matches: "engineer" must catch
#     "engineerING"/"engineerS", "consult" → "consultING/consultANTS/consultANCY", "technolog" →
#     "technologY/technologIES". The old single \b(…)\b pattern silently failed on every such
#     inflection (the trailing \b needs a word→non-word edge, which "engineerS" doesn't have), so
#     ARUP CONSULTING ENGINEERS / CREATIVE TECHNOLOGY etc. were misclassed sole-trader. Fixed
#     2026-06-13; the reclassifier in procurement_payments_consolidate.py carries the same vocab.
_CO_WORDS = (
    "ltd", "limited", "dac", "plc", "clg", "llp", "teo", "teoranta", "t/a", "uc", "inc", "llc",
    "gmbh", "co", "company", "companies", "group", "sons", "bros", "university", "college",
    "council", "hse", "board", "media", "hotel", "ireland",
)
_CO_STEMS = (
    "servic", "solution", "consult", "engineer", "architect", "surveyor", "solicitor", "barrister",
    "accountant", "advis", "contract", "construct", "develop", "enterprise", "industr", "technolog",
    "system", "software", "logistic", "distribut", "manufactur", "pharma", "diagnostic", "laborator",
    "healthcare", "medical", "insuranc", "assuranc", "management", "communicat", "telecom", "propert",
    "holding", "internation", "institut", "foundation", "partner", "associat", "incorporat", "corporat",
    "centre",
)
COMPANY_SUFFIX = re.compile(
    r"\b(?:" + "|".join(_CO_WORDS) + r")\b|&|\b(?:" + "|".join(_CO_STEMS) + r")",
    re.I,
)
FOREIGN_FORM = re.compile(
    r"\b(gmbh|s\.?a\.?|n\.?v\.?|s\.?a\.?s|s\.?p\.?a|inc|llc|\bpty\b|\bab\b|\bbv\b|\boy\b|srl|sl|sarl|aps|kft|ltda)\b",
    re.I,
)
# NOTE: national state agencies named "X Ireland" / "X Éireann" (e.g. Transport
# Infrastructure Ireland, Uisce Éireann) must be caught HERE — _pub is tested before
# _co, otherwise COMPANY_SUFFIX's bare "ireland" token misclassifies them as companies
# and their intergovernmental transfers leak into value_safe_to_sum. Add agencies as
# they surface as transfer recipients.
PUBLIC_BODY = re.compile(
    r"\b(county council|city council|university|institute of technology|department of|office of|\bHSE\b|health service|an garda|údarás|udaras|education and training board|\bETB\b|local authority|national \w+ authority|\bOPW\b|hospital|transport infrastructure ireland|\bTII\b|uisce éireann|irish water|tailte éireann)\b",
    re.I,
)
# Drops title/threshold rows that masquerade as a supplier (the page heading bleeds into the
# supplier column with the literal €20,000 threshold as its amount). Plural "Purchase Orders"
# and "Payments greater than/over €20,000" headings leaked through the singular-only pattern.
CATEGORY_WORD = re.compile(
    r"^\s*(total|category total|sum|subtotal|grand total|all suppliers|various|publication of|purchase orders?|payments? (greater|over|to suppliers)|payments? greater than)\b",
    re.I,
)
# Non-anchored variant: a page-title that bleeds into the supplier column may LEAD with the
# body name (e.g. "TU Dublin Payments and Purchase Orders over €20,000") so ^-anchoring misses
# it. These rows carry the literal threshold as their amount; the phrasing is never a real
# supplier name. Checked in addition to CATEGORY_WORD.
TITLE_ROW = re.compile(
    r"(purchase orders?|payments?)\b.{0,30}(over|greater than)\s*€?\s*20[,.]?000"
    r"|payments?\s+(and|or)\s+purchase orders?"
    # gov.ie department banners phrase the threshold as "Payments for/of €20,000 or above/over"
    # (DFAT, Health, Education) — the literal €20,000 in the banner otherwise leaks as a fake
    # supplier row with amount 20,000. "or above/over/more" only ever appears in a heading.
    r"|payments?\s+(for|of)\s+€?\s*20[,.]?000"
    r"|€?\s*20[,.]?000\s+or\s+(above|over|more)",
    re.I,
)
# exclude policy/guidance/privacy/contract docs when harvesting period data files
POLICY_RE = re.compile(
    r"guide|guidelin|\bplan\b|policy|circular|strategy|manual|terms|fin.?07|privacy|"
    r"prompt.?payment|appendix|procedure|annual.?report|statement|setup|form|charter|scheme",
    re.I,
)
DATA_FILE_RE = re.compile(r"q[1-4]\b|qtr|quarter|20[12]\d|h[12]\b|over.?20|over.?25|payment|purchase|\bpo[s]?\b", re.I)
NAV_HINT = re.compile(
    r"purchase|procure|over.?20|over.?25|20k|payment|quarter|qtr|finance|"
    r"publication|spend|supplier|expenditure|disclosure|financial",
    re.I,
)
MERGE_GAP = 22.0


# ============================================================================ CONFIG
# amount_semantics controlled vocab (PROCUREMENT_INVESTIGATION.md value taxonomy):
#   po_committed     -> "ordered €X"  (PO-over-20k order lists)  summable
#   payment_actual   -> "paid €X"     (payments/paid lists)      summable (true spend)
#   contract_award_value -> "awarded €X" (Tailte contracts)      caution
# listing_url = page to harvest period files from; direct_files = known-good file URLs
# (used as a floor so a publisher still yields data if its listing is JS/awkward).
def cfg(
    pid,
    name,
    ptype,
    sector,
    *,
    listing,
    semantics,
    grain,
    privacy="low",
    tier="A",
    direct=None,
    include=None,
    exclude=None,
    caveat="",
) -> dict:
    return {
        "id": pid,
        "name": name,
        "ptype": ptype,
        "sector": sector,
        "listing_url": listing,
        "amount_semantics": semantics,
        "grain": grain,
        "privacy_risk": privacy,
        "tier": tier,
        "direct_files": direct or [],
        "include": re.compile(include, re.I) if include else None,
        "exclude": re.compile(exclude, re.I) if exclude else None,
        "caveat": caveat,
    }


PUBLISHERS: list[dict] = [
    # ---- Tier A: clean tabular / high-confidence PDF -------------------------------
    cfg(
        "ie_opw",
        "Office of Public Works",
        "state_body",
        "property_land",
        listing="https://www.gov.ie/en/office-of-public-works/collections/payments-greater-than-20000/",
        semantics="payment_actual",
        grain="payment",
        direct=["https://assets.gov.ie/static/documents/b526ff76/OPW_Payments_of_20000_or_over_in_Q1_2026.xlsx"],
    ),
    cfg(
        "dept_climate",
        "Dept of Climate, Energy and the Environment",
        "department",
        "central_government",
        listing="https://www.gov.ie/en/department-of-climate-energy-and-the-environment/collections/payments-over-20000/",
        semantics="payment_actual",
        grain="payment",
        direct=["https://assets.gov.ie/static/documents/ae8b1a0a/DPER_Payments_over_20K_Q1_2026_Report.xlsx"],
    ),
    cfg(
        "dept_defence",
        "Department of Defence",
        "department",
        "central_government",
        listing="https://www.gov.ie/en/department-of-defence/collections/purchase-orders-over-20000/",
        semantics="po_committed",
        grain="purchase_order",
    ),
    cfg(
        "dept_culture",
        "Department of Culture, Communications and Sport",
        "department",
        "central_government",
        listing="https://www.gov.ie/en/department-of-culture-communications-and-sport/collections/purchase-orders/",
        semantics="po_committed",
        grain="purchase_order",
        caveat="contains very large NBI infrastructure POs; check outlier share before any total",
    ),
    # ---- Cheap wins 2026-06-08: gov.ie / enterprise.gov.ie departments already published, files
    # cached in c:/tmp but never wired. Both parse clean with the generic reader (offline-validated).
    cfg(
        "dept_dper",
        "Dept of Public Expenditure, Infrastructure, PSR and Digitalisation",
        "department",
        "central_government",
        listing="https://www.gov.ie/en/department-of-public-expenditure-infrastructure-public-service-reform-and-digitalisation/collections/dpendr-ogcio-and-ogp-purchase-order-payments-2024/",
        semantics="po_committed",
        grain="purchase_order",
        caveat="DPENDR+OGCIO+OGP PO-over-20000 listing; note the SEPARATE pre-existing bug that the "
        "Dept Climate gov.ie collection mis-serves a 'DPER_Payments_over_20K' file (filename/dept "
        "mismatch) — that is payment-grain and ingested under dept_climate, distinct from these POs",
    ),
    cfg(
        "dept_enterprise",
        "Department of Enterprise, Tourism and Employment",
        "department",
        "central_government",
        listing="https://enterprise.gov.ie/en/publications/payments-over-20k.html",
        semantics="payment_actual",
        grain="payment",
        include=r"\.xlsx(\?|$)|\.csv(\?|$)",
        caveat="DETE 'Payments over €20,000'; XLSX only (2024-2026 + 2017). Amount col is 'Total' (the "
        "'Payment Number' ref matches the amount regex via 'payment' but is excluded by NON_AMOUNT_HDR). "
        "The older quarterly PDFs (2016-2025) are DEFERRED: merged 'Supplier Name Total (€)' header + "
        "inconsistent line-wrapping corrupt the supplier column under the generic word-geometry reader "
        "(amount parses fine, supplier becomes 'DELL (IRELAND) 84,255') — would need a bespoke reading-order parser",
    ),
    cfg(
        "ie_teagasc",
        "Teagasc",
        "semi_state",
        "agri_food_marine",
        listing="https://www.teagasc.ie/about/corporate-responsibility/information-for-suppliers/",
        semantics="po_committed",
        grain="purchase_order",
    ),
    cfg(
        "ie_bordbia",
        "Bord Bia",
        "semi_state",
        "agri_food_marine",
        listing="https://www.bordbia.ie/about/governance/corporate-governance/purchase-orders/",
        semantics="po_committed",
        grain="purchase_order",
    ),
    cfg(
        "ie_bim",
        "Bord Iascaigh Mhara (BIM)",
        "semi_state",
        "agri_food_marine",
        listing="https://bim.ie/about/corporate-governance/purchase-orders-over-20k/",
        semantics="po_committed",
        grain="purchase_order",
        caveat="amounts excluding VAT",
    ),
    cfg(
        "ie_cib",
        "Citizens Information Board",
        "agency",
        "social",
        listing="https://www.citizensinformationboard.ie/en/freedom_of_information/financial_information/payments_or_purchase_orders_for_goods_and_services.html",
        semantics="payment_actual",
        grain="payment",
    ),
    cfg(
        "ie_hea",
        "Higher Education Authority",
        "agency",
        "education",
        listing="https://hea.ie/about-us/public-sector-information/",
        semantics="payment_actual",
        grain="payment",
        privacy="low",
    ),
    # ---- Tier F: government departments (gov.ie collections) — discovery sweep 2026-06-13.
    # All seven publish quarterly PO/payment-over-€20k lists as digital PDFs linked DIRECTLY on
    # the collection page (WebFetch-confirmed) — the proven Defence/Culture pattern the generic
    # header-anchored PDF reader already handles. grain per the page title (Purchase Orders ->
    # po_committed, Payments -> payment_actual). --list-verify before a full --merge run.
    cfg(
        "dept_agriculture",
        "Department of Agriculture, Food and the Marine",
        "department",
        "central_government",
        listing="https://www.gov.ie/en/collection/903f95-purchase-orders/",
        semantics="po_committed",
        grain="purchase_order",
        tier="F",
    ),
    cfg(
        "dept_social_protection",
        "Department of Social Protection",
        "department",
        "central_government",
        listing="https://www.gov.ie/en/department-of-social-protection/collections/purchase-orders-for-20000-or-above/",
        semantics="po_committed",
        grain="purchase_order",
        tier="F",
        caveat="PO-over-20000 quarterly, 2012-present",
    ),
    # dept_foreign_affairs (DFAT) DE-SCOPED 2026-06-13 — every "Payments over €20,000" PDF is a
    # single-column READING-ORDER layout ("<GL category> <SUPPLIER> <amount>" on one line), not a
    # column-geometry table, so the generic word-row reader cannot split supplier from category
    # (supplier comes out as "POSTAGE & OTHER COURIER COSTS AN POST"). Amounts ARE recoverable.
    # This is the NTA/SEAI/NPHDB bespoke family — needs a reading-order parser anchored on the
    # trailing amount. Listing (files-directly-linked, 2012-present):
    # https://www.gov.ie/en/department-of-foreign-affairs/organisation-information/payments-over-20000/
    # dept_justice DE-SCOPED 2026-06-13 — the "Purchase Orders Issued over €20,000" PDF is also a
    # single-column reading-order layout ("<PO#> <SUPPLIER> €<amount> <desc> <Y/N>"); the generic
    # reader scores a NOTES paragraph as the header and reads the 6-digit PO number as the amount
    # (€30bn+ garbage). The companion xlsx buries its real header under a 7-row notes preamble that
    # _tabular_from_raw's 8-row window misses. Both need bespoke handling (line-regex on the pdf, or
    # a deeper xlsx header search). Listing (annual, may lag a year):
    # https://www.gov.ie/en/department-of-justice-home-affairs-and-migration/collections/department-of-justice-purchase-orders-issued-over-20000-in-value/
    cfg(
        "dept_health",
        "Department of Health",
        "department",
        "central_government",
        listing="https://www.gov.ie/en/department-of-health/collections/department-of-health-payments-over-20000/",
        semantics="payment_actual",
        grain="payment",
        tier="F",
    ),
    cfg(
        "dept_education",
        "Department of Education and Youth",
        "department",
        "central_government",
        listing="https://www.gov.ie/en/department-of-education/collections/department-of-education-payments-greater-than-20000/",
        semantics="payment_actual",
        grain="payment",
        tier="F",
        caveat="Payments-of-20000-or-over quarterly PDFs, 2013-present (one 2023 quarter also offers xlsx)",
    ),
    # dept_transport DE-SCOPED 2026-06-13 — its PO-over-20000 PDFs (esp. the 2025
    # "Q#_Purchase_Order_20k_or_over.pdf" series) are reading-order, not column-geometry: 73% of
    # rows came out with a null supplier and the total inflated to €15.7bn (PO/value-column bleed).
    # Same bespoke reading-order family as DFAT/Justice. Listing (files directly linked, 2018-):
    # https://www.gov.ie/en/department-of-transport/organisation-information/departmental-purchase-orders-greater-than-20000/
    # ---- Tier B: OWNED BY A SEPARATE CONTEXT (procurement_hse_tusla_parser.py) -----
    # HSE + Tusla need bespoke per-publisher column-x specs (the generic header-anchored
    # reader misparses them: HSE fuses amount+quarter+date, Tusla's vendor bleeds into the
    # amount column). NTPF + SVUH (health, privacy=high) de-scoped here too pending that
    # context's reconciliation. Their output merges into THIS schema later — do not re-add
    # HSE/Tusla here or the generic reader will produce duplicate low-quality rows.
    # ---- Tier C: needed a corrected listing URL or a parser fix --------------------
    cfg(
        "ie_tii",
        "Transport Infrastructure Ireland",
        "agency",
        "transport",
        listing="https://www.tii.ie/en/compliance/payments/",
        semantics="payment_actual",
        grain="payment",
        tier="C",
        direct=["https://websitecms.tii.ie/media/sw3dzt2l/tii-payments-q1-2025-over-20k.csv"],
        caveat="CSV carries a category-total row (~€1.2bn) that must be excluded from any sum",
    ),
    cfg(
        "ie_revenue",
        "Revenue Commissioners",
        "agency",
        "regulator",
        listing="https://www.revenue.ie/en/corporate/statutory-obligations/freedom-of-information/section8/procurement.aspx",
        semantics="payment_actual",
        grain="payment",
        tier="C",
        direct=["https://www.revenue.ie/en/corporate/documents/procurement/payments-over-20000-quarter4-2025.pdf"],
    ),
    cfg(
        "ie_atu",
        "Atlantic Technological University",
        "education_body",
        "education",
        listing="https://www.atu.ie/freedom-of-information/freedom-of-information-financial-information",
        semantics="payment_actual",
        grain="payment",
        privacy="medium",
        tier="C",
        direct=["https://www.atu.ie/app/uploads/2026/03/atu-payments-purchase-orders-q1-2025.pdf"],
        caveat="supplier published with a leading numeric supplier-ID; stripped on read",
    ),
    # ie_nta DE-SCOPED to pipeline_sandbox/procurement_nta_parser.py — every NTA PO PDF is
    # 90deg-rotated (and the layout/date format varies by year), so the generic word-geometry
    # reader clusters a whole €-column into one row and yields 0. The bespoke reading-order
    # parser owns it (9 quarters, ~2.3k rows) and emits THIS schema. Do not re-add here.
    cfg(
        "ie_marine",
        "Marine Institute",
        "agency",
        "agri_food_marine",
        listing="https://www.marine.ie/site-area/about-us/purchase-orders",
        semantics="po_committed",
        grain="purchase_order",
        tier="C",
        direct=["https://marine.ie/sites/default/files/MIFiles/Docs/CS/Purchase%20Orders%20Qtr%201%202026.pdf"],
    ),
    cfg(
        "ie_esbnetworks",
        "ESB Networks DAC",
        "semi_state",
        "energy_utilities",
        listing="https://www.esbnetworks.ie/about-us/company/publication-scheme/financial-information",
        semantics="payment_actual",
        grain="payment",
        tier="C",
        caveat="prior sample was a category-total page; harvesting supplier-level file",
    ),
    cfg(
        "ie_tailte",
        "Tailte Éireann",
        "state_body",
        "property_land",
        listing="https://tailte.ie/category/publications/",
        semantics="po_committed",
        grain="purchase_order",
        tier="C",
        include=r"purchase|payment|po[s]?[-_ ]?over|20[,]?000|over.?20k",
        caveat="Purchase-Orders quarterly files (PO grain); contracts-awarded list excluded",
    ),
    cfg(
        "dept_housing",
        "Department of Housing, Local Government and Heritage",
        "department",
        "central_government",
        listing="https://www.gov.ie/en/department-of-housing-local-government-and-heritage/collections/procurement-related-payments-over-20000-euro/",
        semantics="payment_actual",
        grain="payment",
        tier="C",
        caveat="prior sample was a privacy statement; using the gov.ie payments collection (slug renamed to procurement-related-payments-over-20000-euro 2026-06)",
    ),
    cfg(
        "ie_cdetb",
        "City of Dublin ETB",
        "education_body",
        "education",
        listing="https://www.cityofdublinetb.ie/about-us/finance-and-procurement/procurement/",
        semantics="po_committed",
        grain="purchase_order",
        privacy="medium",
        tier="C",
        include=r"purchase|payment|po[s]?[-_ ]?over|20[,]?000|quarter|q[1-4]",
        caveat="prior sample was the procurement policy; excluding policy docs",
    ),
    cfg(
        "ie_enterprise_ireland",
        "Enterprise Ireland",
        "semi_state",
        "enterprise_tourism",
        listing="https://www.enterprise-ireland.com/en/legal/policies-guidelines/procurement-policy",
        semantics="po_committed",
        grain="purchase_order",
        tier="C",
        include=r"purchase|payment|po[s]?[-_ ]?over|20[,]?000|over.?20k",
        caveat="agency (not DETE dept) procurement-policy page; quarterly XLSX 'Payments over €20,000' 2012-present",
    ),
    # ---- Tier D: discovery sweep 2026-06-04 (doc/PROCUREMENT_SOURCE_DISCOVERY_2026_06_04.md) --
    # Probe-confirmed, generic-reader-clean. Held back for bespoke/render passes (NOT here):
    #   Beaumont + Pobal (dual/MIXED PO+payment grain — need value_kind split),
    #   Coimisiún na Meán + Irish Prison Service (scanned PDFs — need OCR),
    #   Garda (sampler hit a fleet report — needs the right PO subpage),
    #   UCD / SETU / CHI / SEAI / EPA (no links via landing — JS/403, EPA serves .php HTML).
    cfg(
        "ie_ntma",
        "National Treasury Management Agency (NTMA)",
        "state_body",
        "finance",
        listing="https://www.ntma.ie/information-pages/freedom-of-information/freedom-of-information-publication-scheme/financial-information",
        semantics="payment_actual",
        grain="payment",
        tier="D",
        exclude=r"revised-foi-publication|[-_ ]publication\.pdf",
        caveat="one quarterly scheme covers 6 business units incl NDFA (ADM/Nat-Debt/ISIF/NDFA/FIF/ICNF); "
        "do NOT also wire ie_ndfa or its rows double-count. The 6-row 'Revised-FOI-Publication' / "
        "'*-Publication.pdf' files are per-unit SUMMARIES (different grain) that overlap the "
        "line-level Q*-Payments files in 2018-19 — excluded to avoid double-counting. "
        "NOTE: the per-unit Q1-2020..Q2-2024 PDFs currently parse to 0 rows (layout/scan break) — known gap.",
    ),
    cfg(
        "ie_courts",
        "Courts Service of Ireland",
        "agency",
        "justice",
        listing="https://www.courts.ie/publications/purchase-orders-greater-than-20k",
        semantics="po_committed",
        grain="purchase_order",
        tier="D",
        include=r"purchase-order|over-20|po[s]?[-_ ]?over",
    ),
    cfg(
        "ie_sportireland",
        "Sport Ireland",
        "agency",
        "sport",
        listing="https://www.sportireland.ie/about-us/freedom-of-information/financial-information",
        semantics="po_committed",
        grain="purchase_order",
        tier="D",
        caveat="single rolling PO log (not per-quarter); period likely null",
    ),
    cfg(
        "ie_tudublin",
        "Technological University Dublin",
        "education_body",
        "education",
        listing="https://www.tudublin.ie/explore/governance-and-compliance/foi/foi-publication-scheme/",
        semantics="po_committed",
        grain="purchase_order",
        tier="D",
        include=r"po-report|purchase-order|over-?20k",
    ),
    cfg(
        "ie_mtu",
        "Munster Technological University (MTU)",
        "education_body",
        "education",
        listing="https://www.mtu.ie/about-mtu/legal/freedom-of-information/",
        semantics="po_committed",
        grain="purchase_order",
        tier="D",
        include=r"pos?-over-?20k|purchase-order|po[s]?[-_ ]?over",
        # Landing only exposes the tender-register xlsx + FOI logs; the actual PO PDFs live under
        # /media/.../foi/financial-information/ and aren't reachable by the one-hop crawl, so the
        # quarterly files are pinned directly. All 3 byte-verified 2026-06-04 (%PDF, 88-132KB);
        # Q4-2025 parses to 123 rows high-conf. Add more quarters as their URLs are confirmed.
        direct=[
            "https://www.mtu.ie/media/mtu-website/files/foi/financial-information/MTU-POs-over-20k-Q4-2025.pdf",
            "https://www.mtu.ie/media/mtu-website/files/foi/financial-information/MTU-POs-over-20k-Q3-2025.pdf",
            "https://www.mtu.ie/media/mtu-website/files/foi/financial-information/MTU-POs-over-20k-Q2-2025.pdf",
        ],
        caveat="PO PDFs pinned via direct_files (landing exposes only tender-register xlsx + FOI logs)",
    ),
    cfg(
        "ie_chi",
        "Children's Health Ireland (CHI)",
        "state_body",
        "health",
        listing="https://www.childrenshealthireland.ie/about-us/corporate-information/payments-to-suppliers-over-20000/",
        semantics="payment_actual",
        grain="payment",
        privacy="low",
        tier="D",
        # Children's-hospital OPERATOR side (complements NPHDB construction). Landing exposes no
        # direct links → file pinned. xlsx row 0 is a TITLE ("CHI Vendor payments >25K") above the
        # real "Vendor Name/Amount" header; the length-filtered header scorer now skips it (297 rows).
        direct=[
            "https://www.childrenshealthireland.ie/documents/3541/CHI_Paid_Invoices_over_25K_incl_VAT_Qtr_1_2026updated.xlsx"
        ],
        caveat="paid invoices at €25k incl VAT (not €20k); single Q1-2026 file; payment grain",
    ),
    cfg(
        "ie_pobal",
        "Pobal",
        "agency",
        "social",
        listing="https://www.pobal.ie/financial-information/",
        semantics="po_committed",
        grain="purchase_order",
        privacy="medium",
        tier="D",
        # Files titled 'Purchase Order OR Payments over €20k' but rows carry PO/SUPPLIER/TOTAL/PAID
        # columns = POs with a paid-flag (Paid/Not Paid captured in paid_flag), not truly mixed.
        # Generic reader handles it (29 rows/high-conf on Q1-2026). Full 2020-2026 series (25 PDFs).
        caveat="grant-adjacent (privacy=medium); harvest returns oldest-first so a low --max-files "
        "biases to 2020 — raise --max-files for full series",
    ),
    cfg(
        "ie_beaumont",
        "Beaumont Hospital",
        "hospital",
        "health",
        listing="https://www.beaumont.ie/page/financial-statements",
        semantics="payment_actual",
        grain="payment",
        privacy="low",
        tier="D",
        # Landing exposes 3 xlsx: two 'Payments Over €20k' (annual, payment grain, 2024+2025) and
        # one 'POs Greater than €20k' (PO grain). include= grabs ONLY the payment files to keep one
        # grain. Header 'No. of Payments > €20,000' is a COUNT trap; NON_AMOUNT_HDR routes amount to 'Value'.
        include=r"payments.*over",
        caveat="annual supplier payment totals (Value col), €20k threshold; the separate Q1-2026 PO file is excluded",
    ),
    # ---- Tier E: regulators / cultural bodies (discovery sweep 2 — commercial-vs-noncommercial) --
    # Commercial semi-states (ESB/Electric Ireland, daa, An Post, ports, CIÉ group) are FOI-exempt
    # and publish annual reports only — NOT here. EirGrid/GNI/Uisce publish CATEGORY-only rollups
    # (no supplier names) despite "PO over €20k" page titles — NOT here. RTÉ's published file is a
    # category summary (Capital/Communication circuits + counts), NOT supplier-level — NOT here.
    # ABP is supplier-level but its multi-line bilingual (Irish) header bleeds date into supplier —
    # deferred (needs header-wrap handling). These three parse clean with the generic reader:
    cfg(
        "ie_hpra",
        "Health Products Regulatory Authority (HPRA)",
        "agency",
        "regulator",
        listing="https://www.hpra.ie/transparency/financial-information/purchase-orders",
        semantics="po_committed",
        grain="purchase_order",
        tier="E",
        direct=[
            "https://assets.hpra.ie/data/docs/default-source/corporate/purchase-orders/purchase-orders---q3-2025.pdf"
        ],
        caveat="clean assets.hpra.ie CDN, predictable quarterly PDF filenames",
    ),
    cfg(
        "ie_ccpc",
        "Competition and Consumer Protection Commission (CCPC)",
        "agency",
        "regulator",
        listing="https://www.ccpc.ie/about-us/corporate-information/governance/payment-reports",
        semantics="payment_actual",
        grain="payment",
        tier="E",
        direct=[
            "https://assets.ccpc.ie/data/docs/default-source/about-us/corporate-information/governance/payment-reports/payments-over-20k-in-q1-2026.pdf"
        ],
        caveat="quarterly payments >€20k; description column repeats the € amount as a prefix (cosmetic)",
    ),
    cfg(
        "ie_nli",
        "National Library of Ireland",
        "agency",
        "media_culture",
        listing="https://www.nli.ie/corporate-information",
        semantics="payment_actual",
        grain="payment",
        tier="E",
        direct=["https://www.nli.ie/sites/default/files/2025-05/payments-over-eu20000-q1-2025.pdf"],
        caveat="Drupal /sites/default/files PDFs; some files bundle Q1-Q4 annually",
    ),
]


# ============================================================================ fetch
def _curl(url: str) -> bytes | None:
    try:
        p = subprocess.run(
            ["curl", "-sS", "-k", "-L", "--max-time", "90", "-A", H["User-Agent"], url],
            capture_output=True,
            timeout=120,
        )
        return p.stdout if p.returncode == 0 and p.stdout else None
    except Exception:
        return None


def fetch_bytes(url: str) -> bytes | None:
    # some publishers emit hrefs with raw spaces — requests/curl reject them as malformed;
    # '%' stays in the safe set so already-encoded hrefs don't double-encode.
    url = quote(url, safe="!#$%&'()*+,/:;=?@[]~")
    LAST_ERR.clear()
    try:
        r = requests.get(url, headers=H, timeout=90, allow_redirects=True)
        r.raise_for_status()
        return r.content
    except Exception as e:
        ec, status = classify_exception(e)
        LAST_ERR.update({"error_class": ec, "http_status": status})
        b = _curl(url)
        if b:
            LAST_ERR.clear()
        return b


def fetch_text(url: str) -> str | None:
    b = fetch_bytes(url)
    return b.decode("utf-8", "ignore") if b else None


def fetch_to_bronze(pub_id: str, url: str, ext: str, refetch: bool = False) -> tuple[bytes | None, bool]:
    """Self-fetch a source file to bronze/pdfs/public_body_procurement/<id>/ and reuse the
    cached copy on re-runs — quarterly disclosures are immutable, so steady-state runs only
    download newly published files (same shape as the LA extractor). Returns
    ``(bytes, fresh_download)``. The DNN ``?sfvrsn=`` version param drops out of the cache
    key deliberately: same filename = same historical document."""
    dest = (
        BRONZE
        / pub_id
        / (re.sub(r"[^A-Za-z0-9._-]", "_", unquote(url.split("?")[0].rsplit("/", 1)[-1]))[:80] or "file")
    )
    if not dest.suffix:
        dest = dest.with_suffix(ext if ext in DATA_EXT else ".pdf")
    if not refetch and dest.exists() and dest.stat().st_size > 1500:
        return dest.read_bytes(), False
    time.sleep(1.0)  # politeness: only on a real network fetch, never on a cache hit
    b = fetch_bytes(url)
    if b and len(b) > 1500:
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(b)
    return b, True


# ============================================================================ harvest
def harvest_files(cf: dict, crawl_cap: int = 12) -> list[str]:
    """Collect ALL period data-file links for a publisher: landing page + one-hop crawl,
    minus policy/guidance docs. Union with direct_files. Honours an optional include re."""
    found: list[str] = list(cf["direct_files"])
    html = fetch_text(cf["listing_url"])
    if html:

        def scan(page_html: str, base: str) -> list[str]:
            out = []
            for href in HREF_RE.findall(page_html):
                low = href.lower().split("?")[0]
                if not any(low.endswith(e) for e in DATA_EXT):
                    continue
                if POLICY_RE.search(href):
                    continue
                if not DATA_FILE_RE.search(href):
                    continue
                if cf["include"] and not cf["include"].search(href):
                    continue
                if cf["exclude"] and cf["exclude"].search(href):
                    continue
                out.append(urljoin(base, href))
            return out

        hits = scan(html, cf["listing_url"])
        if not hits:  # one-hop crawl same-host nav links
            host = urlparse(cf["listing_url"]).netloc
            subs, seen = [], set()
            for href in HREF_RE.findall(html):
                full = urljoin(cf["listing_url"], href)
                low = full.lower().split("?")[0]
                if urlparse(full).netloc != host or full == cf["listing_url"]:
                    continue
                if any(low.endswith(e) for e in DATA_EXT):
                    continue
                if NAV_HINT.search(href) and full not in seen:
                    seen.add(full)
                    subs.append(full)
            for s in subs[:crawl_cap]:
                sub_html = fetch_text(s)
                if sub_html:
                    hits.extend(scan(sub_html, s))
        found.extend(hits)
    # Dedup by basename STEM (extension stripped), not basename: the same quarterly
    # report is sometimes published in two formats (e.g. dept_climate Q1-2026 as both
    # .xlsx AND .pdf) which a with-extension key let through -> double-counted rows.
    # On a stem collision prefer the cleaner tabular format (xlsx/csv > xls > pdf).
    # (Also still collapses the same file served via two hosts, e.g. TII.)
    fmt_pref = {".xlsx": 0, ".csv": 1, ".xls": 2, ".pdf": 3}

    def stem_ext(u: str) -> tuple[str, str]:
        base = u.rsplit("/", 1)[-1].split("?")[0].lower()
        for e in DATA_EXT:
            if base.endswith(e):
                return base[: -len(e)], e
        return base, ""

    best: dict[str, str] = {}
    order: list[str] = []
    for u in found:
        s, e = stem_ext(u)
        if s not in best:
            best[s] = u
            order.append(s)
        elif fmt_pref.get(e, 9) < fmt_pref.get(stem_ext(best[s])[1], 9):
            best[s] = u  # keep the more reliable format for the same report
    result = [best[s] for s in order]

    # Cross-format same-period dedup: some publishers post the SAME quarter as BOTH a tabular file
    # (csv/xlsx) AND a pdf under DIFFERENT filenames (e.g. dept_agriculture Q4-2025 as
    # "Q4_2025_Purchase_Orders_over_20k.pdf" + "Payments_to_Suppliers…_Q4_2025.csv"). Different
    # stems slip past the stem-dedup above and the quarter is counted twice. When a period carries
    # BOTH a tabular and a pdf, drop the pdf (the tabular has cleaner columns). Crucially this only
    # fires across DIFFERENT formats — same-format repeats in one period are untouched, so NTMA's
    # 6 per-business-unit pdfs/quarter and any genuinely split quarter survive.
    def _is_tab(u: str) -> bool:
        return stem_ext(u)[1] in (".csv", ".xlsx", ".xls")

    period_fmts: dict[str | None, set[bool]] = {}
    for u in result:
        per = period_from_url(u)[0]
        if per:
            period_fmts.setdefault(per, set()).add(_is_tab(u))
    drop_pdf_periods = {p for p, fmts in period_fmts.items() if True in fmts and False in fmts}
    return [u for u in result if _is_tab(u) or period_from_url(u)[0] not in drop_pdf_periods]


# ============================================================================ readers
def to_eur(token) -> float | None:
    if token is None:
        return None
    if isinstance(token, (int, float)):
        return float(token)
    s = str(token).strip()
    neg = s.startswith("(") and s.endswith(")")
    m = NUM_RE.search(s)
    if not m:
        return None
    with contextlib.suppress(ValueError):
        v = float(m.group().replace(",", ""))
        return -v if neg else v
    return None


def clean_supplier(s) -> str:
    return DIGIT_PREFIX.sub("", str(s or "")).strip(" -:|,")


# Month-range filenames (e.g. "jul-sep-2016.pdf") encode a quarter the Q\d patterns miss; many
# older Dept of Defence PO lists are named this way, so their period otherwise collapses to the
# bare year and recurring quarterly payments look like cross-file duplicates. DQ audit 2026-06-06.
MONTH_RANGE_Q = [
    (re.compile(r"jan\w*[-_ .]+mar", re.I), 1),
    (re.compile(r"apr\w*[-_ .]+jun", re.I), 2),
    (re.compile(r"jul\w*[-_ .]+sep", re.I), 3),
    (re.compile(r"oct\w*[-_ .]+dec", re.I), 4),
]


def quarter_from_name(name: str) -> int | None:
    q = QUARTER_RE.search(name)
    if q:
        return next((int(g) for g in q.groups() if g), None)
    for rx, qn in MONTH_RANGE_Q:
        if rx.search(name):
            return qn
    return None


def period_from_url(url: str) -> tuple[str | None, int | None, int | None]:
    name = url.rsplit("/", 1)[-1]
    y = PERIOD_RE.search(name)
    year = int(y.group(1)) if y else None
    quarter = quarter_from_name(name)
    period = f"{year}-Q{quarter}" if year and quarter else (str(year) if year else None)
    return period, year, quarter


# ---- PDF (header-anchored) ----
def cluster_word_rows(page, ytol: float = 3.0) -> list[list]:
    words = page.get_text("words")
    words.sort(key=lambda w: (round(w[1] / ytol), w[0]))
    rows, cur, cur_y = [], [], None
    for w in words:
        y = w[1]
        if cur_y is None or abs(y - cur_y) <= ytol:
            cur.append(w)
            cur_y = y if cur_y is None else cur_y
        else:
            rows.append(cur)
            cur, cur_y = [w], y
    if cur:
        rows.append(cur)
    return rows


def find_header(rows: list[list]):
    best, best_hits = None, 1
    for r in rows[:18]:
        text = " ".join(w[4] for w in r)
        hits = sum(bool(rx.search(text)) for rx in ROLE_RE.values())
        has_anchor = ROLE_RE["supplier"].search(text) or ROLE_RE["amount"].search(text)
        if hits >= 2 and has_anchor and hits > best_hits:
            best, best_hits = r, hits
    return best


def header_columns(header: list) -> list[dict]:
    ws = sorted(header, key=lambda w: w[0])
    cols: list[dict] = []
    for w in ws:
        if cols and w[0] - cols[-1]["x1"] < MERGE_GAP:
            cols[-1]["label"] += " " + w[4]
            cols[-1]["x1"] = max(cols[-1]["x1"], w[2])
        else:
            cols.append({"label": w[4], "x0": w[0], "x1": w[2]})
    for c in cols:
        c["center"] = (c["x0"] + c["x1"]) / 2
    return cols


def assign_role(cols: list[dict]) -> dict[str, int]:
    roles: dict[str, int] = {}
    for role, rx in ROLE_RE.items():
        cands = [i for i, c in enumerate(cols) if rx.search(c["label"])]
        if role == "amount":
            # Drop identifier/count columns that carry a money keyword (DETE PDF "Payment No."
            # holds PO refs ~60,002,179 that dwarf the real "Total" amount). Same guard as the
            # tabular reader; see NON_AMOUNT_HDR.
            cands = [i for i in cands if not NON_AMOUNT_HDR.search(cols[i]["label"])] or cands
        if cands:
            roles[role] = cands[-1] if role == "amount" else cands[0]
    return roles


def row_to_cols(words: list, cols: list[dict]) -> list[str]:
    bounds = [(cols[i]["center"] + cols[i + 1]["center"]) / 2 for i in range(len(cols) - 1)]
    buckets: list[list] = [[] for _ in cols]
    for w in sorted(words, key=lambda w: w[0]):
        c = (w[0] + w[2]) / 2
        idx = 0
        while idx < len(bounds) and c > bounds[idx]:
            idx += 1
        buckets[idx].append(w[4])
    return [" ".join(b).strip(" -:|") for b in buckets]


def refine_roles(cols, roles, records):
    if not records:
        return roles

    def numfrac(i):
        vals = [r[i] for r in records if i < len(r) and r[i]]
        return sum(to_eur(v) is not None for v in vals) / len(vals) if vals else 0.0

    def moneyfrac(i):  # fraction of cells that look like MONEY (thousands/decimals) — a bare
        vals = [r[i] for r in records if i < len(r) and r[i]]  # year col like "2023" won't match
        return sum(bool(MONEY_RE.search(str(v))) for v in vals) / len(vals) if vals else 0.0

    amt_cands = [i for i, c in enumerate(cols) if ROLE_RE["amount"].search(c["label"])]
    # Exclude identifier/count columns before ranking by numeric density — a "Payment No." /
    # "Order No." column is 100% numeric and would otherwise win the numfrac tie-break.
    amt_cands = [i for i in amt_cands if not NON_AMOUNT_HDR.search(cols[i]["label"])] or amt_cands
    if amt_cands:
        roles["amount"] = max(amt_cands, key=numfrac)
    elif "amount" not in roles and cols:
        # No header word matches "amount" (e.g. NTMA's amount column is headed "Q4"/"Q3",
        # the quarter, not a money word). Fall back to the most money-like column by content.
        best = max(range(len(cols)), key=moneyfrac)
        if moneyfrac(best) >= 0.5:
            roles["amount"] = best
    sup_cands = [i for i, c in enumerate(cols) if ROLE_RE["supplier"].search(c["label"])]
    if sup_cands:
        roles["supplier"] = min(sup_cands, key=numfrac)
    return roles


def read_pdf(b: bytes, max_pages: int | None) -> dict:
    doc = fitz.open(stream=b, filetype="pdf")
    npages = doc.page_count
    limit = min(npages, max_pages) if max_pages else npages
    cols, header_label, page0 = [], "", ""
    for i in range(min(npages, 3)):
        rows = cluster_word_rows(doc[i])
        if i == 0:
            page0 = doc[i].get_text("text")
        h = find_header(rows)
        if h:
            cols = header_columns(h)
            header_label = " | ".join(c["label"] for c in cols)
            break
    roles = assign_role(cols) if cols else {}
    out_rows, digital_chars = [], 0
    for i in range(limit):
        page = doc[i]
        digital_chars += len(page.get_text("text").strip())
        if not cols:
            continue
        for wrow in cluster_word_rows(page):
            xs = [w[4] for w in wrow]
            if not any(MONEY_RE.search(t) for t in xs):
                continue
            out_rows.append((i + 1, row_to_cols(wrow, cols)))
    doc.close()
    roles = refine_roles(cols, roles, [r for _, r in out_rows]) if cols else roles
    return {
        "digital": digital_chars > 200,
        "cols": cols,
        "header_label": header_label,
        "roles": roles,
        "rows": out_rows,
        "page0": page0,
        "pages": npages,
    }


# ---- XLSX / XLS / CSV ----
def _tabular_from_raw(raw: list[list]):
    """Shared header-pick + body-trim for any 2D cell grid (openpyxl or xlrd)."""
    full = " ".join(str(c) for row in raw[:6] for c in row if c is not None)

    def score(row):
        # Count SHORT role-matching cells only: a real header is short labels ("Vendor Name",
        # "Amount"); a TITLE row above it is one long sentence with embedded keywords
        # ("CHI Vendor payments >25K (incl VAT)") that should NOT win header detection.
        return sum(
            1
            for c in (row or [])
            if c is not None and len(str(c).strip()) <= 30 and any(rx.search(str(c)) for rx in ROLE_RE.values())
        )

    # Tie-break toward the LATER row — a title/banner row precedes the real header.
    hi = max(range(min(8, len(raw))), key=lambda i: (score(raw[i]), i), default=0)
    header = [str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(raw[hi])]
    rows = [r for r in raw[hi + 1 :] if any(c is not None and str(c).strip() for c in r)]
    return header, rows, full


def read_xlsx(b: bytes):
    import openpyxl

    ws = openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True).active
    raw = [list(r) for r in ws.iter_rows(values_only=True)]
    return _tabular_from_raw(raw)


def read_xls(b: bytes):
    import xlrd  # legacy binary .xls (pre-2021 quarterlies); openpyxl is .xlsx-only

    sh = xlrd.open_workbook(file_contents=b).sheet_by_index(0)
    raw = [sh.row_values(i) for i in range(sh.nrows)]
    return _tabular_from_raw(raw)


def read_csv(b: bytes):
    # gov.ie CSVs frequently carry a TITLE + blank preamble before the real header row, e.g. Social
    # Protection "Purchase Orders over €20,000 - Quarter 1, 2026" / blank / "Supplier name,Order
    # type,Order No,Order amount", or Health "Payments €20,000 or above…" / blank / blank /
    # "Reference,Payee Name,Tran Value,Description,Paid Y/N". polars' default has_header=True takes
    # the title as the header and the real columns get lost. Read header-less and reuse the shared
    # title-skipping header detector (same as xlsx/xls).
    df = pl.read_csv(
        io.BytesIO(b),
        has_header=False,
        infer_schema_length=0,
        truncate_ragged_lines=True,
        ignore_errors=True,
        encoding="utf8-lossy",
    )
    raw = [list(r) for r in df.iter_rows()]
    return _tabular_from_raw(raw)


# A header that carries a money-ish KEYWORD but is really an identifier or a count, never the
# amount: a COUNT column (Beaumont "No. of Payments > €20,000", values 1..300) OR an ID/reference
# column (DETE "Payment Number" — matches the amount regex via 'payment', but holds PO refs like
# 137014941 that look like €137m). Excluded from amount detection so the real money column
# ("Total"/"Value"/"Amount") wins; a true amount column is never titled "...Number/Reference".
NON_AMOUNT_HDR = re.compile(
    r"no\.?\s*of\b|number\b|\bnumbers\b|\bcount\b|\bqty\b|quantit|reference\b|\bref\b|\bid\b|"
    # an "<X> No." identifier column (DETE PDF "Payment No.", "PO No.", "Order No.", "Invoice No.")
    r"(?:payment|order|invoice|po|p\.?o\.?|doc|transaction)\s*no\.?\b",
    re.I,
)


def detect_roles_tab(header, rows):
    roles = {k: None for k in ROLE_RE}
    for role, rx in ROLE_RE.items():
        cands = [i for i, h in enumerate(header) if rx.search(h or "")]
        if not cands:
            continue
        if role == "amount":
            # Drop count/identifier columns that carry money keywords in their label (see
            # NON_AMOUNT_HDR) before ranking by numeric density.
            strong = [i for i in cands if not NON_AMOUNT_HDR.search(header[i] or "")]
            cands = strong or cands
            cands.sort(
                key=lambda i: (
                    -(sum(to_eur(r[i]) is not None for r in rows[:200] if i < len(r)) / max(1, len(rows[:200])))
                )
            )
        roles[role] = cands[0]

    # A category/description column is often headed "Payment Type" / "Order Type" (OPW, several
    # gov.ie bodies): the `paid` regex claims it (its pattern includes "payment type") and the
    # `description` regex — which has no "type" keyword — leaves description empty, so the useful
    # category text (Software, Roofworks, Building Maintenance) is discarded. When the column the
    # `paid` role grabbed actually holds free-text categories rather than a Y/N flag or a payment
    # count, promote it to description (and clear `paid`, so "Software" never lands in paid_flag).
    def _looks_descriptive(i: int) -> bool:
        vals = [str(r[i]).strip() for r in rows[:200] if i < len(r) and r[i] not in (None, "")]
        if len(vals) < 3:
            return False
        if sum(to_eur(v) is not None for v in vals) / len(vals) > 0.3:  # a numeric flag/count, not text
            return False
        has_letters = sum(any(c.isalpha() for c in v) for v in vals) / len(vals)
        distinct = len({v.lower() for v in vals})
        # free-text categories vary (Software/Roofworks/…); a Paid/Not-Paid flag has ≤2 values.
        return has_letters > 0.7 and distinct > 2

    if roles["description"] is None and roles["paid"] is not None and _looks_descriptive(roles["paid"]):
        roles["description"], roles["paid"] = roles["paid"], None
    return roles


# ============================================================================ extract
def emit_rows(cf, file_url, b, fmt, max_pages) -> tuple[list[dict], dict]:
    """Parse one file -> gold-schema row dicts + a small per-file stat block."""
    fhash = hashlib.sha256(b).hexdigest()[:16]
    period, year, quarter = period_from_url(file_url)
    rows_out: list[dict] = []
    caveat_detected = False
    conf = "low"

    def base(srn, page, supplier, amount, desc, po, paid):
        return {
            "publisher_id": cf["id"],
            "publisher_name": cf["name"],
            "publisher_type": cf["ptype"],
            "sector": cf["sector"],
            "source_landing_url": cf["listing_url"],
            "source_file_url": file_url,
            "source_file_hash": fhash,
            "period": period,
            "year": year,
            "quarter": quarter,
            "supplier_raw": supplier,
            "amount_eur": amount,
            "amount_semantics": cf["amount_semantics"],
            "description": desc,
            "po_number": po,
            "paid_flag": paid,
            "source_row_number": srn,
            "source_page_number": page,
            "parser_name": f"public_body_{fmt}",
            "parser_version": PARSER_VERSION,
            "source_caveat": cf["caveat"] or None,
        }

    if fmt == "pdf":
        info = read_pdf(b, max_pages)
        caveat_detected = bool(CAVEAT_RE.search(info["page0"]) or CAVEAT_RE.search(info["header_label"]))
        if not info["digital"] or not info["cols"] or "amount" not in info["roles"]:
            return [], {
                "status": "unparsed",
                "reason": "scanned/no-header/no-amount",
                "rows": 0,
                "confidence": "low",
                "pages": info.get("pages"),
            }
        sup_i = info["roles"].get("supplier")
        amt_i = info["roles"]["amount"]
        desc_i, po_i, paid_i = (info["roles"].get(k) for k in ("description", "po", "paid"))
        good = 0
        for srn, (page, rec) in enumerate(info["rows"]):
            amt = to_eur(rec[amt_i]) if amt_i < len(rec) else None
            if amt is None:
                continue
            sup = clean_supplier(rec[sup_i]) if sup_i is not None and sup_i < len(rec) else None
            desc = rec[desc_i] if desc_i is not None and desc_i < len(rec) else None
            # Drop total/category/title-masquerade rows. The page banner ("... Payments greater
            # than €20,000") splits across cells — "greater than" into the description, the
            # "€20,000" into the amount column — so no single cell holds the whole phrase. Test
            # TITLE_ROW against the JOINED row (bucket order re-adjoins "greater than … 20,000").
            rowtext = " ".join(str(x) for x in rec if x)
            if (sup and CATEGORY_WORD.search(sup)) or TITLE_ROW.search(rowtext):
                continue
            good += 1
            rows_out.append(
                base(
                    srn,
                    page,
                    sup,
                    amt,
                    desc,
                    clean_supplier(rec[po_i]) if po_i is not None and po_i < len(rec) else None,
                    rec[paid_i] if paid_i is not None and paid_i < len(rec) else None,
                )
            )
        conf = "high" if good > 20 else ("medium" if good > 3 else "low")

    else:  # xlsx / xls / csv
        reader = {"xlsx": read_xlsx, "xls": read_xls, "csv": read_csv}[fmt]
        header, rows, full = reader(b)
        caveat_detected = bool(CAVEAT_RE.search(full) or any(CAVEAT_RE.search(h or "") for h in header))
        roles = detect_roles_tab(header, rows)
        sup_i, amt_i = roles["supplier"], roles["amount"]
        if amt_i is None:
            return [], {"status": "unparsed", "reason": "no-amount-col", "rows": 0, "confidence": "low"}
        desc_i, po_i, paid_i = roles["description"], roles["po"], roles["paid"]
        good = 0
        for srn, r in enumerate(rows):
            amt = to_eur(r[amt_i]) if amt_i < len(r) else None
            if amt is None:
                continue
            sup = clean_supplier(r[sup_i]) if sup_i is not None and sup_i < len(r) else None
            if sup and CATEGORY_WORD.search(sup):
                continue
            good += 1
            rows_out.append(
                base(
                    srn,
                    None,
                    sup,
                    amt,
                    r[desc_i] if desc_i is not None and desc_i < len(r) else None,
                    clean_supplier(r[po_i]) if po_i is not None and po_i < len(r) else None,
                    r[paid_i] if paid_i is not None and paid_i < len(r) else None,
                )
            )
        conf = "high" if good > 20 else ("medium" if good > 3 else "low")

    for r in rows_out:
        r["extraction_status"] = "extracted"
        r["extraction_confidence"] = conf
        r["caveat_text_detected"] = caveat_detected
        # Blank-supplier repair: a mis-mapped column can leave supplier_raw empty while the
        # company name sits in po_number ("AN POST", "AIRNAV IRELAND"). Promote it back IF it
        # looks like a multi-word name and carries no big number (which would mean it is a
        # category-total line, e.g. ESB Networks "Meter Reading Services 3,823,410").
        sup = (r.get("supplier_raw") or "").strip()
        po = (r.get("po_number") or "").strip()
        if not sup and po and NAME_LIKE.search(po) and not NUMERIC_NOISE.search(po):
            r["supplier_raw"] = po
            r["po_number"] = None
            sup = po
        # Anything still missing a supplier is NOT a clean supplier-level row (category totals,
        # blank cells) — downgrade so it is filterable and never ranked as a real supplier.
        if not sup:
            r["extraction_confidence"] = "low"
            r["caveat_text_detected"] = True
    return rows_out, {"status": "ok" if rows_out else "empty", "rows": len(rows_out), "confidence": conf}


# Within-source-file duplicate signature: a row identical to another in EVERY extracted field
# (same file + supplier + amount + description + PO + page + paid-flag + period) is an
# indistinguishable repeat the word-row clusterer emitted more than once (a table row captured
# twice, a header re-clustered), and summing both double-counts. A row that differs in ANY field
# — notably `description` — is a DISTINCT payment and MUST be kept (e.g. Courts had 9 lines that
# share a mis-parsed amount + truncated name but carry 9 different descriptions). DQ audit 2026-06-05.
DEDUP_SIG = [
    "source_file_hash",
    "supplier_raw",
    "amount_eur",
    "description",
    "po_number",
    "source_page_number",
    "paid_flag",
    "period",
]


def dedup_source_repeats(df: pl.DataFrame) -> tuple[pl.DataFrame, int]:
    """Drop indistinguishable within-file parser repeats (see DEDUP_SIG). Keeps the first
    occurrence; returns (deduped_df, n_dropped). Errs toward UNDER-deduping: any differing
    field preserves the row, so genuine distinct payments are never collapsed."""
    if df.is_empty():
        return df, 0
    keys = [c for c in DEDUP_SIG if c in df.columns]
    n = df.height
    out = df.unique(subset=keys, keep="first", maintain_order=True)
    return out, n - out.height


# Generic business/industry words that, standing ALONE as the whole normalised supplier name, do
# NOT identify a specific company — they are the remnant after a distinctive leading word was
# truncated at source (published "Construction Ltd" / "Aircraft Ltd" / "Ireland Energy Ltd" ->
# norm "CONSTRUCTION"/"AIRCRAFT"/"ENERGY"), or a too-generic published name ("Infrastructure DAC").
# Distinct from real one-word firms whose distinctive token survives (SODEXO, FUJITSU, ADSTON,
# ACCENTURE). Single-token match only, so "NBI INFRASTRUCTURE" / "DSV LOGISTICS" are NOT flagged.
GENERIC_SUPPLIER_NAME = frozenset(
    {
        "infrastructure",
        "energy",
        "construction",
        "aircraft",
        "shipping",
        "media",
        "technology",
        "partnership",
        "bundle",
        "group",
        "holdings",
        "services",
        "solutions",
        "systems",
        "engineering",
        "logistics",
        "properties",
        "developments",
        "consulting",
        "consultants",
        "management",
        "international",
        "contractors",
        # legal-form / geographic remnants left after the distinctive lead word was truncated
        # (e.g. "Deloitte LLP" -> "LLP", "[Brand] Electric Ltd" -> "ELECTRIC", "X UK Ltd" -> "UK")
        "llp",
        "electric",
        "europe",
        "uk",
        "ireland",
    }
)


def canonicalise_supplier_raw(df: pl.DataFrame) -> pl.DataFrame:
    """Evidence-based merge of known split entities BEFORE normalisation (no name fabrication —
    uses only strings already published in this data + the po_number signal). NBI: the National
    Broadband Plan contractor is published both as 'Infrastructure DAC' (po_number 'NBI', Dept
    Climate) and 'NBI Infrastructure DAC' (Dept Culture) — one legal entity. Rewrite the po=NBI
    'Infrastructure DAC' form to 'NBI Infrastructure DAC' so it merges with the NBI-prefixed rows
    and normalises to the identifiable 'NBI INFRASTRUCTURE' instead of the generic 'INFRASTRUCTURE'.
    DQ audit 2026-06-05 (A2)."""
    if df.is_empty() or "supplier_raw" not in df.columns or "po_number" not in df.columns:
        return df
    is_nbi = (pl.col("po_number").cast(pl.Utf8).str.to_uppercase().str.strip_chars() == "NBI") & (
        pl.col("supplier_raw").str.contains(r"(?i)\binfrastructure dac\b")
    )
    return df.with_columns(
        pl.when(is_nbi).then(pl.lit("NBI Infrastructure DAC")).otherwise(pl.col("supplier_raw")).alias("supplier_raw")
    )


def flag_unidentifiable_suppliers(df: pl.DataFrame) -> pl.DataFrame:
    """Downgrade extraction_confidence to 'low' where the normalised supplier name is empty
    (truncated to just a legal suffix, e.g. 'LTD'/'IRELAND LTD' -> '') or is a single generic
    business word (truncation remnant, GENERIC_SUPPLIER_NAME). Such rows have a real amount —
    they STAY summable (value_safe_to_sum untouched) — but no usable supplier identity, so the
    low-confidence flag lets a supplier ranking filter them out. Real one-word firms keep their
    distinctive token and are unaffected. DQ audit 2026-06-05 (A2/A4)."""
    if df.is_empty() or "supplier_normalised" not in df.columns:
        return df
    norm = pl.col("supplier_normalised")
    unidentifiable = (
        norm.is_null()
        | (norm.str.strip_chars() == "")
        | norm.str.to_lowercase().str.strip_chars().is_in(list(GENERIC_SUPPLIER_NAME))
    )
    conf = pl.col("extraction_confidence") if "extraction_confidence" in df.columns else pl.lit("high")
    return df.with_columns(pl.when(unidentifiable).then(pl.lit("low")).otherwise(conf).alias("extraction_confidence"))


def classify_and_flag(df: pl.DataFrame) -> pl.DataFrame:
    """supplier_normalised + supplier_class + privacy_status; quarantine DEFERRED."""
    if df.is_empty():
        return df
    df = (
        df.with_columns(
            name_norm_expr("supplier_raw").alias("supplier_normalised"),
            pl.col("supplier_raw")
            .map_elements(lambda s: bool(PUBLIC_BODY.search(s or "")), return_dtype=pl.Boolean)
            .alias("_pub"),
            pl.col("supplier_raw")
            .map_elements(lambda s: bool(COMPANY_SUFFIX.search(s or "")), return_dtype=pl.Boolean)
            .alias("_co"),
            pl.col("supplier_raw")
            .map_elements(lambda s: bool(FOREIGN_FORM.search(s or "")), return_dtype=pl.Boolean)
            .alias("_for"),
        )
        .with_columns(
            pl.when(pl.col("_pub"))
            .then(pl.lit("public_body"))
            .when(pl.col("_co"))
            .then(pl.lit("company"))
            .when(pl.col("_for"))
            .then(pl.lit("foreign_company"))
            .when(pl.col("supplier_raw").is_null() | (pl.col("supplier_raw").str.strip_chars() == ""))
            .then(pl.lit("unknown"))
            .otherwise(pl.lit("sole_trader_or_individual"))
            .alias("supplier_class"),
        )
        .with_columns(
            # privacy_status flags likely-personal rows (sole traders / individuals).
            pl.when(pl.col("supplier_class") == "sole_trader_or_individual")
            .then(pl.lit("review_personal_data"))
            .otherwise(pl.lit("ok"))
            .alias("privacy_status"),
            # QUARANTINE APPLIED: a likely-personal supplier is never displayable. Rows are RETAINED
            # for analysis/coverage (nothing dropped) — only the display flag is gated, so a
            # downstream UI / promotion must filter on public_display.
            (pl.col("supplier_class") != "sole_trader_or_individual").alias("public_display"),
            # po_committed / payment_actual are summable; contract_award_value is caution-only.
            # EXCLUDE public_body suppliers: a payment whose recipient is itself a public body is an
            # intergovernmental TRANSFER / grant (e.g. TII -> county-council road grants = €2.5bn /
            # 32% of this fact), NOT private procurement. Summing them inflates "procurement spend"
            # and triple-counts the same euro (TII grant -> council -> contractor in la_payments_fact
            # -> the contractor's eTenders/TED award). They are RETAINED (public_display stays True)
            # but never summed. DQ audit 2026-06-05; supplier_class is derived in the block above.
            (
                pl.col("amount_semantics").is_in(["po_committed", "payment_actual"])
                & pl.col("amount_eur").is_not_null()
                & (pl.col("amount_eur") > 0)
                # Belt-and-braces against an order/PO number read as an amount (the recurring
                # €30bn/€400m bug on reading-order PDFs, e.g. dept_justice's "109245 …" lines): no
                # single public-body PO/payment line is ≥ €100m, so such a value is a parse error,
                # never summable. Mirrors the LA extractor's post-guard. The row is RETAINED (low
                # confidence) for audit but excluded from any total.
                & (pl.col("amount_eur") < 100_000_000)
                # A row with NO identifiable supplier is never summable spend: it is either a
                # category/quarterly SUBTOTAL (e.g. dept_social_protection PDFs emit 4 blank-supplier
                # 'Sum:' rows/year worth €428m that would DOUBLE-COUNT the per-supplier rows) or a
                # parse gap. CATEGORY_WORD misses these because the total label sits in the
                # description, not the blank supplier cell. Rows kept (low-conf) for audit, never
                # totalled. GENERIC_SUPPLIER_NAME single-word firms keep a non-empty normalised name
                # and stay summable — only a TRULY empty supplier_normalised is excluded here.
                & pl.col("supplier_normalised").is_not_null()
                & (pl.col("supplier_normalised").str.strip_chars() != "")
                & (pl.col("supplier_class") != "public_body")
            ).alias("value_safe_to_sum"),
        )
        .drop(["_pub", "_co", "_for"])
    )
    return df


# ============================================================================ main
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true", help="harvest-only: print candidate files, no parse")
    ap.add_argument("--only", default="", help="comma-separated publisher ids")
    ap.add_argument(
        "--max-files", type=int, default=None, help="cap files parsed per publisher (default: all — full history)"
    )
    ap.add_argument("--max-pages", type=int, default=None, help="cap pages per PDF")
    ap.add_argument(
        "--merge",
        action="store_true",
        help="with --only: update JUST those publishers inside the existing fact (all other "
        "publishers kept) instead of overwriting the whole parquet — avoids re-downloading the "
        "full ~600-file history and prevents the clobber-to-N-publishers footgun",
    )
    ap.add_argument("--refetch", action="store_true", help="bypass the bronze cache and re-download every file")
    args = ap.parse_args()
    only = {x.strip() for x in args.only.split(",") if x.strip()} or None
    if args.merge and not only:
        raise SystemExit("--merge requires --only (it updates the named publishers in place)")

    pubs = [p for p in PUBLISHERS if not only or p["id"] in only]
    print(f"{'=' * 80}\nPUBLIC-BODY EXTRACT{' (LIST MODE)' if args.list else ''} — {len(pubs)} publishers\n{'=' * 80}")

    all_rows: list[dict] = []
    per_pub: list[dict] = []
    for cf in pubs:
        files = harvest_files(cf)
        print(f"\n[{cf['id']:<22}] {cf['name']}  (tier {cf['tier']}, {cf['amount_semantics']})")
        print(f"   listing: {cf['listing_url']}")
        print(f"   files harvested: {len(files)}")
        if not files:
            REPORT.record_zero_harvest(publisher_id=cf["id"], publisher_name=cf["name"], listing_url=cf["listing_url"])
        for u in files[:8]:
            print(f"     - {u.rsplit('/', 1)[-1][:70]}")
        if len(files) > 8:
            print(f"     … +{len(files) - 8} more")
        if args.list:
            per_pub.append({"id": cf["id"], "files": len(files)})
            continue

        pub_rows, parsed, ok_files, skipped = [], 0, 0, 0
        breaker = Breaker()
        file_list = files[: args.max_files]
        for i, u in enumerate(file_list):
            ext = next((e for e in DATA_EXT if u.lower().split("?")[0].endswith(e)), "")
            fmt = {".pdf": "pdf", ".xlsx": "xlsx", ".xls": "xls", ".csv": "csv"}.get(ext)
            if not fmt:
                continue
            b, fresh = fetch_to_bronze(cf["id"], u, ext, refetch=args.refetch)
            if not b and fresh:
                time.sleep(30.0)  # cool off once before the single retry
                b, fresh = fetch_to_bronze(cf["id"], u, ext, refetch=True)
            bad = classify_body(b, b"%PDF" if fmt == "pdf" else None) if b else None
            if not b or bad:
                breaker.record(False)
                err = bad or LAST_ERR.get("error_class", "unknown")
                REPORT.record_failure(
                    publisher_id=cf["id"],
                    publisher_name=cf["name"],
                    url=u,
                    listing_url=cf["listing_url"],
                    error_class=err,
                    http_status=LAST_ERR.get("http_status"),
                    attempts=4 if not b else 1,
                )
                print(f"     ! download failed ({err}): {u.rsplit('/', 1)[-1][:50]}")
                if breaker.tripped:
                    rest = len(file_list) - i - 1
                    REPORT.record_breaker_trip(publisher_id=cf["id"], publisher_name=cf["name"], files_skipped=rest)
                    print(
                        f"     !! breaker tripped: {breaker.consecutive} consecutive failures — skipping {rest} remaining files"
                    )
                    break
                continue
            breaker.record(True)
            write_sentinel("public_body", cf["id"], u)
            try:
                rows, stat = emit_rows(cf, u, b, fmt, args.max_pages)
            except Exception as e:  # one malformed file must not abort the run
                skipped += 1
                print(f"     ! parse error ({type(e).__name__}): {u.rsplit('/', 1)[-1][:46]}")
                continue
            parsed += 1
            if rows:
                ok_files += 1
            pub_rows.extend(rows)
            print(
                f"     -> {u.rsplit('/', 1)[-1][:48]:<48} {stat['status']:<8} "
                f"rows={stat['rows']} conf={stat['confidence']}"
            )
        all_rows.extend(pub_rows)
        per_pub.append(
            {
                "id": cf["id"],
                "name": cf["name"],
                "tier": cf["tier"],
                "files_seen": len(files),
                "files_parsed": parsed,
                "files_skipped": skipped,
                "files_with_rows": ok_files,
                "rows": len(pub_rows),
                "amount_semantics": cf["amount_semantics"],
                "privacy_risk": cf["privacy_risk"],
            }
        )

    if args.list:
        print(f"\n{'=' * 80}\nLIST DONE. Lock URLs into config, then run without --list.")
        return

    if not all_rows:
        print("\nno rows extracted")
        REPORT.write()
        for line in REPORT.summary_lines():
            print(line)
        return

    SCHEMA_COLS = [
        "publisher_id",
        "publisher_name",
        "publisher_type",
        "sector",
        "source_landing_url",
        "source_file_url",
        "source_file_hash",
        "period",
        "year",
        "quarter",
        "supplier_raw",
        "supplier_normalised",
        "amount_eur",
        "amount_semantics",
        "value_safe_to_sum",
        "description",
        "po_number",
        "paid_flag",
        "source_row_number",
        "source_page_number",
        "parser_name",
        "parser_version",
        "extraction_status",
        "extraction_confidence",
        "caveat_text_detected",
        "supplier_class",
        "privacy_status",
        "public_display",
        "source_caveat",
    ]
    df = pl.DataFrame(all_rows, infer_schema_length=None)
    df, rows_deduped = dedup_source_repeats(df)
    if rows_deduped:
        print(f"\ndeduped {rows_deduped:,} within-file parser repeats (indistinguishable rows)")
    df = canonicalise_supplier_raw(df)
    df = classify_and_flag(df)
    df = flag_unidentifiable_suppliers(df)
    df = df.select([c for c in SCHEMA_COLS if c in df.columns])

    # MERGE MODE: fold the freshly-parsed publishers into the existing fact instead of replacing
    # it. Drop any prior rows for the publishers we just (re)parsed — idempotent — then concat the
    # untouched publishers back on. The kept rows are already fully classified (same SCHEMA_COLS),
    # so no re-classification is needed; we align dtypes defensively before concat.
    merged_kept_pubs: list[dict] = []
    if args.merge and OUT_FACT.exists():
        existing = pl.read_parquet(OUT_FACT)
        sel_ids = [p["id"] for p in pubs]
        kept = existing.filter(~pl.col("publisher_id").is_in(sel_ids))
        print(
            f"\nMERGE: existing fact {existing.height:,} rows / {existing['publisher_id'].n_unique()} "
            f"publishers; replacing {existing.height - kept.height:,} rows for {sel_ids}, "
            f"keeping {kept.height:,} rows for {kept['publisher_id'].n_unique()} other publishers"
        )
        df = df.select(existing.columns).cast(dict(existing.schema))
        df = pl.concat([kept, df], how="vertical")
        # Preserve the by_publisher coverage entries for the publishers we did NOT reparse.
        if OUT_COV.exists():
            with contextlib.suppress(Exception):
                old_cov = json.loads(OUT_COV.read_text(encoding="utf-8"))
                merged_kept_pubs = [e for e in old_cov.get("by_publisher", []) if e.get("id") not in sel_ids]

    # PRIVACY INVARIANT (runtime, -O-proof): no displayable row may be a likely person.
    leaked = df.filter(pl.col("public_display") & (pl.col("supplier_class") == "sole_trader_or_individual"))
    if leaked.height:
        raise RuntimeError(
            f"privacy quarantine breached: {leaked.height} sole_trader_or_individual rows "
            "left public_display=True; refusing to write public_payments_fact"
        )

    OUT_FACT.parent.mkdir(parents=True, exist_ok=True)
    save_parquet(df, OUT_FACT)

    print(f"\n{'=' * 80}\nGOLD-CANDIDATE WRITTEN\n{'=' * 80}")
    print(f"rows: {df.height:,}  ->  {OUT_FACT}")
    print(df.group_by("supplier_class").len().sort("len", descending=True))
    safe = df.filter(pl.col("value_safe_to_sum"))
    print(
        f"\nvalue_safe_to_sum rows: {safe.height:,}  "
        f"sum=€{(safe['amount_eur'].sum() or 0):,.0f} (po_committed+payment_actual only)"
    )

    by_publisher = merged_kept_pubs + per_pub  # kept (merge mode) + freshly parsed this run
    cov = {
        "publishers_attempted": len(by_publisher),
        "publishers_with_rows": sum(p.get("rows", 0) > 0 for p in by_publisher),
        "rows_extracted": df.height,
        "rows_deduped_within_file": rows_deduped,
        "rows_public_display": int(df["public_display"].sum()),
        "rows_review_personal_data": int((df["privacy_status"] == "review_personal_data").sum()),
        "rows_quarantined": int((~df["public_display"]).sum()),
        "supplier_class_counts": {
            r["supplier_class"]: r["len"] for r in df.group_by("supplier_class").len().iter_rows(named=True)
        },
        "amount_semantics_counts": {
            r["amount_semantics"]: r["len"] for r in df.group_by("amount_semantics").len().iter_rows(named=True)
        },
        "value_safe_to_sum_rows": safe.height,
        "value_safe_to_sum_total_eur": float(safe["amount_eur"].sum() or 0),
        "by_publisher": by_publisher,
        "privacy_quarantine_applied": True,
        "schema_version": 1,
        "parser_version": PARSER_VERSION,
        "generated_at": datetime.now(UTC).isoformat(timespec="seconds"),
        "caveat": "GOLD-CANDIDATE (sandbox, pre-promotion). One row per source line. "
        "amount_semantics distinguishes po_committed/payment_actual/contract_award_value; "
        "only value_safe_to_sum (po_committed+payment_actual, EXCLUDING public_body "
        "recipients which are intergovernmental transfers/grants e.g. TII road grants) "
        "may be totalled, labelled 'ordered/paid', never mixed with award ceilings. "
        "PRIVACY QUARANTINE APPLIED: "
        "rows flagged privacy_status=review_personal_data (likely sole traders / "
        "individuals) are marked public_display=False and must be filtered out before any "
        "UI use; they are retained here for analysis only. "
        "A line is a purchase order or payment record, not evidence of influence.",
    }
    OUT_COV.write_text(json.dumps(cov, indent=2), encoding="utf-8")
    print(f"wrote coverage {OUT_COV}")

    report_path = REPORT.write()
    lines = REPORT.summary_lines()
    if lines:
        print(f"\nfetch-failure report -> {report_path}")
        for line in lines:
            print(line)


if __name__ == "__main__":
    main()
