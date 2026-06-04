"""PHASE 4-6 (PRE-ETL, sandbox): public-body PO/payment extractor -> public_payments_fact.

Promotes the two PRE-ETL sample readers into ONE config-driven extractor:
  - PDF header-anchored reader  <- sample_extract_procurement_pdf.py
  - XLSX/CSV tabular reader      <- sample_extract_procurement.py
  - gold conventions (name_norm, supplier_class, value-safe-to-sum, coverage JSON,
    zstd parquet)                <- procurement_etenders_extract.py

It is NOT wired into pipeline.py and writes a GOLD-CANDIDATE to data/sandbox/parquet/
(LA precedent: promote to data/gold/parquet/ only on a separate go-ahead). One row per
source line, with full provenance (plan PROCUREMENT_SEMISTATE_EXPANSION_PLAN.md Phase 5).

SCOPE / OWNERSHIP (multi-context split, 2026-06-03): this extractor owns the publishers
the GENERIC reader handles cleanly (Tier A + the corrected Tier C). HSE + Tusla are owned
by procurement_hse_tusla_parser.py (bespoke column-x specs — the generic reader misparses
them); local-authority POs and the BUDGET tier are owned by other context windows. All
emit THIS same schema so the layers union at promotion time.

PRIVACY: classification is computed (supplier_class -> privacy_status) but the quarantine
is DEFERRED by request — public_display defaults True for every row and NOTHING is dropped.
coverage records privacy_quarantine_applied=false so the deferral is explicit, not silent.

Run:
  ./.venv/Scripts/python.exe pipeline_sandbox/procurement_public_body_extract.py --list            # harvest-only (lock URLs)
  ./.venv/Scripts/python.exe pipeline_sandbox/procurement_public_body_extract.py --list --only ie_opw,ie_tii
  ./.venv/Scripts/python.exe pipeline_sandbox/procurement_public_body_extract.py                   # full ingest
  ./.venv/Scripts/python.exe pipeline_sandbox/procurement_public_body_extract.py --only ie_hse --max-files 2
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import io
import json
import re
import subprocess
import sys
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

import fitz  # PyMuPDF
import polars as pl
import requests

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8")

from cro_normalise import name_norm_expr  # noqa: E402

H = {"User-Agent": "Mozilla/5.0 (dail-tracker research probe)"}
TMP = Path("c:/tmp/procurement_publishers")
OUT_FACT = ROOT / "data/sandbox/parquet/public_payments_fact.parquet"
OUT_COV = ROOT / "data/_meta/public_payments_coverage.json"
PARSER_VERSION = "0.1.0"

DATA_EXT = (".pdf", ".xlsx", ".xls", ".csv")

# ----------------------------------------------------------------------------- regexes
MONEY_RE = re.compile(r"(?:€|EUR)?\s?\d{1,3}(?:,\d{3})+(?:\.\d{2})?|\d+\.\d{2}")
NUM_RE = re.compile(r"\d[\d,]*(?:\.\d+)?")
HREF_RE = re.compile(r"""href\s*=\s*["']([^"']+)["']""", re.I)
DIGIT_PREFIX = re.compile(r"^(?:\d{3,}\s+){1,3}")  # strips a leading PO/vendor-ID run
PERIOD_RE = re.compile(r"(?:^|[^0-9])(20[12]\d)(?:[^0-9]|$)")
QUARTER_RE = re.compile(r"q\s?([1-4])|quarter[\s_-]?([1-4])|qtr[\s_-]?([1-4])", re.I)

ROLE_RE = {
    "supplier": re.compile(r"supplier|payee|vendor|provider|customer|recipient|\bname\b", re.I),
    "amount": re.compile(r"amount|total|value|gross|\beuro\b|€|\bpaid\b|\bvat\b|ledger|payment\b", re.I),
    "description": re.compile(r"descript|\bdesc\b|detail|categor|service|goods|nature|\bgl\b|main gl", re.I),
    "po": re.compile(r"\border\b|\bpo\b|\bpor\b|referen|\bref\b|\bnumber\b|invoice|\bdoc\b|transaction", re.I),
    "period": re.compile(r"period|quarter|\bqtr\b|\bdate\b|\byear\b|posting|month", re.I),
    "paid": re.compile(r"\bpaid\b|payment type|status|no\.? of payments", re.I),
}
CAVEAT_RE = re.compile(r"\bvat\b|exclud|inclus|indicativ|not (a )?payment|net of|estimate|note:|please note", re.I)
COMPANY_SUFFIX = re.compile(
    r"\b(ltd|limited|dac|plc|clg|llp|teo|teoranta|t/a|uc|inc|llc|gmbh|company|co\.|group|"
    r"services|solutions|consult|engineer|partners|associates|holdings|university|college|"
    r"council|hse|board|institute|ireland|technolog|systems|media|hotel|centre|&)\b", re.I)
FOREIGN_FORM = re.compile(r"\b(gmbh|s\.?a\.?|n\.?v\.?|s\.?a\.?s|s\.?p\.?a|inc|llc|\bpty\b|\bab\b|\bbv\b|\boy\b|srl|sl|sarl|aps|kft|ltda)\b", re.I)
PUBLIC_BODY = re.compile(r"\b(county council|city council|university|institute of technology|department of|office of|\bHSE\b|health service|an garda|údarás|udaras|education and training board|\bETB\b|local authority|national \w+ authority|\bOPW\b|hospital)\b", re.I)
# Drops title/threshold rows that masquerade as a supplier (the page heading bleeds into the
# supplier column with the literal €20,000 threshold as its amount). Plural "Purchase Orders"
# and "Payments greater than/over €20,000" headings leaked through the singular-only pattern.
CATEGORY_WORD = re.compile(r"^\s*(total|category total|sum|subtotal|grand total|all suppliers|various|publication of|purchase orders?|payments? (greater|over|to suppliers)|payments? greater than)\b", re.I)
# exclude policy/guidance/privacy/contract docs when harvesting period data files
POLICY_RE = re.compile(
    r"guide|guidelin|\bplan\b|policy|circular|strategy|manual|terms|fin.?07|privacy|"
    r"prompt.?payment|appendix|procedure|annual.?report|statement|setup|form|charter|scheme",
    re.I)
DATA_FILE_RE = re.compile(r"q[1-4]\b|qtr|quarter|20[12]\d|h[12]\b|over.?20|over.?25|payment|purchase|\bpo[s]?\b", re.I)
NAV_HINT = re.compile(
    r"purchase|procure|over.?20|over.?25|20k|payment|quarter|qtr|finance|"
    r"publication|spend|supplier|expenditure|disclosure|financial", re.I)
MERGE_GAP = 22.0


# ============================================================================ CONFIG
# amount_semantics controlled vocab (PROCUREMENT_INVESTIGATION.md value taxonomy):
#   po_committed     -> "ordered €X"  (PO-over-20k order lists)  summable
#   payment_actual   -> "paid €X"     (payments/paid lists)      summable (true spend)
#   contract_award_value -> "awarded €X" (Tailte contracts)      caution
# listing_url = page to harvest period files from; direct_files = known-good file URLs
# (used as a floor so a publisher still yields data if its listing is JS/awkward).
def cfg(pid, name, ptype, sector, *, listing, semantics, grain, privacy="low",
        tier="A", direct=None, include=None, caveat="") -> dict:
    return {"id": pid, "name": name, "ptype": ptype, "sector": sector,
            "listing_url": listing, "amount_semantics": semantics, "grain": grain,
            "privacy_risk": privacy, "tier": tier, "direct_files": direct or [],
            "include": re.compile(include, re.I) if include else None, "caveat": caveat}


PUBLISHERS: list[dict] = [
    # ---- Tier A: clean tabular / high-confidence PDF -------------------------------
    cfg("ie_opw", "Office of Public Works", "state_body", "property_land",
        listing="https://www.gov.ie/en/office-of-public-works/collections/payments-greater-than-20000/",
        semantics="payment_actual", grain="payment",
        direct=["https://assets.gov.ie/static/documents/b526ff76/OPW_Payments_of_20000_or_over_in_Q1_2026.xlsx"]),
    cfg("dept_climate", "Dept of Climate, Energy and the Environment", "department", "central_government",
        listing="https://www.gov.ie/en/department-of-climate-energy-and-the-environment/collections/payments-over-20000/",
        semantics="payment_actual", grain="payment",
        direct=["https://assets.gov.ie/static/documents/ae8b1a0a/DPER_Payments_over_20K_Q1_2026_Report.xlsx"]),
    cfg("dept_defence", "Department of Defence", "department", "central_government",
        listing="https://www.gov.ie/en/department-of-defence/collections/purchase-orders-over-20000/",
        semantics="po_committed", grain="purchase_order"),
    cfg("dept_culture", "Department of Culture, Communications and Sport", "department", "central_government",
        listing="https://www.gov.ie/en/department-of-culture-communications-and-sport/collections/purchase-orders/",
        semantics="po_committed", grain="purchase_order",
        caveat="contains very large NBI infrastructure POs; check outlier share before any total"),
    cfg("ie_teagasc", "Teagasc", "semi_state", "agri_food_marine",
        listing="https://www.teagasc.ie/about/corporate-responsibility/information-for-suppliers/",
        semantics="po_committed", grain="purchase_order"),
    cfg("ie_bordbia", "Bord Bia", "semi_state", "agri_food_marine",
        listing="https://www.bordbia.ie/about/governance/corporate-governance/purchase-orders/",
        semantics="po_committed", grain="purchase_order"),
    cfg("ie_bim", "Bord Iascaigh Mhara (BIM)", "semi_state", "agri_food_marine",
        listing="https://bim.ie/about/corporate-governance/purchase-orders-over-20k/",
        semantics="po_committed", grain="purchase_order",
        caveat="amounts excluding VAT"),
    cfg("ie_cib", "Citizens Information Board", "agency", "social",
        listing="https://www.citizensinformationboard.ie/en/freedom_of_information/financial_information/payments_or_purchase_orders_for_goods_and_services.html",
        semantics="payment_actual", grain="payment"),
    cfg("ie_hea", "Higher Education Authority", "agency", "education",
        listing="https://hea.ie/about-us/public-sector-information/",
        semantics="payment_actual", grain="payment", privacy="low"),

    # ---- Tier B: OWNED BY A SEPARATE CONTEXT (procurement_hse_tusla_parser.py) -----
    # HSE + Tusla need bespoke per-publisher column-x specs (the generic header-anchored
    # reader misparses them: HSE fuses amount+quarter+date, Tusla's vendor bleeds into the
    # amount column). NTPF + SVUH (health, privacy=high) de-scoped here too pending that
    # context's reconciliation. Their output merges into THIS schema later — do not re-add
    # HSE/Tusla here or the generic reader will produce duplicate low-quality rows.

    # ---- Tier C: needed a corrected listing URL or a parser fix --------------------
    cfg("ie_tii", "Transport Infrastructure Ireland", "agency", "transport",
        listing="https://www.tii.ie/en/compliance/payments/", semantics="payment_actual",
        grain="payment", tier="C",
        direct=["https://websitecms.tii.ie/media/sw3dzt2l/tii-payments-q1-2025-over-20k.csv"],
        caveat="CSV carries a category-total row (~€1.2bn) that must be excluded from any sum"),
    cfg("ie_revenue", "Revenue Commissioners", "agency", "regulator",
        listing="https://www.revenue.ie/en/corporate/statutory-obligations/freedom-of-information/section8/procurement.aspx",
        semantics="payment_actual", grain="payment", tier="C",
        direct=["https://www.revenue.ie/en/corporate/documents/procurement/payments-over-20000-quarter4-2025.pdf"]),
    cfg("ie_atu", "Atlantic Technological University", "education_body", "education",
        listing="https://www.atu.ie/freedom-of-information/freedom-of-information-financial-information",
        semantics="payment_actual", grain="payment", privacy="medium", tier="C",
        direct=["https://www.atu.ie/app/uploads/2026/03/atu-payments-purchase-orders-q1-2025.pdf"],
        caveat="supplier published with a leading numeric supplier-ID; stripped on read"),
    cfg("ie_nta", "National Transport Authority", "agency", "transport",
        listing="https://www.nationaltransport.ie/publications/2026-purchase-orders-e20000-and-over/",
        semantics="po_committed", grain="purchase_order", tier="C",
        direct=["https://www.nationaltransport.ie/wp-content/uploads/2026/05/Purchase-Orders-20k-and-over-Quarter-1-2026.pdf"]),
    cfg("ie_marine", "Marine Institute", "agency", "agri_food_marine",
        listing="https://www.marine.ie/site-area/about-us/purchase-orders",
        semantics="po_committed", grain="purchase_order", tier="C",
        direct=["https://marine.ie/sites/default/files/MIFiles/Docs/CS/Purchase%20Orders%20Qtr%201%202026.pdf"]),
    cfg("ie_esbnetworks", "ESB Networks DAC", "semi_state", "energy_utilities",
        listing="https://www.esbnetworks.ie/about-us/company/publication-scheme/financial-information",
        semantics="payment_actual", grain="payment", tier="C",
        caveat="prior sample was a category-total page; harvesting supplier-level file"),
    cfg("ie_tailte", "Tailte Éireann", "state_body", "property_land",
        listing="https://tailte.ie/category/publications/", semantics="po_committed",
        grain="purchase_order", tier="C",
        include=r"purchase|payment|po[s]?[-_ ]?over|20[,]?000|over.?20k",
        caveat="Purchase-Orders quarterly files (PO grain); contracts-awarded list excluded"),
    cfg("dept_housing", "Department of Housing, Local Government and Heritage", "department",
        "central_government",
        listing="https://www.gov.ie/en/department-of-housing-local-government-and-heritage/collections/purchase-orders-and-payments-over-20000/",
        semantics="payment_actual", grain="payment", tier="C",
        caveat="prior sample was a privacy statement; using the gov.ie payments collection"),
    cfg("ie_cdetb", "City of Dublin ETB", "education_body", "education",
        listing="https://www.cityofdublinetb.ie/about-us/finance-and-procurement/procurement/",
        semantics="po_committed", grain="purchase_order", privacy="medium", tier="C",
        include=r"purchase|payment|po[s]?[-_ ]?over|20[,]?000|quarter|q[1-4]",
        caveat="prior sample was the procurement policy; excluding policy docs"),
    cfg("ie_enterprise_ireland", "Enterprise Ireland", "semi_state", "enterprise_tourism",
        listing="https://www.enterprise-ireland.com/en/about-us/our-policies/purchase-orders-over-20000",
        semantics="po_committed", grain="purchase_order", tier="C",
        include=r"purchase|payment|po[s]?[-_ ]?over|20[,]?000|over.?20k",
        caveat="FOI-only: no clean PO file located at this URL (only sustainability/climate reports surface)"),

    # ---- Tier D: discovery sweep 2026-06-04 (doc/PROCUREMENT_SOURCE_DISCOVERY_2026_06_04.md) --
    # Probe-confirmed, generic-reader-clean. Held back for bespoke/render passes (NOT here):
    #   Beaumont + Pobal (dual/MIXED PO+payment grain — need value_kind split),
    #   Coimisiún na Meán + Irish Prison Service (scanned PDFs — need OCR),
    #   Garda (sampler hit a fleet report — needs the right PO subpage),
    #   UCD / SETU / CHI / SEAI / EPA (no links via landing — JS/403, EPA serves .php HTML).
    cfg("ie_ntma", "National Treasury Management Agency (NTMA)", "state_body", "finance",
        listing="https://www.ntma.ie/information-pages/freedom-of-information/freedom-of-information-publication-scheme/financial-information",
        semantics="payment_actual", grain="payment", tier="D",
        caveat="one quarterly scheme covers 6 business units incl NDFA (ADM/Nat-Debt/ISIF/NDFA/FIF/ICNF); "
               "do NOT also wire ie_ndfa or its rows double-count"),
    cfg("ie_courts", "Courts Service of Ireland", "agency", "justice",
        listing="https://www.courts.ie/publications/purchase-orders-greater-than-20k",
        semantics="po_committed", grain="purchase_order", tier="D",
        include=r"purchase-order|over-20|po[s]?[-_ ]?over"),
    cfg("ie_sportireland", "Sport Ireland", "agency", "sport",
        listing="https://www.sportireland.ie/about-us/freedom-of-information/financial-information",
        semantics="po_committed", grain="purchase_order", tier="D",
        caveat="single rolling PO log (not per-quarter); period likely null"),
    cfg("ie_tudublin", "Technological University Dublin", "education_body", "education",
        listing="https://www.tudublin.ie/explore/governance-and-compliance/foi/foi-publication-scheme/",
        semantics="po_committed", grain="purchase_order", tier="D",
        include=r"po-report|purchase-order|over-?20k"),
    cfg("ie_mtu", "Munster Technological University (MTU)", "education_body", "education",
        listing="https://www.mtu.ie/about-mtu/legal/freedom-of-information/",
        semantics="po_committed", grain="purchase_order", tier="D",
        include=r"pos?-over-?20k|purchase-order|po[s]?[-_ ]?over",
        caveat="TODO harvest gap: probe finds 15 PO PDFs from this URL but the extractor's one-hop "
               "crawl/include misses them (files live under /media/.../foi/financial-information/) — "
               "needs a deeper crawl or a direct sub-page listing_url; 0 rows until fixed. "
               "Also exclude the 'Procurement Listing' tender-register xlsx (Tendered-by columns)"),
]


# ============================================================================ fetch
def _curl(url: str) -> bytes | None:
    try:
        p = subprocess.run(["curl", "-sS", "-k", "-L", "--max-time", "90", "-A", H["User-Agent"], url],
                           capture_output=True, timeout=120)
        return p.stdout if p.returncode == 0 and p.stdout else None
    except Exception:
        return None


def fetch_bytes(url: str) -> bytes | None:
    try:
        r = requests.get(url, headers=H, timeout=90, allow_redirects=True)
        r.raise_for_status()
        return r.content
    except Exception:
        return _curl(url)


def fetch_text(url: str) -> str | None:
    b = fetch_bytes(url)
    return b.decode("utf-8", "ignore") if b else None


# ============================================================================ harvest
def harvest_files(cf: dict, crawl_cap: int = 12) -> list[str]:
    """Collect ALL period data-file links for a publisher: landing page + one-hop crawl,
    minus policy/guidance docs. Union with direct_files. Honours an optional include re."""
    found: list[str] = list(cf["direct_files"])
    html = fetch_text(cf["listing_url"])
    if html:
        def scan(page_html: str, base: str) -> list[str]:
            out = []
            for href in HREF_RE.findall(page_html):
                low = href.lower().split("?")[0]
                if not any(low.endswith(e) for e in DATA_EXT):
                    continue
                if POLICY_RE.search(href):
                    continue
                if not DATA_FILE_RE.search(href):
                    continue
                if cf["include"] and not cf["include"].search(href):
                    continue
                out.append(urljoin(base, href))
            return out
        hits = scan(html, cf["listing_url"])
        if not hits:  # one-hop crawl same-host nav links
            host = urlparse(cf["listing_url"]).netloc
            subs, seen = [], set()
            for href in HREF_RE.findall(html):
                full = urljoin(cf["listing_url"], href)
                low = full.lower().split("?")[0]
                if urlparse(full).netloc != host or full == cf["listing_url"]:
                    continue
                if any(low.endswith(e) for e in DATA_EXT):
                    continue
                if NAV_HINT.search(href) and full not in seen:
                    seen.add(full)
                    subs.append(full)
            for s in subs[:crawl_cap]:
                sub_html = fetch_text(s)
                if sub_html:
                    hits.extend(scan(sub_html, s))
        found.extend(hits)
    # dedup by basename (same file served via two hosts -> one entry, e.g. TII)
    seen, uniq = set(), []
    for u in found:
        key = u.rsplit("/", 1)[-1].split("?")[0].lower()
        if key not in seen:
            seen.add(key)
            uniq.append(u)
    return uniq


# ============================================================================ readers
def to_eur(token) -> float | None:
    if token is None:
        return None
    if isinstance(token, (int, float)):
        return float(token)
    s = str(token).strip()
    neg = s.startswith("(") and s.endswith(")")
    m = NUM_RE.search(s)
    if not m:
        return None
    with contextlib.suppress(ValueError):
        v = float(m.group().replace(",", ""))
        return -v if neg else v
    return None


def clean_supplier(s) -> str:
    return DIGIT_PREFIX.sub("", str(s or "")).strip(" -:|,")


def period_from_url(url: str) -> tuple[str | None, int | None, int | None]:
    name = url.rsplit("/", 1)[-1]
    y = PERIOD_RE.search(name)
    year = int(y.group(1)) if y else None
    q = QUARTER_RE.search(name)
    quarter = next((int(g) for g in (q.groups() if q else []) if g), None) if q else None
    period = (f"{year}-Q{quarter}" if year and quarter else (str(year) if year else None))
    return period, year, quarter


# ---- PDF (header-anchored) ----
def cluster_word_rows(page, ytol: float = 3.0) -> list[list]:
    words = page.get_text("words")
    words.sort(key=lambda w: (round(w[1] / ytol), w[0]))
    rows, cur, cur_y = [], [], None
    for w in words:
        y = w[1]
        if cur_y is None or abs(y - cur_y) <= ytol:
            cur.append(w)
            cur_y = y if cur_y is None else cur_y
        else:
            rows.append(cur)
            cur, cur_y = [w], y
    if cur:
        rows.append(cur)
    return rows


def find_header(rows: list[list]):
    best, best_hits = None, 1
    for r in rows[:18]:
        text = " ".join(w[4] for w in r)
        hits = sum(bool(rx.search(text)) for rx in ROLE_RE.values())
        has_anchor = ROLE_RE["supplier"].search(text) or ROLE_RE["amount"].search(text)
        if hits >= 2 and has_anchor and hits > best_hits:
            best, best_hits = r, hits
    return best


def header_columns(header: list) -> list[dict]:
    ws = sorted(header, key=lambda w: w[0])
    cols: list[dict] = []
    for w in ws:
        if cols and w[0] - cols[-1]["x1"] < MERGE_GAP:
            cols[-1]["label"] += " " + w[4]
            cols[-1]["x1"] = max(cols[-1]["x1"], w[2])
        else:
            cols.append({"label": w[4], "x0": w[0], "x1": w[2]})
    for c in cols:
        c["center"] = (c["x0"] + c["x1"]) / 2
    return cols


def assign_role(cols: list[dict]) -> dict[str, int]:
    roles: dict[str, int] = {}
    for role, rx in ROLE_RE.items():
        cands = [i for i, c in enumerate(cols) if rx.search(c["label"])]
        if cands:
            roles[role] = cands[-1] if role == "amount" else cands[0]
    return roles


def row_to_cols(words: list, cols: list[dict]) -> list[str]:
    bounds = [(cols[i]["center"] + cols[i + 1]["center"]) / 2 for i in range(len(cols) - 1)]
    buckets: list[list] = [[] for _ in cols]
    for w in sorted(words, key=lambda w: w[0]):
        c = (w[0] + w[2]) / 2
        idx = 0
        while idx < len(bounds) and c > bounds[idx]:
            idx += 1
        buckets[idx].append(w[4])
    return [" ".join(b).strip(" -:|") for b in buckets]


def refine_roles(cols, roles, records):
    if not records:
        return roles
    def numfrac(i):
        vals = [r[i] for r in records if i < len(r) and r[i]]
        return sum(to_eur(v) is not None for v in vals) / len(vals) if vals else 0.0
    amt_cands = [i for i, c in enumerate(cols) if ROLE_RE["amount"].search(c["label"])]
    if amt_cands:
        roles["amount"] = max(amt_cands, key=numfrac)
    sup_cands = [i for i, c in enumerate(cols) if ROLE_RE["supplier"].search(c["label"])]
    if sup_cands:
        roles["supplier"] = min(sup_cands, key=numfrac)
    return roles


def read_pdf(b: bytes, max_pages: int | None) -> dict:
    doc = fitz.open(stream=b, filetype="pdf")
    npages = doc.page_count
    limit = min(npages, max_pages) if max_pages else npages
    cols, header_label, page0 = [], "", ""
    for i in range(min(npages, 3)):
        rows = cluster_word_rows(doc[i])
        if i == 0:
            page0 = doc[i].get_text("text")
        h = find_header(rows)
        if h:
            cols = header_columns(h)
            header_label = " | ".join(c["label"] for c in cols)
            break
    roles = assign_role(cols) if cols else {}
    out_rows, digital_chars = [], 0
    for i in range(limit):
        page = doc[i]
        digital_chars += len(page.get_text("text").strip())
        if not cols:
            continue
        for wrow in cluster_word_rows(page):
            xs = [w[4] for w in wrow]
            if not any(MONEY_RE.search(t) for t in xs):
                continue
            out_rows.append((i + 1, row_to_cols(wrow, cols)))
    doc.close()
    roles = refine_roles(cols, roles, [r for _, r in out_rows]) if cols else roles
    return {"digital": digital_chars > 200, "cols": cols, "header_label": header_label,
            "roles": roles, "rows": out_rows, "page0": page0, "pages": npages}


# ---- XLSX / CSV ----
def read_xlsx(b: bytes):
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True).active
    raw = [list(r) for r in ws.iter_rows(values_only=True)]
    full = " ".join(str(c) for row in raw[:6] for c in row if c is not None)
    def score(row):
        return sum(any(rx.search(str(c)) for rx in ROLE_RE.values()) for c in (row or []) if c is not None)
    hi = max(range(min(8, len(raw))), key=lambda i: score(raw[i]), default=0)
    header = [str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(raw[hi])]
    rows = [r for r in raw[hi + 1:] if any(c is not None and str(c).strip() for c in r)]
    return header, rows, full


def read_csv(b: bytes):
    df = pl.read_csv(io.BytesIO(b), infer_schema_length=0, truncate_ragged_lines=True,
                     ignore_errors=True, encoding="utf8-lossy")
    return df.columns, [list(r) for r in df.iter_rows()], " ".join(df.columns)


def detect_roles_tab(header, rows):
    roles = {k: None for k in ROLE_RE}
    for role, rx in ROLE_RE.items():
        cands = [i for i, h in enumerate(header) if rx.search(h or "")]
        if not cands:
            continue
        if role == "amount":
            cands.sort(key=lambda i: -(sum(to_eur(r[i]) is not None for r in rows[:200] if i < len(r))
                                       / max(1, len(rows[:200]))))
        roles[role] = cands[0]
    return roles


# ============================================================================ extract
def emit_rows(cf, file_url, b, fmt, max_pages) -> tuple[list[dict], dict]:
    """Parse one file -> gold-schema row dicts + a small per-file stat block."""
    fhash = hashlib.sha256(b).hexdigest()[:16]
    period, year, quarter = period_from_url(file_url)
    rows_out: list[dict] = []
    caveat_detected = False
    conf = "low"

    def base(srn, page, supplier, amount, desc, po, paid):
        return {
            "publisher_id": cf["id"], "publisher_name": cf["name"],
            "publisher_type": cf["ptype"], "sector": cf["sector"],
            "source_landing_url": cf["listing_url"], "source_file_url": file_url,
            "source_file_hash": fhash, "period": period, "year": year, "quarter": quarter,
            "supplier_raw": supplier, "amount_eur": amount, "amount_semantics": cf["amount_semantics"],
            "description": desc, "po_number": po, "paid_flag": paid,
            "source_row_number": srn, "source_page_number": page,
            "parser_name": f"public_body_{fmt}", "parser_version": PARSER_VERSION,
            "source_caveat": cf["caveat"] or None,
        }

    if fmt == "pdf":
        info = read_pdf(b, max_pages)
        caveat_detected = bool(CAVEAT_RE.search(info["page0"]) or CAVEAT_RE.search(info["header_label"]))
        if not info["digital"] or not info["cols"] or "amount" not in info["roles"]:
            return [], {"status": "unparsed", "reason": "scanned/no-header/no-amount",
                        "rows": 0, "confidence": "low", "pages": info.get("pages")}
        sup_i = info["roles"].get("supplier")
        amt_i = info["roles"]["amount"]
        desc_i, po_i, paid_i = (info["roles"].get(k) for k in ("description", "po", "paid"))
        good = 0
        for srn, (page, rec) in enumerate(info["rows"]):
            amt = to_eur(rec[amt_i]) if amt_i < len(rec) else None
            if amt is None:
                continue
            sup = clean_supplier(rec[sup_i]) if sup_i is not None and sup_i < len(rec) else None
            if sup and CATEGORY_WORD.search(sup):
                continue  # drop total/category-masquerade rows
            good += 1
            rows_out.append(base(
                srn, page, sup, amt,
                rec[desc_i] if desc_i is not None and desc_i < len(rec) else None,
                clean_supplier(rec[po_i]) if po_i is not None and po_i < len(rec) else None,
                rec[paid_i] if paid_i is not None and paid_i < len(rec) else None))
        conf = "high" if good > 20 else ("medium" if good > 3 else "low")

    else:  # xlsx / csv
        header, rows, full = (read_xlsx if fmt in ("xlsx", "xls") else read_csv)(b)
        caveat_detected = bool(CAVEAT_RE.search(full) or any(CAVEAT_RE.search(h or "") for h in header))
        roles = detect_roles_tab(header, rows)
        sup_i, amt_i = roles["supplier"], roles["amount"]
        if amt_i is None:
            return [], {"status": "unparsed", "reason": "no-amount-col", "rows": 0, "confidence": "low"}
        desc_i, po_i, paid_i = roles["description"], roles["po"], roles["paid"]
        good = 0
        for srn, r in enumerate(rows):
            amt = to_eur(r[amt_i]) if amt_i < len(r) else None
            if amt is None:
                continue
            sup = clean_supplier(r[sup_i]) if sup_i is not None and sup_i < len(r) else None
            if sup and CATEGORY_WORD.search(sup):
                continue
            good += 1
            rows_out.append(base(
                srn, None, sup, amt,
                r[desc_i] if desc_i is not None and desc_i < len(r) else None,
                clean_supplier(r[po_i]) if po_i is not None and po_i < len(r) else None,
                r[paid_i] if paid_i is not None and paid_i < len(r) else None))
        conf = "high" if good > 20 else ("medium" if good > 3 else "low")

    for r in rows_out:
        r["extraction_status"] = "extracted"
        r["extraction_confidence"] = conf
        r["caveat_text_detected"] = caveat_detected
    return rows_out, {"status": "ok" if rows_out else "empty", "rows": len(rows_out), "confidence": conf}


def classify_and_flag(df: pl.DataFrame) -> pl.DataFrame:
    """supplier_normalised + supplier_class + privacy_status; quarantine DEFERRED."""
    if df.is_empty():
        return df
    df = df.with_columns(
        name_norm_expr("supplier_raw").alias("supplier_normalised"),
        pl.col("supplier_raw").map_elements(lambda s: bool(PUBLIC_BODY.search(s or "")), return_dtype=pl.Boolean).alias("_pub"),
        pl.col("supplier_raw").map_elements(lambda s: bool(COMPANY_SUFFIX.search(s or "")), return_dtype=pl.Boolean).alias("_co"),
        pl.col("supplier_raw").map_elements(lambda s: bool(FOREIGN_FORM.search(s or "")), return_dtype=pl.Boolean).alias("_for"),
    ).with_columns(
        pl.when(pl.col("_pub")).then(pl.lit("public_body"))
        .when(pl.col("_co")).then(pl.lit("company"))
        .when(pl.col("_for")).then(pl.lit("foreign_company"))
        .when(pl.col("supplier_raw").is_null() | (pl.col("supplier_raw").str.strip_chars() == ""))
        .then(pl.lit("unknown"))
        .otherwise(pl.lit("sole_trader_or_individual")).alias("supplier_class"),
    ).with_columns(
        # privacy_status flags likely-personal rows for a LATER quarantine pass; nothing dropped now.
        pl.when(pl.col("supplier_class") == "sole_trader_or_individual").then(pl.lit("review_personal_data"))
        .otherwise(pl.lit("ok")).alias("privacy_status"),
        pl.lit(True).alias("public_display"),  # DEFERRED: quarantine not applied this run
        # po_committed / payment_actual are summable; contract_award_value is caution-only.
        (pl.col("amount_semantics").is_in(["po_committed", "payment_actual"])
         & pl.col("amount_eur").is_not_null() & (pl.col("amount_eur") > 0)).alias("value_safe_to_sum"),
    ).drop(["_pub", "_co", "_for"])
    return df


# ============================================================================ main
def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true", help="harvest-only: print candidate files, no parse")
    ap.add_argument("--only", default="", help="comma-separated publisher ids")
    ap.add_argument("--max-files", type=int, default=6, help="cap files parsed per publisher")
    ap.add_argument("--max-pages", type=int, default=None, help="cap pages per PDF")
    args = ap.parse_args()
    only = {x.strip() for x in args.only.split(",") if x.strip()} or None

    pubs = [p for p in PUBLISHERS if not only or p["id"] in only]
    print(f"{'=' * 80}\nPUBLIC-BODY EXTRACT{' (LIST MODE)' if args.list else ''} — {len(pubs)} publishers\n{'=' * 80}")

    all_rows: list[dict] = []
    per_pub: list[dict] = []
    for cf in pubs:
        files = harvest_files(cf)
        print(f"\n[{cf['id']:<22}] {cf['name']}  (tier {cf['tier']}, {cf['amount_semantics']})")
        print(f"   listing: {cf['listing_url']}")
        print(f"   files harvested: {len(files)}")
        for u in files[:8]:
            print(f"     - {u.rsplit('/', 1)[-1][:70]}")
        if len(files) > 8:
            print(f"     … +{len(files) - 8} more")
        if args.list:
            per_pub.append({"id": cf["id"], "files": len(files)})
            continue

        pub_rows, parsed, ok_files, skipped = [], 0, 0, 0
        for u in files[:args.max_files]:
            ext = next((e for e in DATA_EXT if u.lower().split("?")[0].endswith(e)), "")
            fmt = {".pdf": "pdf", ".xlsx": "xlsx", ".xls": "xls", ".csv": "csv"}.get(ext)
            if not fmt:
                continue
            if fmt == "xls":  # legacy binary .xls needs xlrd (pipeline-only extra) — deferred gap
                skipped += 1
                print(f"     ~ skip (.xls needs xlrd): {u.rsplit('/', 1)[-1][:50]}")
                continue
            b = fetch_bytes(u)
            if not b:
                print(f"     ! download failed: {u.rsplit('/', 1)[-1][:50]}")
                continue
            if fmt == "pdf" and b[:4] != b"%PDF":
                continue
            try:
                rows, stat = emit_rows(cf, u, b, fmt, args.max_pages)
            except Exception as e:  # one malformed file must not abort the run
                skipped += 1
                print(f"     ! parse error ({type(e).__name__}): {u.rsplit('/', 1)[-1][:46]}")
                continue
            parsed += 1
            if rows:
                ok_files += 1
            pub_rows.extend(rows)
            print(f"     -> {u.rsplit('/', 1)[-1][:48]:<48} {stat['status']:<8} "
                  f"rows={stat['rows']} conf={stat['confidence']}")
        all_rows.extend(pub_rows)
        per_pub.append({"id": cf["id"], "name": cf["name"], "tier": cf["tier"],
                        "files_seen": len(files), "files_parsed": parsed, "files_skipped": skipped,
                        "files_with_rows": ok_files, "rows": len(pub_rows),
                        "amount_semantics": cf["amount_semantics"], "privacy_risk": cf["privacy_risk"]})

    if args.list:
        print(f"\n{'=' * 80}\nLIST DONE. Lock URLs into config, then run without --list.")
        return

    if not all_rows:
        print("\nno rows extracted")
        return

    SCHEMA_COLS = [
        "publisher_id", "publisher_name", "publisher_type", "sector",
        "source_landing_url", "source_file_url", "source_file_hash",
        "period", "year", "quarter", "supplier_raw", "supplier_normalised",
        "amount_eur", "amount_semantics", "value_safe_to_sum", "description",
        "po_number", "paid_flag", "source_row_number", "source_page_number",
        "parser_name", "parser_version", "extraction_status", "extraction_confidence",
        "caveat_text_detected", "supplier_class", "privacy_status", "public_display",
        "source_caveat",
    ]
    df = pl.DataFrame(all_rows, infer_schema_length=None)
    df = classify_and_flag(df)
    df = df.select([c for c in SCHEMA_COLS if c in df.columns])

    OUT_FACT.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(OUT_FACT, compression="zstd", compression_level=3, statistics=True)

    print(f"\n{'=' * 80}\nGOLD-CANDIDATE WRITTEN\n{'=' * 80}")
    print(f"rows: {df.height:,}  ->  {OUT_FACT}")
    print(df.group_by("supplier_class").len().sort("len", descending=True))
    safe = df.filter(pl.col("value_safe_to_sum"))
    print(f"\nvalue_safe_to_sum rows: {safe.height:,}  "
          f"sum=€{(safe['amount_eur'].sum() or 0):,.0f} (po_committed+payment_actual only)")

    cov = {
        "publishers_attempted": len(per_pub),
        "publishers_with_rows": sum(p["rows"] > 0 for p in per_pub if "rows" in p),
        "rows_extracted": df.height,
        "rows_public_display": int(df["public_display"].sum()),
        "rows_review_personal_data": int((df["privacy_status"] == "review_personal_data").sum()),
        "supplier_class_counts": {r["supplier_class"]: r["len"]
                                  for r in df.group_by("supplier_class").len().iter_rows(named=True)},
        "amount_semantics_counts": {r["amount_semantics"]: r["len"]
                                    for r in df.group_by("amount_semantics").len().iter_rows(named=True)},
        "value_safe_to_sum_rows": safe.height,
        "value_safe_to_sum_total_eur": float(safe["amount_eur"].sum() or 0),
        "by_publisher": per_pub,
        "privacy_quarantine_applied": False,
        "schema_version": 1,
        "parser_version": PARSER_VERSION,
        "generated_at": datetime.now(UTC).isoformat(timespec="seconds"),
        "caveat": "GOLD-CANDIDATE (sandbox, pre-promotion). One row per source line. "
                  "amount_semantics distinguishes po_committed/payment_actual/contract_award_value; "
                  "only value_safe_to_sum (po_committed+payment_actual) may be totalled, labelled "
                  "'ordered/paid', never mixed with award ceilings. PRIVACY QUARANTINE IS DEFERRED: "
                  "rows flagged privacy_status=review_personal_data are NOT yet excluded "
                  "(public_display=True for all) — a quarantine pass must run before any UI use. "
                  "A line is a purchase order or payment record, not evidence of influence.",
    }
    OUT_COV.write_text(json.dumps(cov, indent=2), encoding="utf-8")
    print(f"wrote coverage {OUT_COV}")


if __name__ == "__main__":
    main()
