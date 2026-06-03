"""PHASE-3 SAMPLE EXTRACTION (PRE-ETL): parse ONE file per easy publisher and report.

Not a parser, not an ETL. For the easiest tabular leads the probe found
(probe_procurement_publishers.py), this downloads ONE sample file, actually parses the
rows, maps the columns, and runs the plan §6/Phase-3 "strange values" battery so we know
the real schema + landmines BEFORE committing to parser classes.

Targets (tabular, no OCR, supplier+amount columns spotted by the probe):
  - OPW          XLSX   SUPPLIER NAME · PAYMENT TYPE · AMOUNT
  - Dept Climate XLSX   Order Number · Supplier · Description · Amount (€)
  - TII          CSV    Period · PAYMENT · Vendor · Description

Writes c:/tmp/procurement_publishers/sample_extraction_report.json with the plan keys:
  publisher, file_url, format, rows_extracted, columns_detected, supplier_column,
  amount_column, period_column, description_column, po_column, paid_column,
  caveat_text_detected, personal_name_risk, strange_values, parser_confidence.

Run:  ./.venv/Scripts/python.exe pipeline_sandbox/sample_extract_procurement.py
"""

from __future__ import annotations

import contextlib
import io
import json
import re
import subprocess
import sys
from collections import Counter
from pathlib import Path

import requests

with contextlib.suppress(Exception):
    sys.stdout.reconfigure(encoding="utf-8")

TMP = Path("c:/tmp/procurement_publishers")
OUT = TMP / "sample_extraction_report.json"
H = {"User-Agent": "Mozilla/5.0 (dail-tracker research probe)"}

TARGETS = [
    {"publisher_id": "ie_opw", "publisher_name": "Office of Public Works", "fmt": "xlsx",
     "url": "https://assets.gov.ie/static/documents/b526ff76/OPW_Payments_of_20000_or_over_in_Q1_2026.xlsx"},
    {"publisher_id": "dept_climate", "publisher_name": "Dept Climate/Energy/Environment (file labelled DPER)",
     "fmt": "xlsx",
     "url": "https://assets.gov.ie/static/documents/ae8b1a0a/DPER_Payments_over_20K_Q1_2026_Report.xlsx"},
    {"publisher_id": "ie_tii", "publisher_name": "Transport Infrastructure Ireland", "fmt": "csv",
     "url": "https://websitecms.tii.ie/media/sw3dzt2l/tii-payments-q1-2025-over-20k.csv"},
]

# column-role detection (header-text based)
COL_RE = {
    "supplier": re.compile(r"supplier|payee|vendor|provider|customer|recipient|name", re.I),
    "amount": re.compile(r"amount|total|value|gross|\beuro\b|€|\bpaid\b|payment\b", re.I),
    "period": re.compile(r"period|quarter|\bqtr\b|\bdate\b|\byear\b|month", re.I),
    "description": re.compile(r"descript|\bdesc\b|detail|categor|service|goods|\bgl\b|nature", re.I),
    "po": re.compile(r"order|\bpo\b|referen|\bref\b|number|invoice|transaction", re.I),
    "paid": re.compile(r"\bpaid\b|payment type|status", re.I),
}
CAVEAT_RE = re.compile(r"\bvat\b|exclud|inclus|indicativ|not (a )?payment|gross|net of|estimate|note:", re.I)
# personal-name risk: a supplier cell that looks like an individual (no company suffix, 2-3 words)
COMPANY_SUFFIX = re.compile(r"\b(ltd|limited|dac|plc|clg|llp|teo|teoranta|t/a|uc|inc|llc|gmbh|company|group|services|holdings|&|university|council|hse|board|college|institute)\b", re.I)
CATEGORY_WORD = re.compile(r"^\s*(total|category total|sum|subtotal|grand total|all suppliers|various)\b", re.I)


def _curl(url: str) -> bytes | None:
    try:
        p = subprocess.run(["curl", "-sS", "-k", "-L", "--max-time", "60", "-A", H["User-Agent"], url],
                           capture_output=True, timeout=90)
        return p.stdout if p.returncode == 0 and p.stdout else None
    except Exception:
        return None


def fetch(url: str) -> bytes | None:
    try:
        r = requests.get(url, headers=H, timeout=60)
        r.raise_for_status()
        return r.content
    except Exception:
        return _curl(url)


def load_xlsx(b: bytes) -> tuple[list[str], list[list], str]:
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(b), read_only=True, data_only=True).active
    raw = [list(r) for r in ws.iter_rows(values_only=True)]
    full_text = " ".join(str(c) for row in raw[:6] for c in row if c is not None)
    # header = the early row whose cells hit the most known column keywords
    def score(row) -> int:
        cells = [str(c) for c in (row or []) if c is not None]
        return sum(any(rx.search(c) for rx in COL_RE.values()) for c in cells)
    hi = max(range(min(8, len(raw))), key=lambda i: score(raw[i]), default=0)
    header = [str(c).strip() if c is not None else f"col{j}" for j, c in enumerate(raw[hi])]
    rows = [r for r in raw[hi + 1:] if any(c is not None and str(c).strip() for c in r)]
    return header, rows, full_text


def load_csv(b: bytes) -> tuple[list[str], list[list], str]:
    import polars as pl
    df = pl.read_csv(io.BytesIO(b), infer_schema_length=0, truncate_ragged_lines=True,
                     ignore_errors=True, encoding="utf8-lossy")
    header = df.columns
    rows = [list(r) for r in df.iter_rows()]
    full_text = " ".join(header)
    return header, rows, full_text


def detect_roles(header: list[str], rows: list[list]) -> dict:
    """Map header columns to roles. For 'amount' prefer the keyword-matching column whose
    values are mostly numeric (avoids grabbing a 'Payment Type' text column)."""
    roles: dict[str, int | None] = {k: None for k in COL_RE}
    for role, rx in COL_RE.items():
        cands = [i for i, h in enumerate(header) if rx.search(h or "")]
        if not cands:
            continue
        if role == "amount":
            cands.sort(key=lambda i: -numeric_frac([r[i] for r in rows[:200] if i < len(r)]))
        roles[role] = cands[0]
    # don't let the same column be both amount and paid unless it really is numeric
    return roles


def parse_amount(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s.replace(",", ""))
    if s in ("", "-", "."):
        return None
    with contextlib.suppress(ValueError):
        f = float(s)
        return -f if neg else f
    return None


def numeric_frac(vals: list) -> float:
    if not vals:
        return 0.0
    ok = sum(parse_amount(v) is not None for v in vals)
    return ok / len(vals)


def analyse(t: dict) -> dict:
    b = fetch(t["url"])
    if not b:
        return {**t, "error": "download failed (requests+curl)"}
    header, rows, full_text = (load_xlsx if t["fmt"] == "xlsx" else load_csv)(b)
    roles = detect_roles(header, rows)
    sup_i, amt_i = roles["supplier"], roles["amount"]

    suppliers = [r[sup_i] for r in rows if sup_i is not None and sup_i < len(r)]
    amts_raw = [r[amt_i] for r in rows if amt_i is not None and amt_i < len(r)]
    amts = [parse_amount(v) for v in amts_raw]
    amts_ok = [a for a in amts if a is not None]

    # strange-value battery (plan §6/Phase-3)
    strange = {
        "negative_amounts": sum(a < 0 for a in amts_ok),
        "zero_amounts": sum(a == 0 for a in amts_ok),
        "very_large_amounts_gt_10m": sum(a > 10_000_000 for a in amts_ok),
        "max_amount": max(amts_ok) if amts_ok else None,
        "missing_supplier": sum(s is None or not str(s).strip() for s in suppliers),
        "missing_or_unparseable_amount": sum(a is None for a in amts) if amt_i is not None else None,
        "currency_symbol_in_text": sum(bool(re.search(r"[€$£]", str(v))) for v in amts_raw),
        "duplicate_full_rows": len(rows) - len({tuple(map(str, r)) for r in rows}),
        "category_total_masquerade": sum(bool(CATEGORY_WORD.search(str(s or ""))) for s in suppliers),
    }
    # most-repeated amount (framework/ceiling smell)
    amt_counts = Counter(round(a, 2) for a in amts_ok)
    top_amt = amt_counts.most_common(1)[0] if amt_counts else (None, 0)
    strange["most_repeated_amount"] = {"value": top_amt[0], "count": top_amt[1]}
    # single-outlier domination: one bad cell (extra zeros / annual figure) can wreck a sum.
    total = sum(amts_ok)
    biggest = max(amts_ok) if amts_ok else 0
    strange["largest_amount_share_of_total"] = round(biggest / total, 3) if total else None
    strange["outlier_warning"] = bool(total and biggest / total > 0.5)

    # personal-name risk: company-suffix-free supplier cells
    indiv = [s for s in suppliers if s and not COMPANY_SUFFIX.search(str(s)) and 1 <= len(str(s).split()) <= 3]
    person_risk = round(len(indiv) / len(suppliers), 3) if suppliers else None

    vat_caveat = bool(CAVEAT_RE.search(full_text)) or any(
        bool(CAVEAT_RE.search(h or "")) for h in header)

    # confidence heuristic
    conf = "low"
    if sup_i is not None and amt_i is not None:
        good = numeric_frac(amts_raw)
        if good > 0.9 and strange["missing_supplier"] / max(len(suppliers), 1) < 0.05:
            conf = "high"
        elif good > 0.6:
            conf = "medium"

    return {
        "publisher": t["publisher_name"], "publisher_id": t["publisher_id"],
        "file_url": t["url"], "format": t["fmt"].upper(),
        "rows_extracted": len(rows),
        "columns_detected": header,
        "supplier_column": header[sup_i] if sup_i is not None else None,
        "amount_column": header[amt_i] if amt_i is not None else None,
        "period_column": header[roles["period"]] if roles["period"] is not None else None,
        "description_column": header[roles["description"]] if roles["description"] is not None else None,
        "po_column": header[roles["po"]] if roles["po"] is not None else None,
        "paid_column": header[roles["paid"]] if roles["paid"] is not None else None,
        "amount_safe_to_sum_eur": round(sum(amts_ok), 2) if amts_ok else None,
        "caveat_text_detected": vat_caveat,
        "personal_name_risk_frac": person_risk,
        "sample_suppliers": [str(s) for s in suppliers[:5]],
        "strange_values": strange,
        "parser_confidence": conf,
    }


def main() -> None:
    print(f"{'=' * 74}\nPHASE-3 SAMPLE EXTRACTION — {len(TARGETS)} easy tabular leads\n{'=' * 74}")
    report = []
    for t in TARGETS:
        print(f"\n[{t['publisher_id']}] {t['publisher_name']}  ({t['fmt'].upper()})")
        r = analyse(t)
        report.append(r)
        if r.get("error"):
            print(f"  ERROR: {r['error']}")
            continue
        print(f"  rows={r['rows_extracted']}  conf={r['parser_confidence']}")
        print(f"  supplier={r['supplier_column']!r}  amount={r['amount_column']!r}  "
              f"period={r['period_column']!r}  paid={r['paid_column']!r}")
        print(f"  columns={r['columns_detected']}")
        print(f"  safe_to_sum=€{r['amount_safe_to_sum_eur']:,}" if r["amount_safe_to_sum_eur"] else "  safe_to_sum=n/a")
        print(f"  caveat_text={r['caveat_text_detected']}  personal_name_risk={r['personal_name_risk_frac']}")
        s = r["strange_values"]
        print(f"  strange: neg={s['negative_amounts']} zero={s['zero_amounts']} "
              f">10m={s['very_large_amounts_gt_10m']} missing_sup={s['missing_supplier']} "
              f"missing_amt={s['missing_or_unparseable_amount']} dup_rows={s['duplicate_full_rows']} "
              f"cat_total={s['category_total_masquerade']} "
              f"top_amt={s['most_repeated_amount']}")
        if s.get("outlier_warning"):
            print(f"  !! OUTLIER: largest single amount = {s['largest_amount_share_of_total']:.0%} "
                  f"of the safe-to-sum total — one row dominates; do NOT headline this sum")
        print(f"  sample suppliers: {r['sample_suppliers']}")

    TMP.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    print(f"\nwrote {OUT}")
    print("PRE-ETL. Next: only after these schemas are agreed -> parser classes (plan Phase 4).")


if __name__ == "__main__":
    main()
